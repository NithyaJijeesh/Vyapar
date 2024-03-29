from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render,redirect
from django.contrib.auth.models import User, auth
from django.utils.text import capfirst
from django.contrib import messages
from . models import *
import json
from django.http.response import JsonResponse
from django.utils.crypto import get_random_string
from datetime import date
from datetime import timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.template.response import TemplateResponse
from django.db.models import F
from openpyxl import load_workbook
from django.http.response import JsonResponse, HttpResponse
from openpyxl import Workbook
from num2words import num2words
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import JsonResponse
from django.db.models import Sum
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from io import BytesIO
import pandas as pd
from django.db.models import F
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from vyaparapp.models import company
from openpyxl import load_workbook, Workbook
from django.db.models import Max
from django.db.models import Count
from collections import defaultdict
from .models import LoanAccounts, LoanHistory, party
from datetime import datetime
import re
from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from itertools import zip_longest
from openpyxl.styles import Font
import math

# Create your views here.
def home(request):
  return render(request, 'home.html')

# @login_required(login_url='login')
def logout(request):
    auth.logout(request)
    return redirect('/')

def view_profile(request):
  com =  company.objects.get(user = request.user) 
  selected_options = request.session.get('selected_options', None)
  
  context = {
              'company' : com,
              'selected_options': json.dumps(selected_options)
          }
  return render(request,'profile.html',context)
  
def edit_profile(request,pk):
  com= company.objects.get(id = pk)
  user1 = User.objects.get(id = com.user_id)
  selected_options = request.session.get('selected_options', None)

  if request.method == "POST":

      user1.first_name = capfirst(request.POST.get('f_name'))
      user1.last_name  = capfirst(request.POST.get('l_name'))
      user1.email = request.POST.get('email')
      com.contact_number = request.POST.get('cnum')
      com.address = capfirst(request.POST.get('ards'))
      com.company_name = request.POST.get('comp_name')
      com.company_email = request.POST.get('comp_email')
      com.city = request.POST.get('city')
      com.state = request.POST.get('state')
      com.country = request.POST.get('country')
      com.pincode = request.POST.get('pinc')
      com.gst_num = request.POST.get('gst')
      com.pan_num = request.POST.get('pan')
      com.business_name = request.POST.get('bname')
      com.company_type = request.POST.get('comp_type')
      if len(request.FILES)!=0 :
          com.profile_pic = request.FILES.get('file')

      com.save()
      user1.save()
      return redirect('view_profile')

  context = {
      'company' : com,
      'user1' : user1,
      'selected_options': json.dumps(selected_options)
  } 
  return render(request,'company/edit_profile.html',context)


def sale_invoices(request):
  return render(request, 'sale_invoices.html')

def estimate_quotation(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    all_estimates = Estimate.objects.filter(company = com)
    estimates = []
    for est in all_estimates:
      history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
      dict = {'estimate':est,'history':history}
      estimates.append(dict)
    context = {
      'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,
    }
    return render(request, 'company/estimate_quotation.html',context)

def payment_in(request):
  return render(request, 'payment_in.html')
    
# def sale_order(request):
#   return render(request, 'sale_order.html')

def delivery_chellan(request):
  return render(request, 'delivery_chellan.html')

def sale_return_cr(request):
  return render(request, 'sale_return_cr.html')


# created by athul


def hide_options(request):
    
    com =  company.objects.get(user = request.user)
    if request.method == 'POST':
        selected_options = list(request.POST.getlist('selected_options'))

    request.session['selected_options'] = selected_options
    
    context = {'selected_options': json.dumps(selected_options),
               'company' : com}
   
    return render(request, 'company/homepage.html', context)

# ------created by athul------

def company_reg(request):
  return render(request,'company/register.html')


    
    
def company_reg2(request):
  terms=payment_terms.objects.all()
  return render(request,'company/register2.html',{'terms':terms})  

def add_company(request):
  
  if request.method == 'POST':
    email=request.POST['email']
    user=User.objects.get(email=email)
    
    c =company.objects.get(user = user)
    c.company_name=request.POST['cname']

    c.address=request.POST['address']
    c.city=request.POST['city']
    c.state=request.POST['state']
    c.country=request.POST['country']
    c.pincode=request.POST['pincode']
    c.pan_number=request.POST['pannumber']
    c.gst_type=request.POST['gsttype']
    c.gst_no=request.POST['gstno']
    c.profile_pic=request.FILES.get('image')

    select=request.POST['select']
    terms=payment_terms.objects.get(id=select)
    c.dateperiod=terms
    c.start_date=date.today()
    days=int(terms.days)

    
    end= date.today() + timedelta(days=days)
    c.End_date=end

    code=get_random_string(length=6)
    if company.objects.filter(Company_code = code).exists():
       code2=get_random_string(length=6)
       c.Company_code=code2
    else:
      c.Company_code=code

   
    c.save()

    staff = staff_details.objects.get(email=email,position='company',company=c)
    staff.first_name = request.POST['cname']
    staff.last_name = ''
    staff.save()
    # messages.success(request, 'Welcome'+ ' ' +  user.first_name +' '+user.last_name + ' ')

    return redirect('Allmodule',user.id)  
  return render(request,'company/register2.html')   

def staff_register(request):
  com=company.objects.all()

  return render(request, 'staff/staffreg.html',{'company':com})

def staff_registraction(request):
  if request.method == 'POST':
    fn=request.POST['fname']
    ln=request.POST['lname']
    email=request.POST['eid']
    un=request.POST['uname']
    pas=request.POST['pass']
    ph=request.POST['ph']
    code=request.POST['code']

    if company.objects.filter(Company_code=code).exists():
      com=company.objects.get(Company_code=code)
    else:
        messages.info(request, 'Sorry, Company code is Invalide')
        return redirect('staff_register')
    img=request.FILES.get('image')

    if staff_details.objects.filter(user_name=un).exists():
      messages.info(request, 'Sorry, Username already exists')
      return redirect('staff_register')
    elif staff_details.objects.filter(email=email).exists():
      messages.info(request, 'Sorry, Email already exists')
      return redirect('staff_register')
    else:
      
      staff=staff_details(first_name=fn,last_name=ln,email=email,user_name=un,password=pas,contact=ph,img=img,company=com)
      staff.save()
      return redirect('log_page')

  else:
    print(" error")
    return redirect('staff_register')
  





def companyaccept(request,id):
  data=staff_details.objects.filter(id=id).update(Action=1)
  return redirect('staff_request')

def companyreject(request,id):
  data=staff_details.objects.get(id=id)
  
  data.delete()
  return redirect('staff_request')

def client_request(request):
  data = company.objects.filter(superadmin_approval = 0,reg_action='self').order_by('-id')
  
  all = company.objects.filter(superadmin_approval = 1)
  return render(request,'admin/client_request.html',{'data': data,'all':all})

def client_request_overview(request,id): 
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  return render(request,'admin/client_request_overview.html',{'company':com,'allmodules':allmodules})

def client_details(request):
  data = company.objects.filter(superadmin_approval = 1,reg_action='self').order_by('-id')
  return render(request,'admin/client_details.html',{"data":data})
def client_details_overview(request,id): 
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  return render(request,'admin/client_details_overview.html',{'company':com,'allmodules':allmodules})



def payment_term(request):
  terms = payment_terms.objects.all()
  
  return render(request,'admin/payment_terms.html',{'terms':terms})
def add_payment_terms(request):
  if request.method == 'POST':
    num=int(request.POST['num'])
    select=request.POST['select']
    if select == 'Years':
      days=int(num)*365
      pt = payment_terms(payment_terms_number = num,payment_terms_value = select,days = days)
      pt.save()
      messages.info(request, 'Payment term is added')
      return redirect('payment_term')

    else:  
      days=int(num*30)
      pt = payment_terms(payment_terms_number = num,payment_terms_value = select,days = days)
      pt.save()
      messages.info(request, 'Payment term is added')
      return redirect('payment_term')


  return redirect('payment_term')




 







def admin_notification(request):
  data= modules_list.objects.filter(update_action=1,status='Pending')

  return render(request,'admin/admin_notification.html',{'data':data})

def module_updation_details(request,mid):
  data= modules_list.objects.get(id=mid)
  allmodules= modules_list.objects.get(company=data.company,status='Pending')
  old_modules = modules_list.objects.get(company=data.company,status='New')

  return render(request,'admin/module_updation_details.html',{'data':data,'allmodules':allmodules,'old_modules':old_modules})

def module_updation_ok(request,mid):
  
  old=modules_list.objects.get(company=mid,status='New')
  old.delete()

  data=modules_list.objects.get(company=mid,status='Pending')  
  data.status='New'
  data.save()
  data1=modules_list.objects.filter(company=mid).update(update_action=0)
  return redirect('admin_notification')






def distributor_reg(request):
  terms=payment_terms.objects.all()
  return render(request,'distributor/distributor_reg.html',{'terms':terms})
def distributor_reg_action(request):
  if request.method == 'POST':
    first_name = request.POST['fname']
    last_name = request.POST['lname']
    user_name = request.POST['uname']
    email_id = request.POST['eid']
    mobile = request.POST['ph']
    passw = request.POST['pass']
    c_passw = request.POST['cpass']
    pic = request.FILES.get('image')

    select=request.POST['select']
    terms=payment_terms.objects.get(id=select)
    # c.dateperiod=terms
    start_date=date.today()
    days=int(terms.days)

    
    end= date.today() + timedelta(days=days)
    End_date=end

    code=get_random_string(length=6)
    if Distributors_details.objects.filter(distributor_id = code).exists():
       code=get_random_string(length=6)
  
    if passw == c_passw:
      if User.objects.filter(username = user_name).exists():
        messages.info(request, 'Sorry, Username already exists')
        return redirect('distributor_reg')
      

      elif not User.objects.filter(email = email_id).exists():
    
        user_data = User.objects.create_user(first_name = first_name,
                        last_name = last_name,
                        username = user_name,
                        email = email_id,
                        password = passw)
        user_data.save()
        
        data = User.objects.get(id = user_data.id)
        distributor_data = Distributors_details(contact=mobile,distributor_id=code,img=pic,
                                                payment_term=terms,start_date=start_date,End_date=End_date,
                                                user = data)
        distributor_data.save()
        
        return redirect('log_page')
      else:
        messages.info(request, 'Sorry, Email already exists')
        return redirect('distributor_reg')
  return render(request,'distributor/distributor_reg.html')
 
def distributor_home(request):
  distributor =  Distributors_details.objects.get(user = request.user)

  return render(request,'distributor/distributor_home.html',{'distributor':distributor})



      
def clients(request):
  return render(request,'admin/clients.html')

def distributors(request):
  return render(request,'admin/distributors.html')  

def distributor_request(request):
  data = Distributors_details.objects.filter(Log_Action = 0).order_by('-id')
  return render(request,'admin/distributor_request.html',{'data':data})

def admin_distributor_accept(request,id):
  data=Distributors_details.objects.filter(id=id).update(Log_Action=1)
  return redirect('distributor_request')
def admin_distributor_reject(request,id):
  data=Distributors_details.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('distributor_request')

def distributor_request_overview(request,id):
  data=Distributors_details.objects.get(id=id)
  return render(request,'admin/distributor_request_overview.html',{'data':data})

def distributor_details(request):
  data = Distributors_details.objects.filter(Log_Action = 1).order_by('-id')
  return render(request,'admin/distributor_details.html',{'data':data})

def distributor_details_overview(request,id):
  data = Distributors_details.objects.get(id=id)
  return render(request,'admin/distributor_details_overview.html',{'data':data})

def dcompany_request(request):
  
  distributor =  Distributors_details.objects.get(user = request.user)
  data = company.objects.filter(Distributors = distributor,Distributor_approval = 0,reg_action='distributor').order_by('-id')
  return render(request,'distributor/dcompany_request.html',{'data':data,'distributor':distributor})

def dcompany_request_overview(request,id):
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/dcompany_request_overview.html',{'company':com,'allmodules':allmodules,'distributor':distributor})

def distributor_accept_company(request,id):
  data=company.objects.filter(id=id).update(Distributor_approval=1)
  
  return redirect('dcompany_request')
def distributor_reject_company(request,id):
  data=company.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('dcompany_request')

def dcompany_details(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  data = company.objects.filter(Distributors = distributor,Distributor_approval = 1,reg_action='distributor').order_by('-id')
  
  return render(request,'distributor/dcompany_details.html',{'data':data,'distributor':distributor})

def dcompany_details_overview(request,id):
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/dcompany_details_overview.html',{'company':com,'allmodules':allmodules,'distributor':distributor})

def distributor_profile(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/distributor_profile.html',{'distributor':distributor})

# ========================================   ASHIKH V U (START) ======================================================

# @login_required(login_url='login')
def item_create(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules=modules_list.objects.get(company=cmp)
  # item_units = UnitModel.objects.filter(user=request.user.id)
  item_units = UnitModel.objects.filter(company = cmp) #updated - shemeem
  return render(request,'company/item_create.html',{'item_units':item_units,'company':cmp, 'staff':staff, 'allmodules' : allmodules})

# @login_required(login_url='login')
def items_list(request,pk):
  try:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    all_items = ItemModel.objects.filter(company=cmp) #updated - shemeem

    if pk == 0:
      first_item = all_items.filter().first()
    elif pk != 0:
      first_item = all_items.get(id=pk)
    print(first_item)

    item_history= Item_History.objects.filter(
      Item= first_item,
      company=cmp
    ).values('Item', 'action' , 'staff__first_name' , 'staff__last_name').last()

    allmodules= modules_list.objects.get(company=staff.company,status='New')
    transactions = None
    # transactions = TransactionModel.objects.filter(user=request.user.id,item=first_item.id).order_by('-trans_created_date')
    if first_item:
      transactions = TransactionModel.objects.filter(company = cmp,item=first_item.id).order_by('-trans_created_date')

    context = {
                'first_item':first_item,
                'transactions':transactions,
                'company':cmp, 
                'staff':staff, 
                'allmodules' : allmodules, 
                'all_items' : all_items , 
                'Item_History' : item_history
              }

    if all_items == None or all_items == '' or first_item == None or first_item == '' or transactions == None or transactions == '':
      return render(request,'company/items_create_first_item.html', context)
    return render(request,'company/items_list.html',context)
  except:
    return render(request,'company/items_create_first_item.html', context)

# @login_required(login_url='login')
def item_create_new(request):
  if request.method=='POST':
    #updated-shemeem
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' or None :
      item_opening_stock = 0
      item_current_stock = 0
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '' or None:
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == ''  or None:
      item_min_stock_maintain = 0
    item_data = ItemModel(user=staff.company.user,
                          company=cmp,
                          item_name=item_name,
                          item_hsn=item_hsn,
                          item_unit=item_unit,
                          item_taxable=item_taxable,
                          item_gst=item_gst,
                          item_igst=item_igst,
                          item_sale_price=item_sale_price,
                          item_purchase_price=item_purchase_price,
                          item_opening_stock=item_opening_stock,
                          item_current_stock=item_current_stock,
                          item_at_price=item_at_price,
                          item_date=item_date,
                          item_min_stock_maintain=item_min_stock_maintain)
  
    if request.POST.get('save_and_next'):
      if ItemModel.objects.filter(item_name=item_name, item_hsn=item_hsn).exists():
        messages.error(request, 'Item with the same item name and HSN  number already exists.')
      elif ItemModel.objects.filter(item_name=item_name).exists():
        messages.error(request, 'An item can have one HSN Number.')
      elif ItemModel.objects.filter(item_hsn=item_hsn).exists():
        messages.error(request, 'Item with the same HSN  number already exists.')
      else:
        item_data.save()
        Item_History.objects.create(Item = item_data,company=cmp,staff=staff,action='Created').save()
      return redirect('item_create')
    
    elif request.POST.get('save'):
      if ItemModel.objects.filter(item_name=item_name, item_hsn=item_hsn).exists():
        messages.error(request, 'Item with the same item name and HSN  number already exists.')
      elif ItemModel.objects.filter(item_name=item_name).exists():
        messages.error(request, 'An item can have one HSN Number.')
      elif ItemModel.objects.filter(item_hsn=item_hsn).exists():
        messages.error(request, 'Item with the same HSN  number already exists.')
      else:
        item_data.save()
        Item_History.objects.create(Item = item_data,company=cmp,staff=staff,action='Created').save()

      return redirect('items_list',pk=0)
  return redirect('item_create')


# @login_required(login_url='login')
def item_delete(request,pk):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  item_to_delete = ItemModel.objects.get(id=pk)
  print(staff.company.id)

  # # item_to_delete.delete()
  # return redirect('items_list',pk=0)

    # List of models to check
  models_to_check1 = [CreditNoteItem, PurchaseBillItem, purchasedebit1, PurchaseOrderItem]
  models_to_check2 = [SalesInvoiceItem,Estimate_items, DeliveryChallanItems,TransactionModel]
  # Check conditions for each model
  conditions_met1 = any(model.objects.filter(company=staff.company.id, product=item_to_delete).exists() for model in models_to_check1)
  conditions_met2 = any(model.objects.filter(company=staff.company.id, item= item_to_delete).exists() for model in models_to_check2)

  if conditions_met1 or conditions_met2 or sales_item.objects.filter(cmp = staff.company.id, product = item_to_delete):
      
      messages.error(request, 'Cannot delete Item with transactions.')
      return redirect('items_list',pk=0)  # 1 could be a code indicating failure
  else:
      item_to_delete.delete()
      return redirect('items_list',pk=0) 
    


# @login_required(login_url='login')
def item_view_or_edit(request,pk):
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  
  item = ItemModel.objects.get(id=pk)
  # item_units = UnitModel.objects.filter(user=request.user.id)
  item_units = UnitModel.objects.filter(company = cmp)
  return render(request,'company/item_view_or_edit.html',{'item':item,'item_units':item_units,'staff':staff})

  
# @login_required(login_url='login')
def item_unit_create(request):
  if request.method=='POST':
    #updated-shemeem
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    # user = User.objects.get(id=request.user.id)
    # company_user_data = company.objects.get(user=request.user.id)

    item_unit_name = request.POST.get('item_unit_name')
    unit_data = UnitModel(user=cmp.user,company=cmp,unit_name=item_unit_name)
    unit_data.save()
  return JsonResponse({'message':'success'})

  
# @login_required(login_url='login')
def item_update(request,pk):
  if request.method=='POST':
    #updated-shemeem
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    item_data = ItemModel.objects.get(id=pk)
    # user = User.objects.get(id=request.user.id)
    user = cmp.user
    # company_user_data = company.objects.get(user=request.user.id)
    company_user_data = cmp

    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    if item_taxable == 'Non Taxable':
      item_gst = 'GST0[0%]'
      item_igst = 'IGST0[0%]'
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' :
      item_opening_stock = 0
      item_current_stock = 0
    else:
      if int(item_opening_stock) > item_data.item_opening_stock:
        item_data.item_current_stock += (int(item_opening_stock) - item_data.item_opening_stock)
      else:
        item_data.item_current_stock -= (int(item_opening_stock) - item_data.item_opening_stock)
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '':
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == '':
      item_min_stock_maintain = 0

    item_data.user = user
    item_data.company_user_data = company_user_data
    item_data.item_name = item_name
    item_data.item_hsn = item_hsn
    item_data.item_unit = item_unit
    item_data.item_taxable = item_taxable
    item_data.item_gst = item_gst
    item_data.item_igst = item_igst
    item_data.item_sale_price = item_sale_price
    item_data.item_purchase_price = item_purchase_price
    item_data.item_opening_stock = item_opening_stock
    item_data.item_current_stock = int(item_current_stock)
    item_data.item_at_price = item_at_price
    item_data.item_date = item_date
    item_data.item_min_stock_maintain = item_min_stock_maintain

    item_data.save()
    Item_History.objects.create(Item = item_data,company=cmp,staff=staff,action='Updated').save()
    print('\nupdated')
  # return redirect('item_view_or_edit',pk)
  return redirect('items_list',pk=item_data.id)


def item_histories(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  getitem = ItemModel.objects.get(id=id,company=cmp)
  item_histories= Item_History.objects.filter(Item=getitem,company=cmp)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  context = {'staff':staff,'allmodules':allmodules,'item_history': item_histories,'getitem': getitem,}
  return render(request,'company/item_history.html',context)

  
# @login_required(login_url='login')
def item_search_filter(request):
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  search_string = request.POST.get('searching_item')
  # items_filtered = ItemModel.objects.filter(user=request.user.id)
  items_filtered = ItemModel.objects.filter(user=cmp.user)
  items_filtered = items_filtered.filter(Q(item_name__icontains=search_string) | Q(item_current_stock__icontains = search_string))
  item_unit_name = request.POST.get('item_unit_name')
  return TemplateResponse(request,'company/item_search_filter.html',{'all_items':items_filtered})


# @login_required(login_url='login')
def item_get_detail(request,pk):
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  item = ItemModel.objects.get(id=pk)
  transactions = TransactionModel.objects.filter(user=cmp.user,item=item.id).order_by('-trans_created_date')
  return TemplateResponse(request,'company/item_get_detail.html',{"item":item,'transactions':transactions,})

  
# @login_required(login_url='login')
def item_get_details_for_modal_target(request,pk):
  item = ItemModel.objects.get(id=pk)
  return TemplateResponse(request,'company/item_get_details_for_modal_target.html',{"item":item,})

# @login_required(login_url='login')
def ajust_quantity(request,pk):
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    # user = User.objects.get(id=request.user.id)
    user = cmp.user
    # company_user_data = company.objects.get(user=request.user.id)
    company_user_data = cmp

    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    trans_adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    item.item_current_stock = trans_adjusted_qty
    item.save()
    transaction_data = TransactionModel(user=user,
                                        company=company_user_data,
                                        item=item,
                                        trans_type=trans_type,
                                        trans_user_name=trans_user_name,
                                        trans_date=trans_date,
                                        trans_qty=trans_qty,
                                        trans_current_qty=trans_current_qty,
                                        trans_adjusted_qty=trans_adjusted_qty,)
    
    Item_History.objects.create(Item = item,company=cmp,staff=staff,action='Updated').save()
    transaction_data.save()
  return redirect('items_list',pk=item.id)


# @login_required(login_url='login')
def transaction_delete(request,pk):
  transaction = TransactionModel.objects.get(id=pk),k
  item = ItemModel.objects.get(id=transaction.item_id)
  if transaction.trans_type=='add stock':
    item.item_current_stock = item.item_current_stock - transaction.trans_qty
  else:
    item.item_current_stock = item.item_current_stock + transaction.trans_qty
  item.save()
  transaction.delete()
  return redirect('items_list',pk=item.id)

  
# @login_required(login_url='login')
def item_transaction_view_or_edit(request,pk,tran):
  item = ItemModel.objects.get(id=pk)
  transaction = TransactionModel.objects.get(id=tran)
  print('enterd')
  return TemplateResponse(request,'company/item_transaction_view_or_edit.html',{"item":item,"transaction":transaction,})


# @login_required(login_url='login')
def update_adjusted_transaction(request,pk,tran):
  item = ItemModel.objects.get(id=pk)
  transaction = TransactionModel.objects.get(id=tran)
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    # user = User.objects.get(id=request.user.id)
    user = cmp.user
    # company_user_data = company.objects.get(user=request.user.id)
    company_user_data = cmp

    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    if transaction.trans_type == 'reduce stock':
      if trans_type == 'reduce stock':
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  - transaction.trans_qty)
      else:
        item.item_current_stock = item.item_current_stock + transaction.trans_qty + int(trans_qty)
    else:
      if trans_type == 'reduce stock':
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  + transaction.trans_qty)
      else:
        item.item_current_stock = item.item_current_stock - transaction.trans_qty + int(trans_qty)
    # item.item_opening_stock = adjusted_qty
    item.save()
    transaction.trans_type =trans_type
    transaction.trans_date=trans_date
    transaction.trans_qty =trans_qty
    transaction.trans_current_qty=trans_current_qty
    transaction.save()
  return redirect('items_list',pk=item.id)
  
# @login_required(login_url='login')
def item_delete_open_stk(request,pk):
  item = ItemModel.objects.get(id=pk)
  if item.item_opening_stock > item.item_current_stock:
    item.item_current_stock =item.item_opening_stock - item.item_current_stock
  else:
    item.item_current_stock =item.item_current_stock - item.item_opening_stock
  # item.item_current_stock =  item.item_opening_stock - item.item_current_stock
  item.item_opening_stock = 0
  # print(f'{item.item_current_stock }={item.item_opening_stock}-{item.item_current_stock}')
  item.save()
  return redirect('items_list',pk=item.id)
  
# ========================================   ASHIKH V U (END) ======================================================


# ---------------Nithya-------import item-----------------

def downloadItemSampleImportFile(request):
    
    item_table_data = [['Item Name','HSN','Unit','Tax Type','GST Rate','IGST Rate','Sales Price','Purchase Price','Opening Stock','At Price','Date','Minimum Stock to Maintain'],['Check  Unit, Tax Type, GST Rate and IGST Rate sheet for details . And remove this row for add item details']]
    details_table_data = [
          
          ['Unit','NOS','BOX','PACKET'],
          ['Tax Type','Taxable', 'Non Taxable'],
          ['GST Rate','GST0[0%]','GST3[3%]','GST5[5%]','GST12[12%]','GST13[18%]','GST28[28%]'],
          ['IGST Rate','IGST0[0%]','IGST3[3%]','IGST5[5%]','IGST12[12%]','IGST13[18%]','IGST28[28%]'],
    ]  
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'Sample Item Details'
    sheet2 = wb.create_sheet(title=' Unit, Tax Type, GST and IGST')

    for row in item_table_data:
        sheet1.append(row)

    transposed_data = list(zip_longest(*details_table_data))
    for row in transposed_data:
      sheet2.append(row)

    bold_headers = ['Item Name','HSN','Unit','Tax Type','GST Rate','IGST Rate','Sales Price','Purchase Price','Opening Stock','At Price','Date','Minimum Stock to Maintain']
    for col_num, header in enumerate(item_table_data[0], start=1):
      if header in bold_headers:
        sheet1.cell(row=1, column=col_num).font = Font(bold=True)

    for col_num, header in enumerate(transposed_data[0], start=1):
      if header in bold_headers:
        sheet2.cell(row=1, column=col_num).font = Font(bold=True)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=item_sample_file.xlsx'
    wb.save(response)
    return response


def import_items(request):
  if request.method == 'POST':
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    comp = company.objects.get(id=staff.company.id)

    file = request.FILES['itemfile']
    df = pd.read_excel(file)

    errors = []
    count_rows = 0

    try:    
      for index, row in df.iterrows():
        count_rows +=1

        item_name = '' if isinstance(row.get('Item Name'), float) and math.isnan(row.get('Item Name')) else str(row.get('Item Name', '')).capitalize()
        hsn = '' if isinstance(row.get('HSN'), float) and math.isnan(row.get('HSN')) else str(row.get('HSN', ''))
        current_date = date.today() if 'nan' else datetime.strptime(str(row.get('Date')), '%Y-%m-%d').date()

        item_obj = ItemModel(
          item_name= item_name,
          item_hsn = hsn,
          item_unit = '' if isinstance(row.get('Unit'), float) and math.isnan(row.get('Unit')) else str(row.get('Unit', '')),
          item_taxable = '' if isinstance(row.get('Tax Type'), float) and math.isnan(row.get('Tax Type')) else row.get('Tax Type', ''),
          item_gst = '' if isinstance(row.get('GST Rate'), float) and math.isnan(row.get('GST Rate')) else row.get('GST Rate', ''),
          item_igst = '' if isinstance(row.get('IGST Rate'), float) and math.isnan(row.get('IGST Rate')) else row.get('IGST Rate', ''),
          item_sale_price = '' if isinstance(row.get('Sales Price'), float) and math.isnan(row.get('Sales Price')) else row.get('Sales Price', ''),
          item_purchase_price = '' if isinstance(row.get('Purchase Price'), float) and math.isnan(row.get('Purchase Price')) else row.get('Purchase Price', ''),
          item_opening_stock = 0 if isinstance(row.get('Opening Stock'), float) and math.isnan(row.get('Opening Stock')) else row.get('Opening Stock', ''),
          item_current_stock = 0 if isinstance(row.get('Opening Stock'), float) and math.isnan(row.get('Opening Stock')) else row.get('Opening Stock', ''),
          item_at_price = 0 if isinstance(row.get('At Price'), float) and math.isnan(row.get('At Price')) else row.get('At Price', ''),
          item_date = current_date,
          item_min_stock_maintain = 0 if isinstance(row.get('Minimum Stock to Maintain'), float) and math.isnan(row.get('Minimum Stock to Maintain')) else row.get('Minimum Stock to Maintain', ''),
          user= staff.company.user,
          company=comp,
        )

        if not item_name or not hsn or hsn == " ":
          messages.error(request, f'Row "{count_rows}" :Please Enter Item Name and HSN Number.')
        else:
          if ItemModel.objects.filter(item_name=item_name, item_hsn=hsn).exists():
            messages.error(request, 'Item with the same item name and HSN  number already exists.')
          elif ItemModel.objects.filter(item_name=item_name).exists():
            messages.error(request, 'An item can have one HSN Number.')
          elif ItemModel.objects.filter(item_hsn=hsn).exists():
            messages.error(request, 'Item with the same HSN  number already exists.')
          else:
            item_obj.save()
            Item_History.objects.create(Item = item_obj,company=staff.company,staff=staff,action='Created').save()

      item_final = ItemModel.objects.filter(company=comp).last()
      return redirect('items_list', item_final.id)
    
    except Exception as e:
      error_message = f"Error in row {index + 1}: {e}"
      errors.append(error_message)
      return redirect('items_list', 0)
    
  return redirect('items_list', 0)


#_________________Parties(new)_______________Antony Tom_________


def add_parties(request):
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  tod = date.today()
  allmodules= modules_list.objects.get(company=cmp,status='New')


  return render(request, 'company/add_parties.html',{'staff':staff, 'tod' : tod , 'allmodules' : allmodules})


def view_parties(request,pk):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  Party=party.objects.filter(company=staff.company.id)
  if pk == 0:
      getparty = party.objects.filter(company=staff.company.id).first()
  else:
      getparty = party.objects.get(company=staff.company.id, id=pk)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  party_histories= party_history.objects.filter(
    party=getparty,
    company=staff.company
  ).values('party', 'action' , 'staff__first_name' , 'staff__last_name').last()

  context = { 
              'staff':staff,
              'allmodules':allmodules,
              'Party':Party, 
              'getparty' : getparty, 
              'party_history' : party_histories,
             }
  return render(request, 'company/view_parties.html',context)

def party_histories(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  getparty = party.objects.get(id=id,company=cmp)
  party_histories= party_history.objects.filter(party=getparty,company=cmp)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  context = {'staff':staff,'allmodules':allmodules,'party_history': party_histories,'getparty': getparty, 'allmodules' : allmodules}
  return render(request,'company/party_history.html',context)

def save_parties(request):
    if request.method == 'POST':
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        allmodules= modules_list.objects.get(company=staff.company,status='New')
            #updated by Nithya

        party_name = request.POST['partyname'].capitalize()
        gst_no = request.POST.get('gstno')
        contact = request.POST['contact']
        gst_type = request.POST.get('gsttype')
        state = request.POST.get('splystate')
        address = request.POST.get('baddress')
        email = request.POST.get('partyemail')
        openingbalance = request.POST.get('openbalance', '')
        balance = request.POST.get('openbalance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('crd_lmt', '')
        current_date = request.POST.get('partydate')
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST.get('additional1')
        additionalfield2 = request.POST.get('additional2')
        additionalfield3 = request.POST.get('additional3')

        context  = {'staff' : staff, 'tod' : date.today(),'allmodules' : allmodules}
       
        # if ( not party_name ):
        #   return render(request, 'company/add_parties.html',context)

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date, current_balance = balance, End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=staff.company.user,company=staff.company)
        
        if not party_name or not contact:
            messages.error(request, 'Please Enter Party Name and Contact.')
        else:
          if 'save_and_new' in request.POST:
              if party.objects.filter(party_name=party_name, contact=contact).exists() or party.objects.filter(contact=contact).exists():
                  messages.error(request, 'Party with the same party name and contact number already exists.')
              else:
                  part.save()
                  party_history.objects.create(party = part,company=staff.company,staff=staff,action='Created').save()
              
              return render(request, 'company/add_parties.html', context)
          else:
              if party.objects.filter(party_name=party_name, contact=contact).exists() or party.objects.filter(contact=contact).exists():
                  messages.error(request, 'Party with the same party name and contact number already exists.')
              else:
                  part.save()
                  party_history.objects.create(party = part,company=staff.company,staff=staff,action='Created').save()

              return redirect('view_parties', part.id)

    return render(request, 'company/add_parties.html',context)  

def party_histories(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  getparty = party.objects.get(id=id,company=cmp)
  party_histories= party_history.objects.filter(party=getparty,company=cmp)
  print(party_histories)
  context = {'staff':staff,'allmodules':allmodules,'party_histories': party_histories,'getparty': getparty}
  return render(request,'company/party_history.html',context)


def edit_party(request,id):
  #updated-shemeem
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  # Company = company.objects.get(user=request.user)
  Company = cmp
  # user_id = request.user.id
  user_id = cmp.user.id
  getparty=party.objects.get(id=id)
  # Party=party.objects.filter(user=request.user)
  Party=party.objects.filter(user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  tod = date.today()
  context = {
              'Company':Company,
              'user_id':user_id,
              'Party':Party,
              'getparty':getparty,
              'staff':staff, 
              'tod' : datetime.now(),
              'allmodules' : allmodules
            }
  return render(request, 'company/edit_party.html',context)


def edit_saveparty(request, id):
    #updated-shemeem
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    Party=party.objects.filter(user=cmp.user)
    user_id = cmp.user.id
    getparty = party.objects.get(id=id)
    Company = cmp
    #updated by Nithya

    if request.method == 'POST':
        getparty.party_name = request.POST.get('partyname')
        getparty.gst_no = request.POST.get('gstno')
        getparty.contact = request.POST.get('contact')
        getparty.gst_type = request.POST.get('gsttype')
        getparty.state = request.POST.get('splystate')
        getparty.address = request.POST.get('baddress')
        getparty.email = request.POST.get('partyemail')
        getparty.current_balance = request.POST.get('openbalance') if getparty.current_balance == getparty.openingbalance else getparty.current_balance
        getparty.openingbalance = request.POST.get('openbalance')
        getparty.payment = request.POST.get('paymentType')
        getparty.creditlimit = request.POST.get('crd_lmt')
        getparty.current_date = request.POST.get('partydate')
        getparty.additionalfield1 = request.POST.get('additional1')
        getparty.additionalfield2 = request.POST.get('additional2')
        getparty.additionalfield3 = request.POST.get('additional3')

        getparty.save()

        party_history.objects.create(party = getparty,company=staff.company,staff=staff,action='Updated').save()
        # return redirect('view_party', id=getparty.id)
        return redirect('view_parties', pk=getparty.id)
    return render(request, 'edit_party.html', {'getparty': getparty, 'Party': Party, 'Company': Company,'user_id':user_id})


def deleteparty(request, id):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    Party = party.objects.get(id=id)

    # List of models to check
    models_to_check1 = [PurchaseBill, PurchaseOrder, SalesInvoice, purchasedebit, PaymentOut,PaymentIn, CreditNote]
    models_to_check2 = [Estimate, DeliveryChallan]
    # Check conditions for each model
    conditions_met1 = any(model.objects.filter(company=staff.company.id, party=Party).exists() for model in models_to_check1)
    conditions_met2 = any(model.objects.filter(company=staff.company.id, party_name=Party.party_name).exists() for model in models_to_check2)

    if conditions_met1 or conditions_met2 or Expense.objects.filter(staff_id = staff, party_id = Party) or salesorder.objects.filter(comp = staff.company.id, party = Party):
        
        messages.error(request, 'Cannot delete Party with transactions.')
        return redirect('view_parties', 0)  # 1 could be a code indicating failure
    else:
        Party.delete()
        return redirect('view_parties', 0) 
    
# ---------------Nithya-------import party-----------------

def downloadPartySampleImportFile(request):
    
    party_table_data = [['Party Name','Contact','Email','GST No.','GST Type','Supply State','Billing Address','Opening Balance','Payment','Current Date','Credit Limit','Additional Field 1','Additional Field 2','Additional Field 3'],['Check GSTType, Supply States, Payment Sheet for details. And remove this row for add party details']]
    details_table_data = [
          ['GST Type','Registered Party', 'Unregistered or Consumer', 'Registered Business or Composition'],
          ['Supply State','Andaman and Nicobar Islands','Andhra Pradhesh','Arunachal Pradesh','Assam','Bihar','Chandigarh','Chhattisgarh',
           'Dadra and Nagar Haveli','Daman and Diu','Delhi','Goa','Gujarat','Haryana','Himachal Pradesh','Jammu and Kashmir','Jharkhand',
           'Karnataka','Kerala','Ladakh','Lakshadweep','Madhya Pradesh','Maharashtra','Manipur','Meghalaya','Mizoram','Nagaland','Odisha',
           'Pondicherry','Punjab','Rajasthan','Sikkim','Tamil Nadu','Telangana','Tripura','Uttar Pradesh','Uttarakhand','West Bengal','Other Territory'],
          ['Payment','To Pay', 'To Recieve'],
          ['GST Number Format','32AAQFR1222B1ZS'],
          ['Email Format','abc@example.com'],
          ['Conatct Format', '7890555444']
    ]  
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'Sample Party Details'
    sheet2 = wb.create_sheet(title='GSTType, Supply States, Payment')

    # Populate the sheets with data
    for row in party_table_data:
        sheet1.append(row)

    transposed_data = list(zip_longest(*details_table_data))
    for row in transposed_data:
      sheet2.append(row)


    bold_headers = ['Party Name','Contact','Email','GST No.','GST Type','Supply State','Billing Address','Opening Balance','Payment','Current Date','Credit Limit','Additional Field 1','Additional Field 2','Additional Field 3','GST Number Format','Email Format','Conatct Format']
    for col_num, header in enumerate(party_table_data[0], start=1):
        if header in bold_headers:
            sheet1.cell(row=1, column=col_num).font = Font(bold=True)

    for col_num, header in enumerate(transposed_data[0], start=1):
        if header in bold_headers:
            sheet2.cell(row=1, column=col_num).font = Font(bold=True)
    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=party_sample_file.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response

def import_parties(request):
    
  if request.method == 'POST':

      staff_id = request.session['staff_id']
      staff =  staff_details.objects.get(id=staff_id)
      cmp = company.objects.get(id=staff.company.id)
      file = request.FILES['partyfile']

      df = pd.read_excel(file)

      errors = []
      count_rows = 0

      try:    
        for index, row in df.iterrows():
          count_rows +=1

          party_name = row.get('Party Name').capitalize()
          contact = str(row.get('Contact'))

          phone_pattern = re.compile(r'^\d{10}$')
          contact = contact if phone_pattern.match(contact) else None

          current_date = date.today() if 'nan' else datetime.strptime(str(row.get('Current Date')), '%Y-%m-%d').date()

          party_obj = party(
            party_name=party_name,
            contact=contact,
            gst_no='' if isinstance(row.get('GST No.'), float) and math.isnan(row.get('GST No.')) else str(row.get('GST No.', '')),
            gst_type='' if isinstance(row.get('GST Type'), float) and math.isnan(row.get('GST Type')) else row.get('GST Type', ''),
            email='' if isinstance(row.get('Email'), float) and math.isnan(row.get('Email')) else row.get('Email', ''),
            state='' if isinstance(row.get('Supply State'), float) and math.isnan(row.get('Supply State')) else row.get('Supply State', ''),
            address='' if isinstance(row.get('Billing Address'), float) and math.isnan(row.get('Billing Address')) else row.get('Billing Address', ''),
            openingbalance=0 if isinstance(row.get('Opening Balance'), float) and math.isnan(row.get('Opening Balance')) else row.get('Opening Balance', ''),
            payment='' if isinstance(row.get('Payment'), float) and math.isnan(row.get('Payment')) else row.get('Payment', ''),
            creditlimit='' if isinstance(row.get('Credit Limit'), float) and math.isnan(row.get('Credit Limit')) else row.get('Credit Limit', ''),
            current_date=current_date,
            End_date=current_date,
            additionalfield1='' if isinstance(row.get('Additional Field 1'), float) and math.isnan(row.get('Additional Field 1')) else row.get('Additional Field 1', ''),
            additionalfield2='' if isinstance(row.get('Additional Field 2'), float) and math.isnan(row.get('Additional Field 2')) else row.get('Additional Field 2', ''),
            additionalfield3='' if isinstance(row.get('Additional Field 3'), float) and math.isnan(row.get('Additional Field 3')) else row.get('Additional Field 3', ''),
            current_balance=0 if isinstance(row.get('Opening Balance'), float) and math.isnan(row.get('Opening Balance')) else row.get('Opening Balance', ''),
            user=staff.company.user,
            company=staff.company
          )

          if not party_name or not contact or contact == " ":
            messages.error(request, f'Row "{count_rows}" :Please Enter Party Name and Contact Number.')
          else:
            if party.objects.filter(contact=contact).exists(): 
              if party.objects.filter(party_name=party_name, contact=contact).exists():
                messages.error(request, f'Row "{count_rows}" :Party with the same party name "{party_name}"  and contact number "{contact}" already exists.')
              else:
                messages.error(request, f'Row "{count_rows}" :Party with the same contact number "{contact}" already exists.')
            else:
              party_obj.save() 
              party_history.objects.create(party = party_obj,company=staff.company,staff=staff,action='Created').save()
        party_final = party.objects.filter(company=cmp).last()
        return redirect('view_parties', party_final.id)
      
      except Exception as e:
          error_message = f"Error in row {index + 1}: {e}"
          errors.append(error_message)
          return redirect('view_parties', 0)
  return redirect('view_parties', 0)


#End

@login_required(login_url='login')
def adminhome(request):
 
  
  
  return render(request, 'admin/adminhome.html')


#******************************************   ASHIKH V U (start) ****************************************************

from django.http import HttpResponse
import re

# account number validation
def validate_bank_account_number(acc_num):
  regex='^[0-9]{9,18}'
  if re.match(regex,acc_num):
    return True
  else:
    return False

# ifsc code validaion
def validate_ifsc(ifsc_code):
    regex = re.compile(r'^[A-Za-z]{4}\d{7}$')
    if regex.match(ifsc_code):
        return True
    else:
        return False

#@login_required(login_url='login')
def account_num_check(request):
  if request.method=='POST':
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    account_num_valid = validate_bank_account_number(account_num)
    if account_num_valid:
      if BankModel.objects.filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
        return HttpResponse('<small><span class="tr fs-2">Account Number already excist</span></small>')
      else:
        return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">Account Number is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def account_num_check_for_edit(request,pk):
  if request.method=='POST':
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    account_num_valid = validate_bank_account_number(account_num)
    if account_num_valid:
      if BankModel.objects.exclude(id=pk).filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
        return HttpResponse('<small><span class="tr fs-2">Account Number already excist</span></small>')
      else:
        return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">Account Number is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def bank_ifsc_check (request):
  if request.method=='POST':
    bank_ifsc = request.POST.get('ifsc')
    print(bank_ifsc)
    ifsc_valid = validate_ifsc(bank_ifsc)
    if ifsc_valid:
      return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">IFSC Code is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def bank_create(request):
  print('asdasd')
  try:
    staff_id = request.session['staff_id']
    print(staff_id)
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    # permission
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    # permission
    return render(request,'company/bank_create.html',{"allmodules":allmodules})
  except:
    user = User.objects.get(id=request.user.id)
    get_company_id_using_user_id = company.objects.get(user=user)
    # permission
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    # permission
    return render(request,'company/bank_create.html',{"allmodules":allmodules})


#@login_required(login_url='login')
def banks_list(request,pk):
  
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
 

  try:
    all_banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
    if pk == 0:
      first_bank = all_banks.first()
      print(all_banks)
      return redirect('banks_list',pk=first_bank.id)
    else:
      bank = all_banks.get(id=pk)
      transactions_all = BankTransactionModel.objects.filter(company=get_company_id_using_user_id.id)
      transactions = transactions_all.filter(Q(from_here=pk) | Q(to_here=pk))
      tr_history = BankTransactionHistory.objects.filter().order_by('date')
    if all_banks.exists():
      open_bal_last_edited = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED')).last()
      
      if bank.open_balance:
        total = bank.open_balance
      else:
        total = 0
      for i in transactions:
        if i.type == "Cash Withdraw":
          total = total - i.amount
        elif  i.type == 'Adjustment Reduce':
          total = total - i.amount
        elif i.from_here == bank:
          total = total - i.amount
        else:
          total = total + i.amount
        i.current_amount = total
      
      return render(request,'company/banks_list.html',{"allmodules":allmodules,
                                                      "all_banks":all_banks,
                                                      "bank":bank,
                                                      "transactions":transactions,
                                                      "tr_history":tr_history,
                                                      "open_bal_last_edited":open_bal_last_edited,
                                                      "staff":staff}) 
    else:
      return render(request,'company/bank_create_first_bank.html',{"allmodules":allmodules,'staff':staff}) 
  except:
    return render(request,'company/bank_create_first_bank.html',{"allmodules":allmodules,'staff':staff}) 
    

#@login_required(login_url='login')
def get_bank_to_bank(request):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_bank_to_bank.html',{'banks':banks})

#@login_required(login_url='login')
def get_bank_to_cash(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_bank_to_cash.html',{'banks':banks})

#@login_required(login_url='login')
def get_cash_to_bank(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_cash_to_bank.html',{'banks':banks})

#@login_required(login_url='login')
def get_adjust_bank_balance(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_adjust_bank_balance.html',{'banks':banks})

#@login_required(login_url='login')
def bank_create_new(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    print(get_company_id_using_user_id)
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    if BankModel.objects.exclude(company=get_company_id_using_user_id.id).filter(bank_name=bank_name,user=user.id,account_num=account_num).exists():
      parmission_var = 0
    else:
      parmission_var = 1
    if validate_bank_account_number(account_num):
      parmission_var1 = 1
    else:
      parmission_var1 = 0
    ifsc = request.POST.get('ifsc')
    if validate_ifsc(ifsc):
      parmission_var2 = 1
    else:
      parmission_var2 = 0
    branch_name = request.POST['branch_name']
    upi_id = request.POST.get('upi_id')
    as_of_date = request.POST['as_of_date']
    card_type = request.POST.get('card_type')
    open_balance = request.POST['open_balance']
    
    if open_balance == '' or open_balance == None:
      open_balance = 0
    if card_type == "CREDIT":
      open_balance = int(open_balance)*-1
      
    if parmission_var == 1:
      if parmission_var1 == 1:
        if parmission_var2 == 1:
          bank_data = BankModel(user=user,
                                company=get_company_id_using_user_id,
                                bank_name=bank_name,
                                account_num=account_num,
                                ifsc=ifsc,
                                branch_name=branch_name,
                                upi_id=upi_id,
                                as_of_date=as_of_date,
                                card_type=card_type,
                                open_balance=open_balance,
                                current_balance=open_balance,
                                created_by=user.first_name)
          bank_data.save()
          tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                              bank=bank_data,
                                              action="BANK CREATION : "+bank_data.bank_name.upper(),
                                              done_by_name=staff.first_name,
                                              done_by=staff)
          tr_history.save()
          tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                              bank=bank_data,
                                              action="BANK OPEN BALANCE CREATED",
                                              done_by_name=staff.first_name,
                                              done_by=staff)
          tr_history.save()
          if request.POST.get('save_and_next'):
            messages.success(request,'Bank created successfully')
            return redirect('bank_create')
          else:
            return redirect('banks_list',pk=bank_data.id)
        else:
          messages.error(request,'IFSC CODE is not valid')
          return redirect('bank_create')
      else:
        messages.error(request,'Account number is not valid')
        return redirect('bank_create')
    else:
      messages.error(request,'Account number already exist')
      return redirect('bank_create')
  return redirect('banks_list',pk=bank_data.id)

#@login_required(login_url='login')
def bank_delete(request,pk):
  bank = BankModel.objects.get(id=pk)
  bank.delete()
  return redirect('banks_list',pk=0)

#@login_required(login_url='login')
def bank_view_or_edit(request,pk):
  bank = BankModel.objects.get(id=pk)
  return render(request,'company/bank_view_or_edit.html',{"bank":bank})

#@login_required(login_url='login')
def bank_update(request,pk):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    
    bank_data = BankModel.objects.get(id=pk)

    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    if BankModel.objects.exclude(id=pk).filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
      parmission_var = 0
    else:
      parmission_var = 1
    if validate_bank_account_number(account_num):
      parmission_var1 = 1
    else:
      parmission_var1 = 0
    ifsc = request.POST.get('ifsc')
    if validate_ifsc(ifsc):
      parmission_var2 = 1
    else:
      parmission_var2 = 0
    branch_name = request.POST['branch_name']
    upi_id = request.POST.get('upi_id')
    as_of_date = request.POST['as_of_date']
    card_type = request.POST.get('card_type')
    open_balance = request.POST['open_balance']
    
    if open_balance == '' or open_balance == None:
      open_balance = 0
    if card_type == "CREDIT":
      open_balance = int(open_balance)*-1
    if parmission_var == 1:
      if parmission_var1 == 1:
        if parmission_var2 == 1:
          bank_data.user = user
          bank_data.company = get_company_id_using_user_id
          bank_data.bank_name = bank_name
          bank_data.account_num = account_num
          bank_data.ifsc = ifsc
          bank_data.branch_name = branch_name
          bank_data.upi_id = upi_id
          bank_data.as_of_date = as_of_date
          bank_data.card_type = card_type

          if int(bank_data.open_balance) < int(open_balance):
            bank_data.current_balance = int(bank_data.current_balance) + (int(open_balance) - int(bank_data.open_balance))
          elif int(bank_data.open_balance) == int(open_balance):
            bank_data.current_balance = int(open_balance)
          elif int(bank_data.open_balance) > int(open_balance):
            bank_data.current_balance = int(bank_data.current_balance)- (int(bank_data.open_balance) - int(open_balance))

          if bank_data.open_balance != open_balance:
            validity = True
          else:
            validity = False
          old_val = bank_data.open_balance

          bank_data.open_balance = open_balance
          bank_data.user = user
          bank_data.save()

          if validity == True:
            tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank=bank_data,
                                          action="BANK OPEN BALANCE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
            tr_history.save()
        else:
          messages.error(request,'IFSC CODE is not valid')
          return redirect('bank_create')
      else:
        messages.error(request,'Account number is not valid')
        return redirect('bank_create')
    else:
      messages.error(request,'Account number already exist')
      return redirect('bank_create')
  return redirect('banks_list',pk=bank_data.id)


#@login_required(login_url='login')
def bank_to_bank_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    select2 = request.POST.get('to_here')
    to_here = BankModel.objects.get(id=select2)
    type = "BANK TO BANK"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date')
    
    bank1 = BankModel.objects.get(id=from_here.id)
    bank1.current_balance -= int(amount)
    bank1.save()
    bank2 = BankModel.objects.get(id=to_here.id)
    bank2.current_balance += int(amount)
    bank2.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        to_here=to_here,
                                        type=type,
                                        date=date,
                                        name=name,
                                        amount=amount,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK TO BANK TRANSACTION CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def bank_to_cash_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    type = "Cash Withdraw"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank1 = BankModel.objects.get(id=from_here.id)
    bank1.current_balance -= int(amount)
    bank1.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def cash_to_bank_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select2 = request.POST.get('to_here')
    to_here = BankModel.objects.get(id=select2)
    type = "Cash Deposit"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank2 = BankModel.objects.get(id=to_here.id)
    bank2.current_balance += int(amount)
    bank2.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        to_here=to_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank2,
                                        bank_trans=transaction_data,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=to_here.id)


#@login_required(login_url='login')
def get_adjust_bank_balance_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    inc_red = request.POST.get('inc_red')
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank1 = BankModel.objects.get(id=from_here.id)
    if inc_red == 'Increase Balance':
      bank1.current_balance += int(amount) 
      type = "Adjustment Increase"
    else:
      bank1.current_balance -= int(amount)
      type = "Adjustment Reduce"
    bank1.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK BALANCE "+type.upper()+" CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def delete_bank_open_balance(request,pk):
  bank = BankModel.objects.get(id=pk)
  bank.current_balance = bank.current_balance - bank.open_balance
  bank.open_balance = 0
  bank.save()
  if 'banks_list' in request.META.get('HTTP_REFERER',None):
    return redirect('banks_list',pk=pk)
  else:
    return redirect('bank_transaction_statement',bank_id=pk)

#@login_required(login_url='login')
def delete_bank_transaction(request,pk,bank_id):
  print(pk,bank_id)
  try:
    pk = request.POST.get('pk')
    bank_id = request.POST.get('bank_id')
    print(pk,bank_id)
  except:
    pk=pk
    bank_id=bank_id

  try:
    trans = BankTransactionModel.objects.get(id=pk)
    if trans.type == 'BANK TO BANK':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      bank2.current_balance -= trans.amount
      bank2.save()
      trans.delete()
      print('enterd')
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Cash Withdraw' or trans.type == 'CASH WITHDRAW':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Cash Deposit' or trans.type == 'CASH DEPOSIT':
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      bank2.current_balance -= trans.amount
      bank2.save()
      trans.delete()
      print('entered')
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Adjustment Increase' or trans.type == 'ADJUSTMENT INCREASE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance -= trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Adjustment Reduce' or trans.type == 'ADJUSTMENT REDUCE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
  except:
    return redirect('banks_list',pk=bank_id)
  return redirect('banks_list',pk=0)

#@login_required(login_url='login')
def view_or_edit_bank_transaction(request,pk,bank_id):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  transaction = BankTransactionModel.objects.get(id=pk)
  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  bank = BankModel.objects.get(id=bank_id)
  if transaction.type == "BANK TO BANK" or transaction.type == 'Bank to bank':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_to_bank_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Cash Withdraw' or transaction.type == 'Cash withdraw' or transaction.type == 'CASH WITHDRAW':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_to_cash_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Cash Deposit' or transaction.type == 'Cash deposit' or transaction.type == 'CASH DEPOSIT':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/cash_to_bank_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Adjustment Increase' or transaction.type == 'Adjustment increase' or transaction.type == 'Adjustment Reduce' or transaction.type == 'Adjustment reduce' or transaction.type == 'ADJUSTMENT INCREASE' or transaction.type == 'ADJUSTMENT REDUCE':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_adjust_bank_balance_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})

#@login_required(login_url='login')
def update_bank_transaction(request,pk,bank_id):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    amount = request.POST.get('amount')
    date = request.POST.get('date')
    trans = BankTransactionModel.objects.get(id=pk)
    trans.date = date
    if trans.type == 'BANK TO BANK':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      if trans.amount > int(amount):
        bank2.current_balance -= (trans.amount-int(amount))
      else:
        bank2.current_balance += (int(amount)-trans.amount)
      bank2.save()
      old_amount = trans.amount
      if old_amount != amount:
        validity =True
      else:
        validity =False
      trans.amount = amount
      trans.save()
      if validity == True:
        tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK TO BANK TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
        tr_history.save()
        trans.last_action='UPDATED'
        trans.by = staff.first_name
        trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=bank_id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Cash Withdraw' or trans.type == 'CASH WITHDRAW':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK TO CASH TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Cash Deposit'  or trans.type == 'CASH DEPOSIT':
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      if trans.amount > int(amount):
        bank2.current_balance -= (trans.amount-int(amount))
      else:
        bank2.current_balance += (int(amount)-trans.amount)
      bank2.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="CASH TO BANK TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.to_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Adjustment Increase' or trans.type == 'ADJUSTMENT INCREASE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance -= (trans.amount-int(amount))
      else:
        bank1.current_balance += (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK BALANCE ADJUSTMENT INCREASE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Adjustment Reduce' or trans.type == 'ADJUSTMENT REDUCE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK BALANCE ADJUSTMENT REDUCE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    return redirect('banks_list',pk=0)
  return redirect('banks_list',pk=0)

from openpyxl import load_workbook
from django.utils import timezone

#@login_required(login_url='login')
def import_from_excel(request,pk):
    current_datetime = timezone.now()
    date =  current_datetime.date()

    try:
      if request.method == "POST" and 'excel_file' in request.FILES:
        
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        get_company_id_using_user_id = company.objects.get(id=staff.company.id)
        user = get_company_id_using_user_id.user
        allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            TYPE, FROM, TO,NAME,DATE,AMOUNT,ACTION,BY = row

            if TYPE != None:
              TYPE = TYPE.upper()
            
            if AMOUNT != None:
              AMOUNT = AMOUNT.replace(' ','')
              AMOUNT = AMOUNT.replace('₹','')
              AMOUNT = AMOUNT.replace('-','')
              AMOUNT = AMOUNT.replace('+','')
              AMOUNT = int(float(AMOUNT))

            print(f'{TYPE}  {FROM}  {TO}    {NAME}  {DATE}  {AMOUNT}')
            
            if TYPE == "BANK TO BANK" or TYPE == 'Bank to bank':
              from_here = BankModel.objects.get(id=int(FROM))
              to_here = BankModel.objects.get(id=int(TO))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  to_here=to_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              to_here.current_balance += AMOUNT
              to_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  bank_trans=transaction,
                                                  action="BANK TO BANK TRANSACTION CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
              
            elif TYPE == 'Open. Balance' or TYPE == 'OPEN. BALANCE':
              from_here = BankModel.objects.get(id=int(FROM))
              if from_here.open_balance > AMOUNT:
                from_here.current_balance += from_here.open_balance - AMOUNT
              else:
                from_here.current_balance -= from_here.open_balance - AMOUNT
              from_here.open_balance = AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  action="BANK OPEN BALANCE CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Withdraw' or TYPE == 'Cash withdraw' or TYPE == 'CASH WITHDRAW':
              from_here = BankModel.objects.get(id=int(FROM))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Deposit' or TYPE == 'Cash deposit' or TYPE == 'CASH DEPOSIT':
              to_here = BankModel.objects.get(id=int(TO))
              to_here.current_balance += AMOUNT
              to_here.save()

              transaction = BankTransactionModel(user = user,
                                                  company=get_company_id_using_user_id,
                                                  to_here=to_here,
                                                  type=TYPE,
                                                  amount=AMOUNT,
                                                  date=DATE,
                                                  last_action='CREATED',
                                                  by = staff.first_name,
                                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=to_here,
                                        bank_trans=transaction,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Increase' or TYPE == 'Adjustment increase' or TYPE == 'ADJUSTMENT INCREASE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance += AMOUNT
              from_here.save()
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Reduce' or TYPE == 'Adjustment reduce' or TYPE == 'ADJUSTMENT REDUCE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance -= AMOUNT
              from_here.save()
              transaction = BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
    except:
      messages.warning(request,"Table field is missing / you are importing the wrong File.")
    return redirect('banks_list',pk=pk)

#@login_required(login_url='login')
def import_statement_from_excel(request,pk):
    current_datetime = timezone.now()
    date =  current_datetime.date()

    try:
      if request.method == "POST" and 'excel_file' in request.FILES:
        
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        get_company_id_using_user_id = company.objects.get(id=staff.company.id)
        user = get_company_id_using_user_id.user
        allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            TYPE, FROM, TO,NAME,DATE,AMOUNT,BALANCE = row
            # TYPE, FROM, TO,NAME,DATE,AMOUNT,BALANCE,ACTION,BY = row

            if TYPE != None:
              TYPE = TYPE.upper()
            
            if AMOUNT != None:
              AMOUNT = AMOUNT.replace(' ','')
              AMOUNT = AMOUNT.replace('₹','')
              AMOUNT = AMOUNT.replace('-','')
              AMOUNT = AMOUNT.replace('+','')
              AMOUNT = int(float(AMOUNT))

            print(f'{TYPE}  {FROM}  {TO}    {NAME}  {DATE}  {AMOUNT}')
            
            if TYPE == "BANK TO BANK" or TYPE == 'Bank to bank':
              from_here = BankModel.objects.get(id=int(FROM))
              to_here = BankModel.objects.get(id=int(TO))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  to_here=to_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              to_here.current_balance += AMOUNT
              to_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  bank_trans=transaction,
                                                  action="BANK TO BANK TRANSACTION CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
              
            elif TYPE == 'Open. Balance' or TYPE == 'OPEN. BALANCE':
              from_here = BankModel.objects.get(id=int(FROM))
              if from_here.open_balance > AMOUNT:
                from_here.current_balance += from_here.open_balance - AMOUNT
              else:
                from_here.current_balance -= from_here.open_balance - AMOUNT
              from_here.open_balance = AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  action="BANK OPEN BALANCE CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Withdraw' or TYPE == 'Cash withdraw' or TYPE == 'CASH WITHDRAW':
              from_here = BankModel.objects.get(id=int(FROM))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Deposit' or TYPE == 'Cash deposit' or TYPE == 'CASH DEPOSIT':
              to_here = BankModel.objects.get(id=int(TO))
              to_here.current_balance += AMOUNT
              to_here.save()

              transaction = BankTransactionModel(user = user,
                                                  company=get_company_id_using_user_id,
                                                  to_here=to_here,
                                                  type=TYPE,
                                                  amount=AMOUNT,
                                                  date=DATE,
                                                  last_action='CREATED',
                                                  by = staff.first_name,
                                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=to_here,
                                        bank_trans=transaction,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Increase' or TYPE == 'Adjustment increase' or TYPE == 'ADJUSTMENT INCREASE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance += AMOUNT
              from_here.save()
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Reduce' or TYPE == 'Adjustment reduce' or TYPE == 'ADJUSTMENT REDUCE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance -= AMOUNT
              from_here.save()
              transaction = BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
    except:
      messages.warning(request,"Table field is missing / you are importing the wrong File.")
    return redirect('bank_transaction_statement',bank_id=pk) 

#@login_required(login_url='login')
def transaction_history(request,pk,bank_id):
    
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    all_banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)

    # tr_history1 = BankTransactionHistory.objects.filter(action__contains='BANK CREATION',bank=bank_id)
    tr_history2 = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED'),bank=bank_id)    
    tr_history = BankTransactionHistory.objects.filter(bank_trans=pk)
    if pk != 0:
      # tr_historys = tr_history | tr_history1
      tr_historys = tr_history
    else:
      # tr_historys = tr_history1 | tr_history1 | tr_history2
      tr_historys =  tr_history | tr_history2
    # print(tr_history)
    
    return render(request,'company/bank_transaction_history.html',{"allmodules":allmodules,
                                                                   "all_banks":all_banks,
                                                                    "tr_historys":tr_historys,
                                                                    "bank_id":bank_id,
                                                                    "staff":staff})


#@login_required(login_url='login')
def bank_transaction_statement(request,bank_id):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
  bank = BankModel.objects.get(id=bank_id)

  transactions_all = BankTransactionModel.objects.filter(company=get_company_id_using_user_id.id)
  transactions = transactions_all.filter(Q(from_here=bank_id) | Q(to_here=bank_id))
  tr_history = BankTransactionHistory.objects.filter().order_by('date')

  open_bal_last_edited = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED')).last()
  
  if bank.open_balance:
    total = bank.open_balance
  else:
    total = 0
  for i in transactions:
    if i.type == "Cash Withdraw":
      total = total - i.amount
    elif  i.type == 'Adjustment Reduce':
      total = total - i.amount
    elif i.from_here == bank:
      total = total - i.amount
    else:
      total = total + i.amount
    i.current_amount = total

  return render(request,'company/bank_transaction_statement.html',{"allmodules":allmodules,
                                                  "bank":bank,
                                                  "transactions":transactions,
                                                  "tr_history":tr_history,
                                                  "open_bal_last_edited":open_bal_last_edited,
                                                  "staff":staff})

#******************************************   ASHIKH V U (end) ****************************************************


def view_purchasebill(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  pbill = PurchaseBill.objects.filter(company=cmp)
  
  if not pbill:
    context = {'staff':staff, 'allmodules':allmodules}
    return render(request,'company/purchasebillempty.html',context)
  context = {'staff':staff,'allmodules':allmodules,'pbill':pbill}
  return render(request,'company/purchasebilllist.html',context)


def add_purchasebill(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  last_bill = PurchaseBill.objects.filter(company=cmp).last()

  if last_bill:
    bill_no = last_bill.tot_bill_no + 1 
  else:
    bill_no = 1

  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)

  context = {'staff':staff, 'allmodules':allmodules, 'cust':cust, 'cmp':cmp,'bill_no':bill_no, 'tod':tod, 'item':item, 'item_units':item_units,'bank':bank}
  return render(request,'company/purchasebilladd.html',context)


def create_purchasebill(request):
  if request.method == 'POST': 
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    part = party.objects.get(id=request.POST.get('customername'))
    pbill = PurchaseBill(party=part, 
                          billno=request.POST.get('bill_no'),
                          billdate=request.POST.get('billdate'),
                          supplyplace =request.POST.get('placosupply'),
                          pay_method=request.POST.get("method"),
                          cheque_no=request.POST.get("cheque_id"),
                          upi_no=request.POST.get("upi_id"),
                          advance = request.POST.get("advance"),
                          balance = request.POST.get("balance"),
                          subtotal=float(request.POST.get('subtotal')),
                          igst = request.POST.get('igst'),
                          cgst = request.POST.get('cgst'),
                          sgst = request.POST.get('sgst'),
                          adjust = request.POST.get("adj"),
                          taxamount = request.POST.get("taxamount"),
                          grandtotal=request.POST.get('grandtotal'),
                          company=cmp,staff=staff)
    pbill.save()
        
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    billno = PurchaseBill.objects.get(billno =pbill.billno,company=cmp)

    if len(product)==len(qty)==len(discount)==len(total):
        mapped=zip(product,qty,discount,total)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          PurchaseBillItem.objects.create(product = itm,qty=ele[1],discount=ele[2],total=ele[3],purchasebill=billno,company=cmp)

    PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=F('tot_bill_no') + 1)
    
    pbill.tot_bill_no = pbill.billno
    pbill.save()

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Created')

    if 'Next' in request.POST:
      return redirect('add_purchasebill')
    
    if "Save" in request.POST:
      return redirect('view_purchasebill')
    
  else:
    return render(request,'company/purchasebilladd.html')


def edit_purchasebill(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  billprd = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)

  if pbill.pay_method != 'Cash' and pbill.pay_method != 'Cheque' and pbill.pay_method != 'UPI':
    bankno = BankModel.objects.get(id= pbill.pay_method,company=cmp,user=cmp.user)
  else:
    bankno = 0

  bdate = pbill.billdate.strftime("%Y-%m-%d")
  context = {'staff':staff, 'allmodules':allmodules, 'pbill':pbill, 'billprd':billprd,'tod':tod,
             'cust':cust, 'item':item, 'item_units':item_units, 'bdate':bdate,'bank':bank, 'bankno':bankno}
  return render(request,'company/purchasebilledit.html',context)


def update_purchasebill(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = party.objects.get(id=request.POST.get('customername'))
    pbill = PurchaseBill.objects.get(id=id,company=cmp)
    pbill.party = part
    pbill.billdate = request.POST.get('billdate')
    pbill.supplyplace  = request.POST.get('placosupply')
    pbill.subtotal =float(request.POST.get('subtotal'))
    pbill.grandtotal = request.POST.get('grandtotal')
    pbill.igst = request.POST.get('igst')
    pbill.cgst = request.POST.get('cgst')
    pbill.sgst = request.POST.get('sgst')
    pbill.taxamount = request.POST.get("taxamount")
    pbill.adjust = request.POST.get("adj")
    pbill.pay_method = request.POST.get("method")
    pbill.cheque_no = request.POST.get("cheque_id")
    pbill.upi_no = request.POST.get("upi_id")
    pbill.advance = request.POST.get("advance")
    pbill.balance = request.POST.get("balance")

    pbill.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
    if len(total)==len(discount)==len(qty):
      mapped=zip(product,qty,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        PurchaseBillItem.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],purchasebill=pbill,company=cmp)

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasebill')

  return redirect('view_purchasebill')


def details_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  pitm = PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'allmodules':allmodules,'pbill':pbill,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'company/purchasebilldetails.html',context)


def history_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pbill = PurchaseBill.objects.get(id=id,company=cmp)
  hst= PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp)

  context = {'staff':staff,'allmodules':allmodules,'hst':hst,'pbill':pbill}
  return render(request,'company/purchasebillhistory.html',context)


def delete_purchasebill(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(id=id)
  PurchaseBillItem.objects.filter(purchasebill=pbill,company=cmp).delete()
  pbill.delete()
  return redirect('view_purchasebill')


def import_purchase_bill(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(PurchaseBill.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      PurchaseBill.objects.create(party=part,billno=totval,
                                  billdate=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      pbill = PurchaseBill.objects.last()
      if billsheet[4] == 'Cheque':
        pbill.pay_method = 'Cheque'
        pbill.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        pbill.pay_method = 'UPI'
        pbill.upi_no = billsheet[5]
      else:
        if billsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=billsheet[4],company=cmp)
          pbill.pay_method = bank
        else:
          pbill.pay_method = 'Cash'
      pbill.save()

      PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=int(prdsheet[2]),company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          PurchaseBillItem.objects.create(purchasebill=pbill,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

          if billsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if billsheet[3]=='State':
            gst = round((taxamount/2),2)
            pbill.sgst=gst
            pbill.cgst=gst
            pbill.igst=0

          else:
            gst=round(taxamount,2)
            pbill.igst=gst
            pbill.cgst=0
            pbill.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pbill.subtotal=round(subtotal,2)
      pbill.taxamount=round(taxamount,2)
      pbill.adjust=round(billsheet[6],2)
      pbill.grandtotal=gtotal
      pbill.advance=round(billsheet[7],2)
      pbill.balance=balance
      pbill.save()

      PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,staff=pbill.staff,company=pbill.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})


def billhistory(request):
  pid = request.POST['id']
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pbill = PurchaseBill.objects.get(billno=pid,company=cmp)
  hst = PurchaseBillTransactionHistory.objects.filter(purchasebill=pbill,company=cmp).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid})


def bankdata(request):
  bid = request.POST['id']
  bank = BankModel.objects.get(id=bid) 
  bank_no = bank.account_num
  bank_name = bank.bank_name
  return JsonResponse({'bank_no':bank_no,'bank_name':bank_name})


def savecustomer(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  party_name = request.POST['name']
  email = request.POST['email']
  contact = request.POST['mobile']
  state = request.POST['splystate']
  address = request.POST['baddress']
  gst_type = request.POST['gsttype']
  gst_no = request.POST['gstin']
  current_date = request.POST['partydate']
  openingbalance = request.POST.get('openbalance')
  payment = request.POST.get('paytype')
  creditlimit = request.POST.get('credit_limit')
  End_date = request.POST.get('enddate', None)
  additionalfield1 = request.POST['add1']
  additionalfield2 = request.POST['add2']
  additionalfield3 = request.POST['add3']

  part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,
                payment=payment,creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,
                additionalfield3=additionalfield3,company=cmp,user=cmp.user)
  part.save() 
  return JsonResponse({'success': True})


def cust_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = party.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  party_list = []
  for p in part:
    id_list.append(p.id)
    party_list.append(p.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })


def saveitem(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  if not hsn:
    hsn = None

  itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})


def item_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
  return JsonResponse({'id_list':id_list, 'product_list':product_list})


def custdata(request):
  cid = request.POST['id']
  part = party.objects.get(id=cid)
  phno = part.contact
  address = part.address
  pay = part.payment
  bal = part.openingbalance
  return JsonResponse({'phno':phno, 'address':address, 'pay':pay, 'bal':bal})


def itemdetails(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})

def add_purchaseorder(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  last_ord = PurchaseOrder.objects.filter(company=cmp).last()

  if last_ord:
    ord_no = last_ord.tot_ord_no + 1 
  else:
    ord_no = 1

  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)

  context = {'staff':staff, 'allmodules':allmodules, 'cust':cust, 'cmp':cmp,'ord_no':ord_no, 'tod':tod, 'item':item, 'item_units':item_units,'bank':bank}
  return render(request,'company/purchaseorderadd.html',context)
  
  
def view_purchaseorder(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  pord = PurchaseOrder.objects.filter(company=cmp)

  if not pord:
    context = {'staff':staff, 'allmodules':allmodules}
    return render(request,'company/purchaseorderempty.html',context)
  
  context = {'staff':staff, 'allmodules':allmodules,'pord':pord}
  return render(request,'company/purchaseorderlist.html',context)

# ===========  estimate & delivery challan ===========shemeem==================   
   
def delivery_challan(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    all_challan = DeliveryChallan.objects.filter(company = com)
    challan = []
    for dc in all_challan:
      history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
      dict = {'challan':dc,'history':history}
      challan.append(dict)
    context = {
      'staff':staff, 'company':com,'allmodules':allmodules, 'challan':challan,
    }
    return render(request, 'company/delivery_challan.html',context)



def addEstParty(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    allmodules= modules_list.objects.get(company=staff.company,status='New')

    if request.method == 'POST':
      Company = company.objects.get(id = staff.company.id)
      user_id = request.user.id
      
      party_name = request.POST['partyname']
      gst_no = request.POST['gstno']
      contact = request.POST['contact']
      gst_type = request.POST['gst']
      state = request.POST['state']
      address = request.POST['address']
      email = request.POST['email']
      openingbalance = request.POST.get('balance', '')
      payment = request.POST.get('paymentType', '')
      creditlimit = request.POST.get('creditlimit', '')
      current_date = request.POST['currentdate']
      End_date = request.POST.get('enddate', None)
      additionalfield1 = request.POST['additionalfield1']
      additionalfield2 = request.POST['additionalfield2']
      additionalfield3 = request.POST['additionalfield3']
      comp=Company


      # print(party_name)
      # print(gst_no)
      # print(contact)
      # print(gst_type)
      # print(state)
      # print(address)
      # print(email)
      # print(openingbalance)
      # print(payment)
      # print(creditlimit)
      # print(current_date)
      # print(End_date)
      # print(additionalfield1)
      # print(additionalfield2)
      # print(additionalfield3)


      context  = {'staff' : staff, 'tod' : date.today(),'allmodules' : allmodules}
       

      part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                      creditlimit=creditlimit,current_date=current_date, current_balance = openingbalance, End_date=End_date,additionalfield1=additionalfield1,
                      additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=staff.company.user,company=staff.company)
        
      if not party_name or not contact:
          messages.error(request, 'Please Enter Party Name and Contact.')
      else:
        if 'save_and_new' in request.POST:
            if party.objects.filter(party_name=party_name, contact=contact).exists() or party.objects.filter(contact=contact).exists():
                messages.error(request, 'Party with the same party name and contact number already exists.')
            else:
                part.save()
                party_history.objects.create(party = part,company=staff.company,staff=staff,action='Created').save()
            
            return render(request, 'company/add_parties.html', context)
        else:
            if party.objects.filter(party_name=party_name, contact=contact).exists() or party.objects.filter(contact=contact).exists():
                messages.error(request, 'Party with the same party name and contact number already exists.')
            else:
                part.save()
                party_history.objects.create(party = part,company=staff.company,staff=staff,action='Created').save()

            return JsonResponse({'status':True})

    return render(request, 'company/create_estimate.html',context)  


      # if (not party_name):
      #   return render(request, 'add_parties.html')

      # part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
      #                 creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,company=comp)
      # part.save()

      # return JsonResponse({'status':True})

  
def addEstItem(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    if request.method=='POST':
      company_user_data = com
      item_name = request.POST.get('item_name')
      item_hsn = request.POST.get('item_hsn')
      item_unit = request.POST.get('item_unit')
      item_taxable = request.POST.get('item_taxable')
      item_gst = request.POST.get('item_gst')
      item_igst = request.POST.get('item_igst')
      item_sale_price = 0 if request.POST.get('item_sale_price') is '' or None else request.POST.get('item_sale_price')
      item_purchase_price = 0 if request.POST.get('item_purchase_price') is '' or None else request.POST.get('item_purchase_price')
      item_opening_stock = request.POST.get('item_opening_stock')
      item_current_stock = item_opening_stock
      if item_opening_stock == '' or None :
        item_opening_stock = 0
        item_current_stock = 0
      item_at_price = request.POST.get('item_at_price')
      if item_at_price == '' or None:
        item_at_price =0
      item_date = date.today() if request.POST.get('item_date') == "" or None  else request.POST.get('item_date')

      item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
      if item_min_stock_maintain == ''  or None:
        item_min_stock_maintain = 0

      item_data = ItemModel(company=company_user_data,
        item_name=item_name,
        item_hsn=item_hsn,
        item_unit=item_unit,
        item_taxable=item_taxable,
        item_gst=item_gst,
        item_igst=item_igst,
        item_sale_price=item_sale_price,
        item_purchase_price=item_purchase_price,
        item_opening_stock=item_opening_stock,
        item_current_stock=item_current_stock,
        item_at_price=item_at_price,
        item_date=item_date,
        item_min_stock_maintain=item_min_stock_maintain
      )
      if ItemModel.objects.filter(item_name=item_name, item_hsn=item_hsn).exists():
        print('Item with the same item name and HSN  number already exists.')
        messages.error(request, 'Item with the same item name and HSN  number already exists.')
        val = 0
      elif ItemModel.objects.filter(item_name=item_name).exists():
        print('An item can have one HSN Number.')
        val = 0
        messages.error(request, 'An item can have one HSN Number.')
      elif ItemModel.objects.filter(item_hsn=item_hsn).exists():
        print('Item with the same HSN  number already exists.')
        val = 0
        messages.error(request, 'Item with the same HSN  number already exists.')
      else:
        print('yes')
        item_data.save()
        val = 1
        Item_History.objects.create(Item = item_data,company=com,staff=staff,action='Created').save()

      return JsonResponse({'status':True , 'value' : val})


def create_estimate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)

      latest_bill = Estimate.objects.filter(company = com).order_by('-id').first()

      if latest_bill:
          last_number = int(latest_bill.ref_no)
          new_number = last_number + 1
      else:
          new_number = 1

      if DeletedEstimate.objects.filter(company = com).exists():
          deleted = DeletedEstimate.objects.get(company = com)
          
          if deleted:
              while int(deleted.ref_no) >= new_number:
                  new_number+=1
      
      context = {
        'staff':staff, 'company':com,'allmodules':allmodules, 'parties':parties, 'ref_no':new_number,'items':items,'item_units':item_units,
      }
      return render(request, 'company/create_estimate.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def getPartyDetails(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  
    party_id = request.POST.get('id')
    party_details = party.objects.get(id = party_id)

    list = []
    dict = {
      'contact': party_details.contact,
      'address':party_details.address,
      'state': party_details.state,
      'balance':party_details.openingbalance,
      'payment':party_details.payment,
    }
    list.append(dict)
    return JsonResponse(json.dumps(list), content_type="application/json", safe=False)
    

def getItemData(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  
    try:
        id = request.GET.get('id')

        # item = ItemModel.objects.get(item_name = id, company = com)
        item = ItemModel.objects.filter(item_name = id, company = com).first()
        hsn = item.item_hsn
        pur_rate = item.item_purchase_price
        sale_rate = item.item_sale_price
        tax = True if item.item_taxable == "Taxable" else False
        gst = item.item_gst
        igst = item.item_igst

        return JsonResponse({"status":True,'id':item.id,'hsn':hsn,'pur_rate':pur_rate,'sale_rate':sale_rate, 'tax':tax, 'gst':gst, 'igst':igst})
    except Exception as e:
        print(e)
        return JsonResponse({"status":False})
  

def createNewEstimate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
        if request.method == 'POST':
            estimate = Estimate(
              staff = staff,
              company = com,
              date = request.POST['date'],
              ref_no = request.POST['ref_no'],
              party_name = party.objects.get(id = request.POST['party_name']).party_name,
              contact = request.POST['contact'],
              billing_address = request.POST['address'],
              state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State',
              description = request.POST['description'],
              subtotal = request.POST['subtotal'],
              cgst = request.POST['cgst_tax'],
              sgst = request.POST['sgst_tax'],
              igst = request.POST['igst_tax'],
              tax_amount = request.POST['tax_amount'],
              adjustment = request.POST['adjustment'],
              total_amount = request.POST['grand_total'],
              balance = 0,
              status = 'Open',
              is_converted = False
            )
            estimate.save()
            
            ids = request.POST.getlist('estItems[]')
            item = request.POST.getlist("item[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("price[]")
            tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
            discount = request.POST.getlist("discount[]")
            total = request.POST.getlist("total[]")

            est_id = Estimate.objects.get( id = estimate.id)

            if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total) and ids and item and hsn and qty and price and tax and discount and total:
                mapped = zip(ids,item,hsn,qty,price,tax,discount,total)
                mapped = list(mapped)
                for ele in mapped:
                  estItems = Estimate_items.objects.create(staff = staff, eid = est_id, company = com, item = ItemModel.objects.get(company = com, id = ele[0]),name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],discount = ele[6],total=ele[7])
            
            # Transaction history

            history = EstimateTransactionHistory(
              staff = staff,
              estimate = estimate,
              company = com,
              action = "Create"
            )
            history.save()

            if 'save_and_next' in request.POST:
                return redirect(create_estimate)
            return redirect(estimate_quotation)
    except Exception as e:
        print(e)
        return redirect(create_estimate)
  return redirect('/')


def getPartyList(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    options = {}
    option_objects = party.objects.filter(company = com)
    for option in option_objects:
        options[option.id] = [option.id , option.party_name]

    return JsonResponse(options)


def getItemList(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    options = {}
    option_objects = ItemModel.objects.filter(company = com)
    for option in option_objects:
        options[option.id] = [option.item_name]

    return JsonResponse(options)
  

def estimateFilterWithDate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      date = request.GET['date_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, date = date)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)      
      
      if not all_estimates:
        messages.warning(request, f'No Estimates found on {date}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'date_value':date,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

def estimateFilterWithRef(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      ref = request.GET['ref_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, ref_no = ref)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Ref No. {ref}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'ref_value':ref,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def estimateFilterWithBal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      bal = request.GET['bal_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, balance = bal)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Balance amount {bal}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'bal_value':bal,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def estimateFilterWithName(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      name = request.GET['name_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, party_name = name)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)      

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Party Name {name}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'name_value':name,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def estimateFilterWithTotal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      tot = request.GET['total_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, total_amount = tot)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Total Amount {tot}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)

      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'total_value':tot,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    
  
def estimateFilterWithStat(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      stat = request.GET['status']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com, status = stat)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)

      if not all_estimates:
        messages.warning(request, f'No Estimates found with Status {stat}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'stat_value':stat,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
   


def estimateInBetween(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      fromDate = request.GET['from_date']
      toDate = request.GET['to_date']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_estimates = Estimate.objects.filter(company = com).filter(date__gte = fromDate, date__lte = toDate)
      estimates = []
      for est in all_estimates:
        history = EstimateTransactionHistory.objects.filter(company = com, estimate = est).last()
        dict = {'estimate':est,'history':history}
        estimates.append(dict)
      
      if not all_estimates:
        messages.warning(request, f'No Estimates found in between {fromDate} to {toDate}.!')
        # estimates = Estimate.objects.filter(company = com)
        return redirect(estimate_quotation)      
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'estimates':estimates,'from':fromDate, 'to':toDate,
      }
      return render(request, 'company/estimate_quotation.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)


def deleteEstimate(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      est = Estimate.objects.get(company = com, id = id)

      # Storing ref number to deleted table
      # if entry exists and lesser than the current, update and save => Only one entry per company

      if DeletedEstimate.objects.filter(company = com).exists():
          deleted = DeletedEstimate.objects.get(company = com)
          if deleted:
              if int(est.ref_no) > int(deleted.ref_no):
                  deleted.ref_no = est.ref_no
                  deleted.save()
          
      else:
          deleted = DeletedEstimate(company = com, staff = staff, ref_no = est.ref_no)
          deleted.save()
      
      Estimate_items.objects.filter(company = com , eid = est).delete()
      est.delete()
      messages.success(request, 'Estimate deleted successfully.!')
      return redirect(estimate_quotation)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
  return redirect('/')


def editEstimate(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      est = Estimate.objects.get(company = com , id = id)
      est_items = Estimate_items.objects.filter(company = com , eid = est)
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'parties':parties,'items':items,'item_units':item_units, 'estimate':est, 'estItems':est_items,
      }
      return render(request, 'company/edit_estimate.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

def updateEstimate(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      estimate = Estimate.objects.get(company = com, id = id)
      if request.method == 'POST':
        estimate.date = request.POST['date']
        estimate.ref_no = request.POST['ref_no']
        estimate.party_name = party.objects.get(id = request.POST['party_name']).party_name
        estimate.contact = request.POST['contact']
        estimate.billing_address = request.POST['address']
        estimate.state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State'
        estimate.description = request.POST['description']
        estimate.subtotal = request.POST['subtotal']
        estimate.cgst = request.POST['cgst_tax']
        estimate.sgst = request.POST['sgst_tax']
        estimate.igst = request.POST['igst_tax']
        estimate.tax_amount = request.POST['tax_amount']
        estimate.adjustment = request.POST['adjustment']
        estimate.total_amount = request.POST['grand_total']
        estimate.balance = 0
        estimate.status = 'Open'
        estimate.is_converted = False

        estimate.save()

        ids = request.POST.getlist('estItems[]')
        item = request.POST.getlist("item[]")
        hsn  = request.POST.getlist("hsn[]")
        qty = request.POST.getlist("qty[]")
        price = request.POST.getlist("price[]")
        tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
        discount = request.POST.getlist("discount[]")
        total = request.POST.getlist("total[]")
        est_item_ids = request.POST.getlist("id[]")
        
        item_ids = [int(id) for id in est_item_ids]

        
        est_item = Estimate_items.objects.filter(eid = estimate)
        object_ids = [obj.id for obj in est_item]

        ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in item_ids]

        Estimate_items.objects.filter(id__in=ids_to_delete).delete()
        
        count = Estimate_items.objects.filter(eid = estimate, company = com).count()
        if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total):
            try:
                mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                mapped=list(mapped)
                
                for ele in mapped:
                    if int(len(item))>int(count):
                        if ele[8] == 0:
                            itemAdd= Estimate_items.objects.create(name = ele[1], hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7] ,staff = staff ,eid = estimate ,company = com, item = ItemModel.objects.get(company = com, id = ele[0]))
                        else:
                            itemAdd = Estimate_items.objects.filter( id = ele[8],company = com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
                    else:
                        itemAdd = Estimate_items.objects.filter( id = ele[8],company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
            except Exception as e:
                    print(e)
                    mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                    mapped=list(mapped)
                    
                    for ele in mapped:
                        created =Estimate_items.objects.filter(id=ele[8] ,company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
        
        history = EstimateTransactionHistory(
          staff = staff,
          estimate = estimate,
          company = com,
          action = "Edit"
        )
        history.save()

        return redirect(viewEstimate,id)
    except Exception as e:
      print(e)
      return redirect(editEstimate, id)
    

def estimateTransactionHistory(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      est = Estimate.objects.get(company = com, id = id)
      history = EstimateTransactionHistory.objects.filter(company = com, estimate = est)
      context = {
        'staff':staff, 'company':com,'allmodules':allmodules,'history':history,
      }
      return render(request, 'company/estimate_transaction_history.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

# DELIVERY CHALLAN

def createDeliveryChallan(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)

      # Fetching last bill and assigning upcoming bill no as current + 1
      # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
      latest_bill = DeliveryChallan.objects.filter(company = com).order_by('-id').first()

      if latest_bill:
          last_number = int(latest_bill.challan_no)
          new_number = last_number + 1
      else:
          new_number = 1

      if DeletedDeliveryChallan.objects.filter(company = com).exists():
          deleted = DeletedDeliveryChallan.objects.get(company = com)
          
          if deleted:
              while int(deleted.challan_no) >= new_number:
                  new_number+=1

      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'parties':parties, 'challan_no':new_number,'items':items,'item_units':item_units,
      }
      return render(request, 'company/create_delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def createNewDeliveryChallan(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
        if request.method == 'POST':
            challan = DeliveryChallan(
              company = com,
              staff = staff,
              date = request.POST['date'],
              due_date = request.POST['due_date'],
              challan_no = request.POST['challan_no'],
              party_name = party.objects.get(id = request.POST['party_name']).party_name,
              contact = request.POST['contact'],
              billing_address = request.POST['address'],
              state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State',
              description = request.POST['description'],
              subtotal = request.POST['subtotal'],
              cgst = request.POST['cgst_tax'],
              sgst = request.POST['sgst_tax'],
              igst = request.POST['igst_tax'],
              tax_amount = request.POST['tax_amount'],
              adjustment = request.POST['adjustment'],
              total_amount = request.POST['grand_total'],
              balance = 0,
              status = 'Open',
              is_converted = False
            )
            challan.save()
            
            ids = request.POST.getlist('dcItems[]')
            item = request.POST.getlist("item[]")
            hsn  = request.POST.getlist("hsn[]")
            qty = request.POST.getlist("qty[]")
            price = request.POST.getlist("price[]")
            tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
            discount = request.POST.getlist("discount[]")
            total = request.POST.getlist("total[]")

            chl_id = DeliveryChallan.objects.get( id = challan.id)

            if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total) and ids and item and hsn and qty and price and tax and discount and total:
                mapped = zip(ids,item,hsn,qty,price,tax,discount,total)
                mapped = list(mapped)
                for ele in mapped:
                  dcItems = DeliveryChallanItems.objects.create(staff = staff,cid = chl_id, company = com, item = ItemModel.objects.get(company = com, id = ele[0]),name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],discount = ele[6],total=ele[7])
            
            history = DeliveryChallanTransactionHistory(
              staff = staff,
              challan = challan,
              company = com,
              action = "Create"
            )
            history.save()

            if 'save_and_next' in request.POST:
                return redirect(createDeliveryChallan)
            return redirect(delivery_challan)
    except Exception as e:
        print(e)
        return redirect(createDeliveryChallan)
  return redirect('/')


def challanInBetween(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      fromDate = request.GET['from_date']
      toDate = request.GET['to_date']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com).filter(date__gte = fromDate, date__lte = toDate)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)
      if not all_challan:
        messages.warning(request, f'No Challans found in between {fromDate} to {toDate}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'from':fromDate, 'to':toDate,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithDate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      date = request.GET['date_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, date = date)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)
      if not all_challan:
        messages.warning(request, f'No Challans found on {date}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'date_value':date,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    

def challanFilterWithDueDate(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      date = request.GET['due_date_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, due_date = date)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Due Date {date}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'duedate_value':date,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    

def challanFilterWithChallanNo(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      chl = request.GET['challan_no_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, challan_no = chl)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Challan No. {chl}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'chno_value':chl,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithBal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      bal = request.GET['bal_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, balance = bal)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Balance amount {bal}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'bal_value':bal,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithName(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      name = request.GET['name_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, party_name = name)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Party Name {name}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'name_value':name,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def challanFilterWithTotal(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      tot = request.GET['total_filter_value']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, total_amount = tot)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)

      if not all_challan:
        messages.warning(request, f'No Challans found with Total Amount {tot}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)

      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'total_value':tot,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    
  
def challanFilterWithStat(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      stat = request.GET['status']
      allmodules= modules_list.objects.get(company=com.id,status='New')
      all_challan = DeliveryChallan.objects.filter(company = com, status = stat)
      challan = []
      for dc in all_challan:
        history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc).last()
        dict = {'challan':dc,'history':history}
        challan.append(dict)
        
      if not all_challan:
        messages.warning(request, f'No Challans found with Status {stat}.!')
        # challan = DeliveryChallan.objects.filter(company = com)
        return redirect(delivery_challan)
      
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'challan':challan,'stat_value':stat,
      }
      return render(request, 'company/delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def deleteChallan(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      challan = DeliveryChallan.objects.get(company = com, id = id)

      # Storing ref number to deleted table
      # if entry exists and lesser than the current, update and save => Only one entry per company

      if DeletedDeliveryChallan.objects.filter(company = com).exists():
          deleted = DeletedDeliveryChallan.objects.get(company = com)
          if deleted:
              if int(challan.challan_no) > int(deleted.challan_no):
                  deleted.challan_no = challan.challan_no
                  deleted.save()
          
      else:
          deleted = DeletedDeliveryChallan(company = com, staff = staff, challan_no = challan.challan_no)
          deleted.save()
      
      DeliveryChallanItems.objects.filter(company = com , cid = challan).delete()
      challan.delete()
      messages.success(request, 'Challan deleted successfully.!')
      return redirect(delivery_challan)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
  return redirect('/')


def editChallan(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      dc = DeliveryChallan.objects.get(company = com , id = id)
      dc_items = DeliveryChallanItems.objects.filter(company = com , cid = dc)
      allmodules= modules_list.objects.get(company=com.id,status='New')
      parties = party.objects.filter(company = com)
      items = ItemModel.objects.filter(company = com)
      item_units = UnitModel.objects.filter(company=com)
      context = {
        'staff':staff,'company':com,'allmodules':allmodules, 'parties':parties,'items':items,'item_units':item_units, 'challan':dc, 'dcItems':dc_items,
      }
      return render(request, 'company/edit_delivery_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)



def updateChallan(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      challan = DeliveryChallan.objects.get(company = com, id = id)
      if request.method == 'POST':
        challan.date = request.POST['date']
        challan.due_date = request.POST['due_date']
        challan.challan_no = request.POST['challan_no']
        challan.party_name = party.objects.get(id = request.POST['party_name']).party_name
        challan.contact = request.POST['contact']
        challan.billing_address = request.POST['address']
        challan.state_of_supply = 'State' if request.POST['state_supply'] == 'state' else 'Other State'
        challan.description = request.POST['description']
        challan.subtotal = request.POST['subtotal']
        challan.cgst = request.POST['cgst_tax']
        challan.sgst = request.POST['sgst_tax']
        challan.igst = request.POST['igst_tax']
        challan.tax_amount = request.POST['tax_amount']
        challan.adjustment = request.POST['adjustment']
        challan.total_amount = request.POST['grand_total']
        challan.balance = 0
        challan.status = 'Open'
        challan.is_converted = False

        challan.save()

        ids = request.POST.getlist('dcItems[]')
        item = request.POST.getlist("item[]")
        hsn  = request.POST.getlist("hsn[]")
        qty = request.POST.getlist("qty[]")
        price = request.POST.getlist("price[]")
        tax = request.POST.getlist("taxgst[]") if request.POST['state_supply'] == 'state' else request.POST.getlist("taxigst[]")
        discount = request.POST.getlist("discount[]")
        total = request.POST.getlist("total[]")
        dc_item_ids = request.POST.getlist("id[]")
        
        item_ids = [int(id) for id in dc_item_ids]

        
        dc_item = DeliveryChallanItems.objects.filter(cid = challan)
        object_ids = [obj.id for obj in dc_item]

        ids_to_delete = [obj_id for obj_id in object_ids if obj_id not in item_ids]

        DeliveryChallanItems.objects.filter(id__in=ids_to_delete).delete()
        
        count = DeliveryChallanItems.objects.filter(cid = challan, company = com).count()
        if len(ids)==len(item)==len(hsn)==len(qty)==len(price)==len(tax)==len(discount)==len(total):
            try:
                mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                mapped=list(mapped)
                
                for ele in mapped:
                    if int(len(item))>int(count):
                        if ele[8] == 0:
                            itemAdd= DeliveryChallanItems.objects.create(name = ele[1], hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7] ,cid = challan, staff = staff, company = com, item = ItemModel.objects.get(company = com, id = ele[0]))
                        else:
                            itemAdd = DeliveryChallanItems.objects.filter( id = ele[8],company = com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
                    else:
                        itemAdd = DeliveryChallanItems.objects.filter( id = ele[8],company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))
            except Exception as e:
                    print(e)
                    mapped=zip(ids,item,hsn,qty,price,tax,total,discount,item_ids)
                    mapped=list(mapped)
                    
                    for ele in mapped:
                        created =DeliveryChallanItems.objects.filter(id=ele[8] ,company=com).update(name = ele[1],hsn=ele[2],quantity=ele[3],price=ele[4],tax=ele[5],total=ele[6],discount=ele[7], item = ItemModel.objects.get(company = com, id = ele[0]))

        history = DeliveryChallanTransactionHistory(
          staff = staff,
          challan = challan,
          company = com,
          action = "Edit"
        )
        history.save()

        return redirect(viewChallan,id)
    except Exception as e:
      print(e)
      return redirect(editChallan, id)


def challanTransactionHistory(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      dc = DeliveryChallan.objects.get(company = com, id = id)
      history = DeliveryChallanTransactionHistory.objects.filter(company = com, challan = dc)
      context = {
        'staff':staff, 'company':com, 'allmodules':allmodules, 'history':history,
      }
      return render(request, 'company/delivery_challan_transaction_history.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)


def importEstimateFromExcel(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)    
    
    current_datetime = timezone.now()
    dateToday =  current_datetime.date()

    if request.method == "POST" and 'excel_file' in request.FILES:
    
        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)

        # checking estimate sheet columns
        try:
          ws = wb["estimate"]
        except:
          print('sheet not found')
          messages.error(request,'`estimate` sheet not found.! Please check.')
          return redirect(estimate_quotation)

        try:
          ws = wb["items"]
        except:
          print('sheet not found')
          messages.error(request,'`items` sheet not found.! Please check.')
          return redirect(estimate_quotation)
        
        ws = wb["estimate"]
        estimate_columns = ['SLNO','DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL']
        estimate_sheet = [cell.value for cell in ws[1]]
        if estimate_sheet != estimate_columns:
          print('invalid sheet')
          messages.error(request,'`estimate` sheet column names or order is not in the required formate.! Please check.')
          return redirect(estimate_quotation)

        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          if slno is None or state_of_supply is None or taxamount is None or grandtotal is None:
            print('estimate == invalid data')
            messages.error(request,'`estimate` sheet entries missing required fields.! Please check.')
            return redirect(estimate_quotation)
        
        # checking items sheet columns
        ws = wb["items"]
        items_columns = ['ESTIMATE NO','NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL']
        items_sheet = [cell.value for cell in ws[1]]
        if items_sheet != items_columns:
          print('invalid sheet')
          messages.error(request,'`items` sheet column names or order is not in the required formate.! Please check.')
          return redirect(estimate_quotation)

        for row in ws.iter_rows(min_row=2, values_only=True):
          est_no,name,hsn,quantity,price,tax_percentage,discount,total = row
          if est_no is None or name is None or quantity is None or tax_percentage is None or total is None:
            print('items == invalid data')
            messages.error(request,'`items` sheet entries missing required fields.! Please check.')
            return redirect(estimate_quotation)
        
        # getting data from estimate sheet and create estimate.
        incorrect_data = []
        ws = wb['estimate']
        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          estNo = slno
          if slno is None:
            continue
          # Fetching last bill and assigning upcoming bill no as current + 1
          # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
          latest_bill = Estimate.objects.filter(company = com).order_by('-id').first()
          
          if latest_bill:
              last_number = int(latest_bill.ref_no)
              new_number = last_number + 1
          else:
              new_number = 1

          if DeletedEstimate.objects.filter(company = com).exists():
              deleted = DeletedEstimate.objects.get(company = com)
              
              if deleted:
                  while int(deleted.ref_no) >= new_number:
                      new_number+=1
          if not party.objects.filter(company = com, party_name = name).exists():
            incorrect_data.append(slno)
            continue
          try:
            cntct = party.objects.get(company = com, party_name = name).contact
            adrs = party.objects.get(company = com, party_name = name).address
          except:
            pass

          if date is None:
            date = dateToday

          print(date,new_number,name,cntct,adrs,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal)

          estimate = Estimate(
              staff = staff,
              company = com,
              date = date,
              ref_no = new_number,
              party_name = name,
              contact = cntct,
              billing_address = adrs,
              state_of_supply = 'State' if str(state_of_supply).lower() == 'state' else 'Other State',
              description = description,
              subtotal = subtotal,
              cgst = cgst,
              sgst = sgst,
              igst = igst,
              tax_amount = taxamount,
              adjustment = adjustment,
              total_amount = grandtotal,
              balance = 0,
              status = 'Open',
              is_converted = False
          )
          estimate.save()

          # Transaction history
          history = EstimateTransactionHistory(
            staff = staff,
            estimate = estimate,
            company = com,
            action = "Create"
          )
          history.save()

          # Items for the estimate
          ws = wb['items']
          for row in ws.iter_rows(min_row=2, values_only=True):
            est_no,name,hsn,quantity,price,tax_percentage,discount,total = row
            if int(est_no) == int(estNo):
              print(row)
              if estimate.state_of_supply == 'State' and tax_percentage:
                tx = 'GST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              elif estimate.state_of_supply == 'Other State' and tax_percentage:
                tx = 'IGST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              if discount is None:
                discount=0
              if price is None:
                price=0
              if not ItemModel.objects.filter(company = com, item_name = name).exists():
                incorrect_data.append(est_no)
                continue
              try:
                itm = ItemModel.objects.get(company = com, item_name = name)
              except:
                pass
              Estimate_items.objects.create(staff = staff, eid = estimate, company = com, item = itm,name = name,hsn=hsn,quantity=int(quantity),price = float(price),tax=tx, discount = float(discount),total=float(total))
    messages.success(request, 'Data imported successfully.!')
    if incorrect_data:
      messages.warning(request, f'Data with following SlNo could not import due to incorrect data provided - {", ".join(str(item) for item in incorrect_data)}')
    return redirect(estimate_quotation)
  

def importChallanFromExcel(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)    
    
    current_datetime = timezone.now()
    dateToday =  current_datetime.date()

    if request.method == "POST" and 'excel_file' in request.FILES:
    
        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)

        # checking challan sheet columns
        try:
          ws = wb["challan"]
        except:
          print('sheet not found')
          messages.error(request,'`challan` sheet not found.! Please check.')
          return redirect(delivery_challan)

        try:
          ws = wb["items"]
        except:
          print('sheet not found')
          messages.error(request,'`items` sheet not found.! Please check.')
          return redirect(delivery_challan)
        
        ws = wb["challan"]
        estimate_columns = ['SLNO','DATE','DUE DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL']
        estimate_sheet = [cell.value for cell in ws[1]]
        if estimate_sheet != estimate_columns:
          print('invalid sheet')
          messages.error(request,'`challan` sheet column names or order is not in the required formate.! Please check.')
          return redirect(delivery_challan)

        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,due_date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          if slno is None or state_of_supply is None or taxamount is None or grandtotal is None:
            print('challan == invalid data')
            messages.error(request,'`challan` sheet entries missing required fields.! Please check.')
            return redirect(delivery_challan)
        
        # checking items sheet columns
        ws = wb["items"]
        items_columns = ['CHALLAN NO','NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL']
        items_sheet = [cell.value for cell in ws[1]]
        if items_sheet != items_columns:
          print('invalid sheet')
          messages.error(request,'`items` sheet column names or order is not in the required formate.! Please check.')
          return redirect(delivery_challan)

        for row in ws.iter_rows(min_row=2, values_only=True):
          chl_no,name,hsn,quantity,price,tax_percentage,discount,total = row
          if chl_no is None or name is None or quantity is None or tax_percentage is None or total is None:
            print('items == invalid data')
            messages.error(request,'`items` sheet entries missing required fields.! Please check.')
            return redirect(delivery_challan)
        
        # getting data from estimate sheet and create estimate.
        incorrect_data = []
        ws = wb['challan']
        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,due_date,name,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal = row
          dcNo = slno
          if slno is None:
            continue
          # Fetching last bill and assigning upcoming bill no as current + 1
          # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
          latest_bill = DeliveryChallan.objects.filter(company = com).order_by('-id').first()
          
          if latest_bill:
              last_number = int(latest_bill.challan_no)
              new_number = last_number + 1
          else:
              new_number = 1

          if DeletedDeliveryChallan.objects.filter(company = com).exists():
              deleted = DeletedDeliveryChallan.objects.get(company = com)
              
              if deleted:
                  while int(deleted.challan_no) >= new_number:
                      new_number+=1
          if not party.objects.filter(company = com, party_name = name).exists():
            incorrect_data.append(slno)
            continue
          try:
            cntct = party.objects.get(company = com, party_name = name).contact
            adrs = party.objects.get(company = com, party_name = name).address
          except:
            pass

          if date is None:
            date = dateToday

          if due_date is None:
            due_date = dateToday

          print(date,due_date,new_number,name,cntct,adrs,state_of_supply,description,subtotal,igst,cgst,sgst,taxamount,adjustment,grandtotal)

          challan = DeliveryChallan(
              staff = staff,
              company = com,
              date = date,
              due_date = due_date,
              challan_no = new_number,
              party_name = name,
              contact = cntct,
              billing_address = adrs,
              state_of_supply = 'State' if str(state_of_supply).lower() == 'state' else 'Other State',
              description = description,
              subtotal = subtotal,
              cgst = cgst,
              sgst = sgst,
              igst = igst,
              tax_amount = taxamount,
              adjustment = adjustment,
              total_amount = grandtotal,
              balance = 0,
              status = 'Open',
              is_converted = False
          )
          challan.save()

          # Transaction history
          history = DeliveryChallanTransactionHistory(
            staff = staff,
            challan = challan,
            company = com,
            action = "Create"
          )
          history.save()

          # Items for the estimate
          ws = wb['items']
          for row in ws.iter_rows(min_row=2, values_only=True):
            chl_no,name,hsn,quantity,price,tax_percentage,discount,total = row
            if int(chl_no) == int(dcNo):
              print(row)
              if challan.state_of_supply == 'State' and tax_percentage:
                tx = 'GST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              elif challan.state_of_supply == 'Other State' and tax_percentage:
                tx = 'IGST'+str(tax_percentage)+'['+str(tax_percentage)+'%]'
              if discount is None:
                discount=0
              if price is None:
                price=0
              if not ItemModel.objects.filter(company = com, item_name = name).exists():
                incorrect_data.append(chl_no)
                continue
              try:
                itm = ItemModel.objects.get(company = com, item_name = name)
              except:
                pass
              DeliveryChallanItems.objects.create(staff = staff, cid = challan, company = com, item = itm,name = name,hsn=hsn,quantity=int(quantity),price = float(price),tax=tx, discount = float(discount),total=float(total))
    messages.success(request, 'Data imported successfully.!')
    if incorrect_data:
      messages.warning(request, f'Data with following SlNo could not import due to incorrect data provided - {", ".join(str(item) for item in incorrect_data)}')
    return redirect(delivery_challan)



def downloadEstimateSampleImportFile(request):
    
    estimate_table_data = [['SLNO','DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL'], ['1', '2023-11-20', 'Alwin', 'State', 'Sample Description','1000','0','25','25','50','0','1050']]
    items_table_data = [['ESTIMATE NO', 'NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL'], ['1', 'Test Item 1','789987','1','1000','5','0','1000']]

    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'estimate'
    sheet2 = wb.create_sheet(title='items')

    # Populate the sheets with data
    for row in estimate_table_data:
        sheet1.append(row)

    for row in items_table_data:
        sheet2.append(row)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=estimate_sample_file.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def downloadChallanSampleImportFile(request):
    
    challan_table_data = [['SLNO','DATE','DUE DATE','NAME','STATE OF SUPPLY','DESCRIPTION','SUB TOTAL','IGST','CGST','SGST','TAX AMOUNT','ADJUSTMENT','GRAND TOTAL'], ['1', '2023-11-20', '2023-11-20', 'Alwin', 'State', 'Sample Description','1000','0','25','25','50','0','1050']]
    items_table_data = [['CHALLAN NO', 'NAME','HSN','QUANTITY','PRICE','TAX PERCENTAGE','DISCOUNT','TOTAL'], ['1', 'Test Item 1','788987','1','1000','5','0','1000']]

    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = 'challan'
    sheet2 = wb.create_sheet(title='items')

    # Populate the sheets with data
    for row in challan_table_data:
        sheet1.append(row)

    for row in items_table_data:
        sheet2.append(row)

    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=challan_sample_file.xlsx'

    # Save the workbook to the response
    wb.save(response)

    return response


def estimateBillPdf(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  

    bill = Estimate.objects.get(company = com, id = id)
    items = Estimate_items.objects.filter(company = com, eid = bill)

    total = bill.total_amount
    words_total = num2words(total)
    
    context = {'staff':staff,'bill': bill, 'company': com,'items':items, 'total':words_total}
    
    template_path = 'company/estimate_bill_pdf.html'
    fname = 'bill'+str(bill.ref_no)

    # return render(request, 'staff/estimate_bill_pdf.html',context)
    # Create a Django response object, and specify content_type as pdftemp_creditnote
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] =f'attachment; filename = Estimate_{fname}.pdf'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def challanBillPdf(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)  

    bill = DeliveryChallan.objects.get(company = com, id = id)
    items = DeliveryChallanItems.objects.filter(company = com, cid = bill)

    total = bill.total_amount
    words_total = num2words(total)
    
    context = {'staff':staff,'bill': bill, 'company': com,'items':items, 'total':words_total}
    
    template_path = 'company/challan_bill_pdf.html'
    fname = 'bill'+str(bill.challan_no)

    # return render(request, 'staff/challan_bill_pdf.html',context)
    # Create a Django response object, and specify content_type as pdftemp_creditnote
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] =f'attachment; filename = DeliveryChallan_{fname}.pdf'
    # find the template and render it.
    template = get_template(template_path)
    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funny view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def viewEstimate(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      bill = Estimate.objects.get(company = com, id = id)
      items = Estimate_items.objects.filter(company = com , eid = bill)
      context= {
        'staff':staff, 'company':com, 'bill':bill, 'items': items,'allmodules':allmodules
      }
      return render(request, 'company/view_estimate.html',context)
    except Exception as e:
      print(e)
      return redirect(estimate_quotation)
    

def viewChallan(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com.id,status='New')
    try:
      bill = DeliveryChallan.objects.get(company = com, id = id)
      items = DeliveryChallanItems.objects.filter(company = com , cid = bill)
      context= {
        'staff':staff, 'company':com, 'bill':bill, 'items': items,'allmodules':allmodules
      }
      return render(request, 'company/view_challan.html',context)
    except Exception as e:
      print(e)
      return redirect(delivery_challan)
    

def addNewParty(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)

    if request.method == 'POST':
      Company = company.objects.get(id = staff.company.id)
      user_id = request.user.id
      
      party_name = request.POST['partyname']
      gst_no = request.POST['gstno']
      contact = request.POST['contact']
      gst_type = request.POST['gst']
      state = request.POST['state']
      address = request.POST['address']
      email = request.POST['email']
      openingbalance = request.POST.get('balance', '')
      payment = request.POST.get('paymentType', '')
      creditlimit = request.POST.get('creditlimit', '')
      current_date = request.POST['currentdate']
      End_date = request.POST.get('enddate', None)
      additionalfield1 = request.POST['additionalfield1']
      additionalfield2 = request.POST['additionalfield2']
      additionalfield3 = request.POST['additionalfield3']
      comp=Company
      if (not party_name):
        return render(request, 'add_parties.html')

      part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                      creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,company=comp)
      part.save()

      return JsonResponse({'status':True})


def addNewItem(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    if request.method=='POST':
      company_user_data = com
      item_name = request.POST.get('item_name')
      item_hsn = request.POST.get('item_hsn')
      item_unit = request.POST.get('item_unit')
      item_taxable = request.POST.get('item_taxable')
      item_gst = request.POST.get('item_gst')
      item_igst = request.POST.get('item_igst')
      item_sale_price = 0 if request.POST.get('item_sale_price') is '' or None else request.POST.get('item_sale_price')
      item_purchase_price = 0 if request.POST.get('item_purchase_price') is '' or None else request.POST.get('item_purchase_price')
      item_opening_stock = request.POST.get('item_opening_stock')
      item_current_stock = item_opening_stock
      if item_opening_stock == '' or None :
        item_opening_stock = 0
        item_current_stock = 0
      item_at_price = request.POST.get('item_at_price')
      if item_at_price == '' or None:
        item_at_price =0
      item_date = request.POST.get('item_date')
      item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
      if item_min_stock_maintain == ''  or None:
        item_min_stock_maintain = 0
      item_data = ItemModel(company=company_user_data,
        item_name=item_name,
        item_hsn=item_hsn,
        item_unit=item_unit,
        item_taxable=item_taxable,
        item_gst=item_gst,
        item_igst=item_igst,
        item_sale_price=item_sale_price,
        item_purchase_price=item_purchase_price,
        item_opening_stock=item_opening_stock,
        item_current_stock=item_current_stock,
        item_at_price=item_at_price,
        item_date=item_date,
        item_min_stock_maintain=item_min_stock_maintain
      )
      item_data.save()

      return JsonResponse({'status':True})
    

# ===================end ---shemeem =============================


# ----------athul-22-11-2023--------

def register(request):
  if request.method == 'POST':
    first_name = request.POST['fname']
    last_name = request.POST['lname']
    user_name = request.POST['uname']
    email_id = request.POST['eid']
    mobile = request.POST['ph']
    passw = request.POST['pass']
    c_passw = request.POST['cpass']
    action = request.POST['r']
    did = request.POST['did']
    if did != '':
      if Distributors_details.objects.filter(distributor_id=did).exists():
        distributor = Distributors_details.objects.get(distributor_id=did)
      else :
          messages.info(request, 'Sorry, distributor id does not exists')
          return redirect('company_reg')
    

    
    if passw == c_passw:
      if User.objects.filter(username = user_name).exists():
        messages.info(request, 'Sorry, Username already exists')
        return redirect('company_reg')
      

      elif not User.objects.filter(email = email_id).exists():
        
        user_data = User.objects.create_user(first_name = first_name,
                        last_name = last_name,
                        username = user_name,
                        email = email_id,
                        password = passw)
        user_data.save()
        if did != '':
          data = User.objects.get(id = user_data.id)
          cust_data = company( contact=mobile,
                             user = data,reg_action=action,Distributors=distributor)
          cust_data.save()
          demo_staff=staff_details(company=cust_data,
                                   email=email_id,
                                   position='company',
                                   user_name=user_name,
                                   password=passw,
                                   contact=mobile)
          demo_staff.save()
          return redirect('company_reg2')
        else:
          data = User.objects.get(id = user_data.id)
          cust_data = company( contact=mobile,
                             user = data,reg_action=action)
          cust_data.save()
          demo_staff=staff_details(company=cust_data,
                                   email=email_id,
                                   position='company',
                                   user_name=user_name,
                                   password=passw,
                                   contact=mobile)
          demo_staff.save()

          print(demo_staff.company.company_name)
        
          return redirect('company_reg2')
      else:
        messages.info(request, 'Sorry, Email already exists')
        return redirect('company_reg')
    return render(request,'company/register.html')
  
def Allmodule(request,uid):
  user=User.objects.get(id=uid)
  return render(request,'company/modules.html',{'user':user})

def addmodules(request,uid):
  if request.method == 'POST':
    com=company.objects.get(user=uid)
    c1=request.POST.get('c1')
    c2=request.POST.get('c2')
    c3=request.POST.get('c3')
    c4=request.POST.get('c4')
    c5=request.POST.get('c5')
    c6=request.POST.get('c6')
    c7=request.POST.get('c7')
    c8=request.POST.get('c8')
    c9=request.POST.get('c9')
    c10=request.POST.get('c10')
    c11=request.POST.get('c11')
    c12=request.POST.get('c12')
    c13=request.POST.get('c13')
    c14=request.POST.get('c14')
    c15=request.POST.get('c15')
    
    data=modules_list(company=com,sales_invoice = c1,
                      Estimate=c2,Payment_in=c3,sales_order=c4,
                      Delivery_challan=c5,sales_return=c6,Purchase_bills=c7,
                      Payment_out=c8,Purchase_order=c9,Purchase_return=c10,
                      Bank_account=c11,Cash_in_hand=c12, cheques=c13,Loan_account=c14,Upi=c15)
    data.save()

    return redirect('log_page')
    
def adminaccept(request,id):
  data=company.objects.filter(id=id).update(superadmin_approval=1)
  data1=staff_details.objects.filter(company=id,position='company').update(Action=1)
  return redirect('client_request')
def adminreject(request,id):
  data1=staff_details.objects.get(company=id,position='company')
  data1.delete()
  data=company.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('client_request')


def log_page(request):
  return render(request, 'log.html')
  
def login(request):
  if request.method == 'POST':
    user_name = request.POST['username']
    passw = request.POST['password']
    
    log_user = auth.authenticate(username = user_name,
                                  password = passw)
    
    if log_user is not None:
      auth.login(request, log_user)

      # ---super admin---

      if request.user.is_staff==1:
        return redirect('adminhome')
      
      if Distributors_details.objects.filter(user=request.user).exists():
        data=Distributors_details.objects.get(user=request.user)
        if data.Log_Action == 1:
            return redirect('distributor_home')
        else:
            messages.info(request, 'Approval is Pending..')
            return redirect('log_page')
        
    if staff_details.objects.filter(user_name=user_name,password=passw,position='company').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='company') 

      if data.company.superadmin_approval == 1 or data.company.Distributor_approval == 1:
        request.session["staff_id"]=data.id
        if 'staff_id' in request.session:
          if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
            print(staff_id)
 
          return redirect('homepage')  
      else:
        messages.info(request, 'Approval is Pending..')
        return redirect('log_page')
      
    if staff_details.objects.filter(user_name=user_name,password=passw,position='staff').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='staff')   
      if data.Action == 1:
        request.session["staff_id"]=data.id
        if 'staff_id' in request.session:
          if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
            print(staff_id)
 
            return redirect('staffhome')  
      else:
        messages.info(request, 'Approval is Pending..')
        return redirect('log_page')
    else:
      messages.info(request, 'Invalid Username or Password. Try Again.')
      return redirect('log_page')  
  else:  
   return redirect('log_page')   
  

def homepage(request):
 
  staff_id = request.session['staff_id']
  print(staff_id)
       
  staff =  staff_details.objects.get(id = staff_id)
  print(staff.position)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules
          }
  return render(request, 'company/homepage.html', context)

def staff_request(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  data = staff_details.objects.filter(company=staff.company.id,Action=0,position='staff').order_by('-id')
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/staff_request.html',{'staff':staff,'data':data,'allmodules':allmodules}) 

# @login_required(login_url='login')
def staffhome(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  

  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request, 'staff/staffhome.html', context)


 
def View_staff(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  data = staff_details.objects.filter(company=staff.company.id,Action=1,position='staff').order_by('-id')
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')

  return render(request, 'company/view_staff.html',{'staff':staff,'data':data,'allmodules':allmodules})

def Companyprofile(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/companyprofile.html',{'staff':staff,'allmodules':allmodules})

def editcompanyprofile(request):
  staff_id = request.session['staff_id']
  print(staff_id)    
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  terms=payment_terms.objects.all()
  return render(request,'company/editcompanyprofile.html',{'staff':staff,'allmodules':allmodules,'terms':terms})

def editcompanyprofile_action(request):
  staff_id = request.session['staff_id']
  print(staff_id) 
  staff =  staff_details.objects.get(id = staff_id)
 
  if request.method == 'POST':
    staff.company.company_name = request.POST['cname']
    staff.company.user.email = request.POST['email']

    staff.email = request.POST['email']

    staff.company.contact = request.POST['ph']

    staff.contact = request.POST['ph']

    staff.company.address = request.POST['address']
    staff.company.city = request.POST['city']
    staff.company.state = request.POST['state']
    staff.company.country = request.POST['country']
    staff.company.pincode = request.POST['pincode']

    t = request.POST['select']
    terms = payment_terms.objects.get(id=t)
    staff.company.dateperiod = terms
    staff.company.start_date=date.today()
    days=int(terms.days)

    end= date.today() + timedelta(days=days)
    staff.company.End_date=end

    old=staff.company.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      staff.company.profile_pic=old
    else:
      staff.company.profile_pic=new
    
    staff.company.save() 
    staff.company.user.save() 
    staff.save()
    return redirect('Companyprofile') 



  return redirect('Companyprofile')


def editmodule(request):
  staff_id = request.session['staff_id']
  print(staff_id) 
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/editmodule.html',{'staff':staff,'allmodules':allmodules})

def editmodule_action(request):
  if request.method == 'POST':
    staff_id = request.session['staff_id']
    print(staff_id) 
    staff =  staff_details.objects.get(id = staff_id)
    com = company.objects.get(id = staff.company.id)
    # if modules_list.objects.filter(company=com.id,status='Old').exists():
    #   old=modules_list.objects.filter(company=com.id,status='Old')
    #   old.delete()

    # old_data=modules_list.objects.get(company=com.id,status='New')  
    # old_data.status='Old'
    # old_data.save()



    c1=request.POST.get('c1')
    c2=request.POST.get('c2')
    c3=request.POST.get('c3')
    c4=request.POST.get('c4')
    c5=request.POST.get('c5')
    c6=request.POST.get('c6')
    c7=request.POST.get('c7')
    c8=request.POST.get('c8')
    c9=request.POST.get('c9')
    c10=request.POST.get('c10')
    c11=request.POST.get('c11')
    c12=request.POST.get('c12')
    c13=request.POST.get('c13')
    c14=request.POST.get('c14')
    c15=request.POST.get('c15')
    
    data=modules_list(company=com,sales_invoice = c1,
                      Estimate=c2,Payment_in=c3,sales_order=c4,
                      Delivery_challan=c5,sales_return=c6,Purchase_bills=c7,
                      Payment_out=c8,Purchase_order=c9,Purchase_return=c10,
                      Bank_account=c11,Cash_in_hand=c12, cheques=c13,Loan_account=c14,Upi=c15,status='Pending')
    data.save()
    data1=modules_list.objects.filter(company=com.id,status='Pending').update(update_action=1)
    return redirect('Companyprofile')
    
    
  return redirect('Companyprofile')


def companyreport(request):
  staff_id = request.session['staff_id']
  print(staff_id) 
  staff =  staff_details.objects.get(id = staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request,'company/companyreport.html',{'staff':staff,'allmodules':allmodules}) 





def staff_profile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'staff/staff_profile.html',context)

def editstaff_profile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'staff/editstaff_profile.html',context)

def editstaff_profile_action(request):
  if request.method == 'POST':
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    staff.first_name = request.POST['fname']
    staff.last_name = request.POST['lname']
    staff.user_name = request.POST['uname']
    staff.email = request.POST['email']
    staff.contact = request.POST['ph']
    old=staff.img
    new=request.FILES.get('image')
    if old!=None and new==None:
      staff.img=old
    else:
      staff.img=new

    staff.save()  

    return redirect ('staff_profile')
  return redirect ('staff_profile')



def view_party(request,id):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  getparty=party.objects.get(id=id,company=staff.company.id)
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'company/view_party.html',{'staff':staff,'allmodules':allmodules,'Party':Party,'getparty':getparty})



#______________Sales Invoice_________________Antony Tom___________________________

def itemdetailinvoice(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  price = itm.item_sale_price
  return JsonResponse({'hsn':hsn, 'price':price}) 

def add_salesinvoice(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  company_instance = company.objects.get(id=staff.company.id)

  Party=party.objects.filter(company=company_instance)
  item=ItemModel.objects.filter(company=company_instance)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  bank=BankModel.objects.filter(company=company_instance)
  toda = date.today()
  todate = toda.strftime("%Y-%m-%d")
  if SalesInvoice.objects.filter(company=company_instance).exists():
        invoice_count = SalesInvoice.objects.last().invoice_no
        next_count = invoice_count+1
  else:
        next_count=1

  return render(request, 'company/add_salesinvoice.html',{'staff':staff,'Party':Party,'item':item,'bank':bank,'count':next_count,'allmodules':allmodules,'todate':todate})

def party_details(request, party_name):
    try:
        Party = party.objects.get(party_name=party_name)
        data = {
            'contact': Party.contact,
            'address': Party.address,
            'openingbalance': Party.openingbalance,
            'payment': Party.payment,
        }
        return JsonResponse(data)
    except party.DoesNotExist:
        return JsonResponse({'error': 'Party not found'},status=404)

def itemdata_salesinvoice(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_sale_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})
    

def itemdata_salesinvoiceedit(request):
  itmid = request.GET['id']
  print(itmid)
  itm = ItemModel.objects.get(id=itmid)
  print(itm)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_sale_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})


def save_sales_invoice(request):

    if request.method == 'POST':
        
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = staff.company 
        
        party_name = request.POST.get('partyname')
        contact = request.POST.get('contact')
        address = request.POST.get('address')
        invoice_no = request.POST.get('invoiceno')
        date = request.POST.get('date')
        state_of_supply = request.POST.get('state_of_supply')
        paymenttype = request.POST.get('bank')
        cheque = request.POST.get('chequeNumber')
        upi = request.POST.get('upiNumber')
        accountno = request.POST.get('accountNumber')
        product = tuple(request.POST.getlist("product[]"))
        hsn =  tuple(request.POST.getlist("hsn[]"))
        qty =  tuple(request.POST.getlist("qty[]"))
        rate =  tuple(request.POST.getlist("price[]"))
        discount =  tuple(request.POST.getlist("discount[]"))
        tax =  tuple(request.POST.getlist("tax[]"))
        total =  tuple(request.POST.getlist("total[]"))
        description = request.POST.get('description')
        advance = request.POST.get("advance")
        balance = request.POST.get("balance")
        subtotal = float(request.POST.get('subtotal'))
        igst = request.POST.get('igst')
        cgst = request.POST.get('cgst')
        sgst = request.POST.get('sgst')
        adjust = request.POST.get("adj")
        taxamount = request.POST.get("taxamount")
        grandtotal=request.POST.get('grandtotal')

        party_instance=party.objects.get(party_name=party_name)
        
      
        sales_invoice = SalesInvoice(
            staff=staff,
            company=company_instance,
            party=party_instance,
            party_name=party_name,
            contact=contact,
            address=address,
            invoice_no=invoice_no,
            date=date,
            state_of_supply=state_of_supply,
            paymenttype=paymenttype,
            cheque=cheque,
            upi=upi,
            accountno=accountno,
            description=description,
            subtotal=subtotal,
            igst=igst,
            cgst=cgst,
            sgst=sgst,
            total_taxamount=taxamount,
            adjustment=adjust,
            grandtotal=grandtotal,
            paidoff=advance,
            totalbalance=balance,
        )
    
        sales_invoice.save()

        tr_history = SalesInvoiceTransactionHistory(company=company_instance,
                                              staff=staff,      
                                              salesinvoice=sales_invoice,
                                              action="CREATED",
                                              done_by_name=staff.first_name,
                                              )
        tr_history.save()

        invoice = SalesInvoice.objects.get(id=sales_invoice.id)
        mapped = []  # Initialize mapped
        if len(product)==len(hsn)==len(qty)==len(rate)==len(discount)==len(tax)==len(total):
          mapped=zip(product, hsn, qty, rate, discount, tax, total)
          mapped=list(mapped)
        for ele in mapped: 
          itm = ItemModel.objects.get(id=ele[0])
          SalesInvoiceItem.objects.create(item=itm, hsn=ele[1], quantity=ele[2], rate=ele[3], discount=ele[4], tax=ele[5], totalamount=ele[6], salesinvoice=invoice, company=company_instance)


        

        if 'save_and_new' in request.POST:
            return render(request, 'staff/add_salesinvoice.html')
        else:
            return redirect('view_salesinvoice')

    return render(request, 'company/add_salesinvoice.html')

def view_salesinvoice(request):
    if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
        else:
            return redirect('/')
    
    staff = staff_details.objects.get(id=staff_id)
    print(staff)
    company_instance = company.objects.get(id=staff.company.id)
    print(company_instance)
    Party = party.objects.filter(company=company_instance)
    item = ItemModel.objects.filter(company=company_instance)
    allmodules= modules_list.objects.get(company=staff.company.id,status='New')
    
    salesinvoice = SalesInvoice.objects.filter(company=company_instance)
    for i in salesinvoice:
        last_transaction = SalesInvoiceTransactionHistory.objects.filter(salesinvoice=i).last()
        if last_transaction:
            i.action = last_transaction.action
            i.done_by_name = last_transaction.done_by_name
        else:
            i.action = None
            i.done_by_name = None

    return render(request, 'company/view_salesinvoice.html', {'staff':staff,'Party': Party, 'item': item, 'salesinvoice': salesinvoice,'allmodules':allmodules})



def get_bank_details(request, bank_name):
    try:
        bank = BankModel.objects.get(bank_name=bank_name)
        data = {
            'accountNumber': bank.account_num, 
          
        }
        return JsonResponse(data)
    except ItemModel.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def get_total_balance(request):
    total_balance = SalesInvoice.objects.aggregate(models.Sum('totalbalance'))['totalbalance__sum']
    total_balance = total_balance if total_balance is not None else 0.00

    paid_off = SalesInvoice.objects.aggregate(models.Sum('paidoff'))['paidoff__sum']
    paid_off = paid_off if paid_off is not None else 0.00

    grand_total = SalesInvoice.objects.aggregate(models.Sum('grandtotal'))['grandtotal__sum']
    grand_total = grand_total if grand_total is not None else 0.00
    return JsonResponse({'total_balance': total_balance,'paid_off':paid_off,'grand_total':grand_total})


def edit_salesinvoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  company_instance = company.objects.get(id=staff.company.id)
  getinoice=SalesInvoice.objects.get(id=id,company=company_instance)
  getitem=SalesInvoiceItem.objects.filter(salesinvoice=id,company=company_instance)
  Party=party.objects.filter(company=company_instance)
  item=ItemModel.objects.filter(company=company_instance)
  bank=BankModel.objects.filter(company=company_instance)
  toda = date.today()
  todate = toda.strftime("%Y-%m-%d")
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')

  return render(request, 'company/edit_salesinvoice.html',{'staff':staff,'getinoice':getinoice,'todate':todate,'getitem':getitem,'Party':Party,'item':item,'bank':bank,'allmodules':allmodules})


def editsave_salesinvoice(request,id):

    if request.method == 'POST':
        
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = company.objects.get(id=staff.company.id)
        sales_invoice=SalesInvoice.objects.get(id=id,company=company_instance,staff=staff)
        
        sales_invoice.party_name = request.POST.get('partyname')
        sales_invoice.contact = request.POST.get('contact')
        sales_invoice.address = request.POST.get('address')
        sales_invoice.invoice_no = request.POST.get('invoiceno')
        sales_invoice.date = request.POST.get('date')
        sales_invoice.state_of_supply = request.POST.get('state_of_supply')
        sales_invoice.paymenttype = request.POST.get('bank')
        sales_invoice.cheque = request.POST.get('chequeNumber')
        sales_invoice.upi = request.POST.get('upiNumber')
        sales_invoice.accountno = request.POST.get('accountNumber')
        sales_invoice.description = request.POST.get('description')
        sales_invoice.subtotal =float(request.POST.get('subtotal'))
        sales_invoice.igst = request.POST.get('igst')
        sales_invoice.cgst = request.POST.get('cgst')
        sales_invoice.sgst = request.POST.get('sgst')
        sales_invoice.total_taxamount = request.POST.get('taxamount')
        sales_invoice.adjustment = request.POST.get('adj')
        sales_invoice.grandtotal = request.POST.get('grandtotal')
        sales_invoice.paidoff = request.POST.get('advance')
        sales_invoice.totalbalance = request.POST.get('balance')
    
        sales_invoice.save()

        product = tuple(request.POST.getlist("product[]"))
        qty = tuple(request.POST.getlist("qty[]"))
        tax =tuple( request.POST.getlist("tax[]"))
        discount = tuple(request.POST.getlist("discount[]"))
        total = tuple(request.POST.getlist("total[]"))
        SalesInvoiceItem.objects.filter(salesinvoice=sales_invoice,company=company_instance).delete()
        if len(product)==len(qty)==len(qty)==len(tax):
          mapped=zip(product,qty,tax,discount,total)
          mapped=list(mapped)
          for ele in mapped:
            itm = ItemModel.objects.get(id=ele[0])
            SalesInvoiceItem.objects.create(item =itm,quantity=ele[1], tax=ele[2],discount=ele[3],totalamount=ele[4],salesinvoice=sales_invoice,company=company_instance)

        tr_history = SalesInvoiceTransactionHistory(company=company_instance,
                                              staff=staff,      
                                              salesinvoice=sales_invoice,
                                              action="UPDATED",
                                              done_by_name=staff.first_name,
                                              )
        tr_history.save()

        
        return redirect('view_salesinvoice')

    return render(request, 'company/edit_salesinvoice.html')


def salesinvoice_save_parties(request):
    if request.method == 'POST':
        if 'staff_id' in request.session:
            staff_id = request.session['staff_id']
        else:
            return redirect('/')

        staff = staff_details.objects.get(id=staff_id)
        company_instance = staff.company 
        
        party_name = request.POST['partyname']
        gst_no = request.POST['gstin']
        contact = request.POST['partyphno']
        gst_type = request.POST['modalgsttype']
        state = request.POST['splystate']
        address = request.POST['baddress']
        email = request.POST['partyemail']
        openingbalance = request.POST.get('openbalance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('crd_lmt', '')
        current_date = request.POST['partydate']
        additionalfield1 = request.POST['additional1']
        additionalfield2 = request.POST['additional2']
        additionalfield3 = request.POST['additional3']
        comp=company_instance
        if (
          not party_name
          
      ):
          return render(request, 'add_salesinvoice.html')

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,company=comp)
        part.save() 

        return redirect('add_salesinvoice')

    return render(request, 'company/add_salesinvoice.html')  


def deletesalesinvoice(request,id):
    salesinvoice=SalesInvoice.objects.get(id=id)
    salesinvoiceitem = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice)
    salesinvoice.delete()
    salesinvoiceitem.delete()
    return redirect('view_salesinvoice')


from django.http import JsonResponse

from django.db.models import Sum

def profit_loss_data(request, year=None):
    if 'staff_id' in request.session:
        staff_id = request.session['staff_id']
    else:
        return redirect('/')

    staff = staff_details.objects.get(id=staff_id)
    company_instance = company.objects.get(id=staff.company.id)
    labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sept", "Oct", "Nov", "Dec"]


    sales_data = (
        SalesInvoice.objects.filter(date__year=year, company=company_instance)
        .values('date__month')
        .annotate(grandtotal_sum=Sum('grandtotal'))
    )


    sales_dict = {item['date__month']: item['grandtotal_sum'] for item in sales_data}

 
    sales = [sales_dict.get(month, 0) for month in range(1, 13)]

    data = {'labels': labels, 'sales': sales}
    return JsonResponse(data)



from django.db.models import F

def graph_salesinvoice(request):
  if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
            staff_id = request.session['staff_id']
        else:
            return redirect('/')
  staff = staff_details.objects.get(id=staff_id)
  # company_instance = staff.company
  Company = company.objects.get(id=staff.company.id)
  user = Company.user
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
    
  salesinvoice = SalesInvoiceItem.objects.filter(company=Company)

  years = list(range(2022, 2031))

  return render(request, 'company/graph_salesinvoice.html',{'staff':staff,'allmodules':allmodules,'years':years})


def salesinvoicehistory(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  Company = company.objects.get(id=staff.company.id)
  history= SalesInvoiceTransactionHistory.objects.filter(salesinvoice=id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  return render(request, 'company/salesinvoicehistory.html',{'staff':staff,'history':history,'allmodules':allmodules})


def salesinvoice_billtemplate(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  Company = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  history= SalesInvoiceTransactionHistory.objects.filter(salesinvoice=id)
  salesinvoice = SalesInvoice.objects.get(id=id)
  salesinvoiceitem = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice)
  dis = 0
  for itm in salesinvoiceitem:
    dis += int(itm.discount)
  itm_len = len(salesinvoiceitem)
  return render(request, 'company/salesinvoice_billtemplate.html',{'staff':staff,'allmodules':allmodules,'history':history,'salesinvoice':salesinvoice,'salesinvoiceitem':salesinvoiceitem,'dis':dis,'itm_len':itm_len})

from openpyxl import Workbook
from django.http import HttpResponse

from openpyxl import load_workbook
from django.contrib import messages
from django.utils import timezone



def importsalesinvoice_excel(request):
    if request.method == 'POST' and request.FILES['billfile'] and request.FILES['prdfile']:
        if 'staff_id' in request.session:
            if request.session.has_key('staff_id'):
                staff_id = request.session['staff_id']
            else:
                return redirect('/')
        staff = staff_details.objects.get(id=staff_id)
        cmp = company.objects.get(id=staff.company.id)
        totval = int(SalesInvoice.objects.filter(company=cmp).last().invoice_no) + 1

        excel_bill = request.FILES['billfile']
        excel_b = load_workbook(excel_bill)
        eb = excel_b['Sheet1']
        excel_prd = request.FILES['prdfile']
        excel_p = load_workbook(excel_prd)
        ep = excel_p['Sheet1']

        for row_number1 in range(2, eb.max_row + 1):
            billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
            part = party.objects.get(party_name=billsheet[0], email=billsheet[1], company=cmp)
            SalesInvoice.objects.create(party=part,
                                        date=billsheet[2],
                                        state_of_supply=billsheet[3],
                                        invoice_no=totval,
                                        company=cmp, staff=staff)

            invoice = SalesInvoice.objects.last()
            if billsheet[4] == 'Cheque':
                invoice.paymenttype = 'Cheque'
                invoice.cheque = billsheet[5]
            elif billsheet[4] == 'UPI':
                invoice.paymenttype = 'UPI'
                invoice.upi = billsheet[5]
            else:
                if billsheet[4] != 'Cash':
                    bank = BankModel.objects.get(bank_name=billsheet[4], company=cmp)
                    invoice.paymenttype = bank
                else:
                    invoice.paymenttype = 'Cash'
            invoice.save()

            SalesInvoice.objects.filter(company=cmp)
            totval += 1
            subtotal = 0
            total_taxamount = 0
            for row_number2 in range(2, ep.max_row + 1):
                prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
                if prdsheet[0] == row_number1:
                    itm = ItemModel.objects.get(item_name=prdsheet[1], item_hsn=prdsheet[2], company=cmp)
                    total = int(prdsheet[3]) * int(itm.item_sale_price) - int(prdsheet[4])
                    SalesInvoiceItem.objects.create(salesinvoice=invoice,
                                                    company=cmp,
                                                    item=itm,
                                                    staff=staff,
                                                    quantity=prdsheet[3],
                                                    discount=prdsheet[4],
                                                    tax=prdsheet[5],
                                                    totalamount=total)
                   
                    tax=int(prdsheet[5])

                    subtotal += total
                    tamount = total * (tax / 100)
                    total_taxamount += tamount

                    if billsheet[3] == 'state':
                        gst = round((total_taxamount / 2), 2)
                        invoice.sgst = gst
                        invoice.cgst = gst
                        invoice.igst = 0
                    else:
                        gst = round(total_taxamount, 2)
                        invoice.igst = gst
                        invoice.cgst = 0
                        invoice.sgst = 0

            gtotal = subtotal + total_taxamount + float(billsheet[6])
            balance = gtotal - float(billsheet[7])
            gtotal = round(gtotal, 2)
            balance = round(balance, 2)

            invoice.subtotal = round(subtotal, 2)
            invoice.total_taxamount = round(total_taxamount, 2)
            invoice.adjustment = round(billsheet[6], 2)
            invoice.grandtotal = gtotal
            invoice.paidoff = round(billsheet[7], 2)
            invoice.totalbalance = balance
            invoice.save()

        SalesInvoiceTransactionHistory.objects.create(salesinvoice=invoice, staff=invoice.staff, company=invoice.company,
                                                      action='Created', done_by_name=invoice.staff.first_name)

        return JsonResponse({'message': 'File uploaded successfully!'})
    else:
        return JsonResponse({'message': 'File upload Failed!'})

    return render(request, 'company/view_salesinvoice.html')    
#End

#---Haripriya--
def view_purchasedebit(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  print("hello")
  print(staff)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  pdebt = purchasedebit.objects.filter(company=cmp)

  if not pdebt:
    context = {'staff':staff, 'allmodules':allmodules}
    return render(request,'company/emptydebit.html',context)
  
  context = {'staff':staff,'allmodules':allmodules,'pdebt':pdebt}
  return render(request,'company/purchase_return_dr.html',context)


def add_debitnote(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  print("hii")
  print(staff)
  cmp = company.objects.get(id=staff.company.id)
  Party=party.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  item=ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)
  bank=BankModel.objects.filter(company=cmp,user=cmp.user)
  debt_count = purchasedebit.objects.filter(company=cmp).order_by('-pdebitid').first()
  
  if debt_count:
    next_count = int(debt_count.reference_number) + 1
  else:
    next_count=1

  return render(request,'company/adddebitnotes.html',{'staff':staff,'allmodules':allmodules,'Party':Party,'item':item,'count':next_count,'tod':tod,'item_units':item_units,'bank':bank,'cmp':cmp})

def create_debitnotes(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  partys=party.objects.get(id=request.POST.get('customername'))
  if request.method == 'POST': 
    
    pdebt = purchasedebit(party=partys,
                      pdebitid=request.POST.get('pdebitid'),
                      debitdate=request.POST.get('debitdate'),
                      supply=request.POST.get('placosupply'),
                      payment_type=request.POST.get("method"),
                      cheque_no=request.POST.get("cheque_id"),
                      upi_no=request.POST.get("upi_id"),
                      billno=request.POST.get("bill_no"),
                      billdate=request.POST.get("billdate"), 
                      reference_number=request.POST.get("pdebitid"),
                      paid_amount = request.POST.get("advance"),
                      balance_amount = request.POST.get("balance"),
                      subtotal=float(request.POST.get('subtotal')),
                      igst = request.POST.get('igst'),
                      cgst = request.POST.get('cgst'),
                      sgst = request.POST.get('sgst'),
                      adjustment = request.POST.get("adj"),
                      taxamount = request.POST.get("taxamount"),
                      grandtotal=request.POST.get('grandtotal'),
                      company=cmp,staff=staff)
    pdebt.save()

    print(pdebt)
          
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    pdebitid = purchasedebit.objects.get(pdebitid =pdebt.pdebitid,company=cmp)

    if len(product)==len(qty)==len(discount)==len(total):
        mapped=zip(product,qty,discount,total)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],pdebit=pdebitid,company=cmp)

    purchasedebit.objects.filter(company=cmp).update(tot_debt_no=F('tot_debt_no') + 1)
          
    pdebt.tot_debt_no = pdebt.pdebitid
    pdebt.save()

    DebitnoteTransactionHistory.objects.create(debitnote=pdebt,staff=staff,company=cmp,action='Created')

    if 'Next' in request.POST:
      return redirect('add_debitnote')
    
    if "Save" in request.POST:
      return redirect('view_purchasedebit')
    
  else:
    return render(request,'company/adddebitnotes.html')


def edit_debitnote(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  partys = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  debtitem = purchasedebit1.objects.filter(pdebit=id,company=cmp)
  
  if pdebt.payment_type != 'Cash' and pdebt.payment_type != 'Cheque' and pdebt.payment_type != 'UPI':
    bankno = BankModel.objects.get(id= pdebt.payment_type,company=cmp,user=cmp.user)
  else:
    bankno = 0
  

  ddate = pdebt.debitdate.strftime("%Y-%m-%d")
  context = {'staff':staff,  'allmodules':allmodules, 'pdebt':pdebt, 'debtitem':debtitem, 'partys':partys, 'item':item, 'item_units':item_units, 'ddate':ddate,'bank':bank,'bankno':bankno,'tod':tod}
  return render(request,'company/debitnoteedit.html',context)

def update_debitnote(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    partys = party.objects.get(id=request.POST.get('customername'))
    pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
    pdebt.party = partys
    pdebt.debitdate = request.POST.get('debitdate')
    pdebt.billno = request.POST.get('bill_no')
    pdebt.billdate = request.POST.get('billdate')
    pdebt.supply  = request.POST.get('placosupply')
    pdebt.subtotal =float(request.POST.get('subtotal'))
    pdebt.grandtotal = request.POST.get('grandtotal')
    pdebt.igst = request.POST.get('igst')
    pdebt.cgst = request.POST.get('cgst')
    pdebt.sgst = request.POST.get('sgst')
    pdebt.taxamount = request.POST.get("taxamount")
    pdebt.adjustment = request.POST.get("adj")
    pdebt.payment_type = request.POST.get("method")
    pdebt.cheque_no = request.POST.get("cheque_id")
    pdebt.upi_no = request.POST.get("upi_id")
    pdebt.paid_amount = request.POST.get("advance")
    pdebt.balance_amount = request.POST.get("balance")

    pdebt.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    purchasedebit1.objects.filter(pdebit=pdebt,company=cmp).delete()
    if len(total)==len(discount)==len(qty):
      mapped=zip(product,qty,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],pdebit=pdebt,company=cmp)

    DebitnoteTransactionHistory.objects.create(debitnote=pdebt,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasedebit')

  return redirect('view_purchasedebit')

def history_debitnote(request,id):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)  
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  hsty= DebitnoteTransactionHistory.objects.filter(debitnote=id,company=cmp)
  context = {'staff':staff,'allmodules':allmodules,'hsty':hsty,'id':id}
  return render(request,'company/debitnotehistory.html',context)

def debthistory(request):
  pid = request.POST['id']
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pdebt = purchasedebit.objects.get(pdebitid=pid,company=cmp)
  hsty = DebitnoteTransactionHistory.objects.filter(debitnote=pdebt,company=cmp).last()
  name = hsty.staff.first_name + ' ' + hsty.staff.last_name 
  action = hsty.action
  return JsonResponse({'name':name,'action':action,'pid':pid})

def delete_debit(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  pdebt = purchasedebit.objects.get(pdebitid=id)
  purchasedebit1.objects.filter(pdebit=pdebt,company=cmp).delete()
  pdebt.delete()
  return redirect('view_purchasedebit')
  


def cust_dropdown1(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = party.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  party_list = []
  for p in part:
    id_list.append(p.id)
    party_list.append(p.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })


def savecustomer1(request):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    party_name = request.POST['name']
    email = request.POST['email']
    contact = request.POST['mobile']
    state = request.POST['splystate']
    address = request.POST['baddress']
    gst_type = request.POST['gsttype']
    gst_no = request.POST['gstin']
    current_date = request.POST['partydate']
    openingbalance = request.POST.get('openbalance')
    payment = request.POST.get('paytype')
    creditlimit = request.POST.get('credit_limit')
    End_date = request.POST.get('enddate', None)
    additionalfield1 = request.POST['add1']
    additionalfield2 = request.POST['add2']
    additionalfield3 = request.POST['add3']

    part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,
                  payment=payment,creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,
                  additionalfield3=additionalfield3,company=cmp,user=cmp.user)
    part.save() 
  return JsonResponse({'success': True})



def details_debitnote(request,id):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.get(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  pitm = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'allmodules':allmodules,'pdebt':pdebt,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'company/debitnotedetails.html',context)


def custdata1(request):
  cid = request.POST['id']
  part = party.objects.get(id=cid)
  # email = part.email
  phno = part.contact
  address = part.address
  pay = part.payment
  bal = part.openingbalance
  return JsonResponse({ 'phno':phno, 'address':address, 'pay':pay, 'bal':bal})


def purchasebilldata(request):
    try:
        party_name = request.POST['id']
        party_instance = party.objects.get(id=party_name)

        # Initialize lists to store multiple bill numbers and dates
        bill_numbers = []
        bill_dates = []

        try:
            # Retrieve all PurchaseBill instances for the party
            bill_instances = PurchaseBill.objects.filter(party=party_instance)

            # Loop through each PurchaseBill instance and collect bill numbers and dates
            for bill_instance in bill_instances:
                bill_numbers.append(bill_instance.billno)
                bill_dates.append(bill_instance.billdate)

        except PurchaseBill.DoesNotExist:
            pass

        # Return a JSON response with the list of bill numbers and dates
        if not bill_numbers and not bill_dates:
            return JsonResponse({'bill_numbers': ['nobill'], 'bill_dates': ['nodate']})

        return JsonResponse({'bill_numbers': bill_numbers, 'bill_dates': bill_dates})

    except KeyError:
        return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

    except party.DoesNotExist:
        return JsonResponse({'error': 'Party not found.'})




def import_debitnote(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(purchasedebit.objects.filter(company=cmp).last().tot_debt_no)

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      debitsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=debitsheet[0],email=debitsheet[1],company=cmp)
      purchasedebit.objects.create(party=part,pdebitid = totval,
                                  debitdate=debitsheet[2],
                                  supply =debitsheet[3],
                                  tot_debt_no = totval,
                                  company=cmp,staff=staff)
      
      pdebt = purchasedebit.objects.last()
      if debitsheet[4] == 'Cheque':
        pdebt.payment_type = 'Cheque'
        pdebt.cheque_no = debitsheet[5]
      elif debitsheet[4] == 'UPI':
        pdebt.upi_no = debitsheet[5]
      else:
        if debitsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=debitsheet[4],company=cmp)
          pdebt.payment_type = bank
        else:
          pdebt.payment_type = 'Cash'
      pdebt.save()

      purchasedebit.objects.filter(company=cmp).update(tot_debt_no=totval )
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=prdsheet[2],company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          
          purchasedebit1.objects.create(pdebit=pdebt,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

       
          if debitsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if debitsheet[3]=='State':
            gst = round((taxamount/2),2)
            pdebt.sgst=gst
            pdebt.cgst=gst
            pdebt.igst=0

          else:
            gst=round(taxamount,2)
            pdebt.igst=gst
            pdebt.cgst=0
            pdebt.sgst=0

      gtotal = subtotal + taxamount + float(debitsheet[6])
      balance = gtotal- float(debitsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pdebt.subtotal=round(subtotal,2)
      pbpdebtill.taxamount=round(taxamount,2)
      pdebt.adjustment=round(debitsheet[6],2)
      pdebt.grandtotal=gtotal
      pdebt.paid_amount=round(debitsheet[7],2)
      pdebt.balance_amount=balance
      pdebt.save()

      DebitnoteTransactionHistory.objects.create(debitnote=pdebt,staff=pdebt.staff,company=pdebt.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})


def saveitem1(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})

def item_dropdowns(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  product_list = []
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
  return JsonResponse({'id_list':id_list, 'product_list':product_list})



def itemdetail(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})


def bankdata1(request):
  bid = request.POST['id']
  bank = BankModel.objects.get(id=bid) 
  bank_no = bank.account_num
  return JsonResponse({'bank_no':bank_no})


# ========================================   Haripriya b Nair (END) ======================================================    


def sharedebitToEmail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)

                sid = request.session.get('staff_id')
                staff =  staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
               
                pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
                pitm = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)
                        
                context = {'pdebt':pdebt, 'cmp':cmp,'pitm':pitm}
                template_path = 'company/debitnote_file_mail.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'DEBIT NOTE - {pdebt.pdebitid}.pdf'
                subject = f"DEBIT NOTE - {pdebt.pdebitid}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached DEBIT NOTE - File-{pdebt.pdebitid}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
                return redirect(details_debitnote,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(details_debitnote, id)
            
            
def distributor_notification(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  data= modules_list.objects.filter(update_action=1,status='Pending',company__reg_action='distributor',company__Distributors=distributor.id)
  return render(request,'distributor/distributor_notification.html',{'distributor':distributor,'data':data})

def distributor_module_updation(request,mid):
  distributor =  Distributors_details.objects.get(user = request.user)
  data= modules_list.objects.get(id=mid)
  allmodules= modules_list.objects.get(company=data.company,status='Pending')
  old_modules = modules_list.objects.get(company=data.company,status='New')
  return render(request,'distributor/distributor_module_updation.html',{'distributor':distributor,'data':data,'allmodules':allmodules,'old_modules':old_modules})

def distributor_module_updation_ok(request,mid):
  
  old=modules_list.objects.get(company=mid,status='New')
  old.delete()

  data=modules_list.objects.get(company=mid,status='Pending')  
  data.status='New'
  data.save()
  data1=modules_list.objects.filter(company=mid).update(update_action=0)
  return redirect('distributor_notification')

def expense(request):

  staff_id = request.session['staff_id']
  staff = staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  
  expenses = Expense.objects.filter(staff_id__company=staff.company).order_by('-id')
  allcat = Expense_Category.objects.filter(staff__company=staff.company).order_by('-id')


  category_totals = defaultdict(float)
  category_balances = defaultdict(float)


  for expense in expenses:
    category_totals[expense.expense_category_id.id] += expense.total
    category_balances[expense.expense_category_id.id] += expense.balance

  for category in allcat:
    category.grant = category_totals.get(category.id, 0)
    category.balance = category_balances.get(category.id, 0)
    category.save() 



  first=allcat[:1]
  print(first)
  
 
  ex = Expense.objects.filter(staff_id__company=staff.company)


  context={'staff':staff,
           'allmodules':allmodules,
           'expenses':expenses,
           'first':first,
           'allcat':allcat,
           'ex':ex
           }
  return render(request,'company/expense.html',context)

def newexpenses(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  parties=party.objects.filter(company=staff.company)
  Category=Expense_Category.objects.filter(staff__company=staff.company)
  bank = BankModel.objects.filter(company=staff.company)
  ex=Expense.objects.filter(staff_id__company=staff.company)
  if ex:
    en = int(Expense.objects.filter(staff_id__company=staff.company).last().EXP_NO)
    enpno=en+1 
    
  else:
     enpno=1  

  context={'staff':staff,
           'allmodules':allmodules,
           'parties':parties,
           'Category':Category,
           'bank':bank,
           'enpno':enpno
           }
  return render(request,'company/newexpenses.html',context)

    
def partydata(request):
    party_id = request.GET.get('id')
    p = party.objects.get(id=party_id)
    data7 = {'email': p.email,'openingbalance':p.openingbalance,'address':p.address,'contact':p.contact,'payment':p.payment}
    return JsonResponse(data7)

def add_party_in_expense(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    party_name = request.POST['name']
    email = request.POST['email']
    contact = request.POST['mobile']
    state = request.POST['splystate']
    address = request.POST['baddress']
    gst_type = request.POST['gsttype']
    gst_no = request.POST['gstin']
    current_date = request.POST['partydate']
    openingbalance = request.POST.get('openbalance')
    payment = request.POST.get('paytype')
    creditlimit = request.POST.get('credit_limit')
    End_date = request.POST.get('enddate', None)
    additionalfield1 = request.POST['add1']
    additionalfield2 = request.POST['add2']
    additionalfield3 = request.POST['add3']

    part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,
                payment=payment,creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,
                additionalfield3=additionalfield3,company=cmp,user=cmp.user)
    part.save() 
    options = {}
    option_objects =party.objects.filter(company=staff.company.id)
    for option in option_objects:
      options[option.id] = [option.party_name]
    return JsonResponse(options) 
  else:
    return JsonResponse({'error': 'Invalid request'},status=400)
  


def create_expense_category(request):
  if request.method=='POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    
    category_name = request.POST.get('item_unit_name')
    category_data = Expense_Category(staff = staff,expense_category=category_name)
    category_data.save()
  return JsonResponse({'message':'asdasd'})


def create_expense(request):
  if request.method=='POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    

    customer = request.POST['party']
    if customer != '':
      pid= party.objects.get(id=customer)


    ex_category = request.POST['ex_category']  
    cid=Expense_Category.objects.get(id=ex_category)
    ex_date = request.POST['ex_date']  
    payment_type = request.POST['pay_method']  
    subtotal = request.POST['sub_total']  
    cgst = request.POST['cgst']  
    sgst = request.POST['sgst']  
    taxamount = request.POST['tax_amount']  
    adjustment = request.POST['adjustment']  
    total = request.POST['total']  
    paid = request.POST['paid']  
    balance = request.POST['balance']
    ex=Expense.objects.filter(staff_id=staff)
    if ex:
      en = int(Expense.objects.filter(staff_id__company=staff.company).last().EXP_NO)
      expno=en+1 
    
    else:
      expno=1  
      

    print(cgst)

    if customer == '' and cgst != 0 and sgst != 0:
      
      
      data = Expense(staff_id = staff,expense_category_id=cid,Sub_total=subtotal,
                   cgst=cgst,sgst=sgst,tax_amount=taxamount,adjustment=adjustment,total=total,paid=paid,balance=balance,payment_type=payment_type,expense_date=ex_date)
      data.save()
    elif customer != '' and cgst == 0 and sgst == 0:

      data = Expense(staff_id = staff,party_id = pid,expense_category_id=cid,Sub_total=subtotal,
                   adjustment=adjustment,total=total,paid=paid,balance=balance,payment_type=payment_type,expense_date=ex_date)
      data.save() 

    elif customer == '' and cgst == 0 and sgst == 0:   
      data = Expense(staff_id = staff,expense_category_id=cid,Sub_total=subtotal,
                   adjustment=adjustment,total=total,paid=paid,balance=balance,payment_type=payment_type,expense_date=ex_date)
      data.save() 

    else :
      data = Expense(staff_id = staff,party_id = pid,expense_category_id=cid,Sub_total=subtotal,
                   cgst=cgst,sgst=sgst,tax_amount=taxamount,adjustment=adjustment,total=total,paid=paid,balance=balance,payment_type=payment_type,expense_date=ex_date)
      data.save()

    data1=Expense.objects.filter(id=data.id).update(EXP_NO=expno) 


    tax = tuple(request.POST.getlist("tax[]"))
    if tax == '':
      tax_value = 0.0
      tax = float(tax_value)
    dis = tuple(request.POST.getlist("dis[]"))
    amount = tuple(request.POST.getlist("amount[]"))

    print("Tax:", tax)  
    print("Dis:", dis)
    print("Amount:", amount)
    print("All POST data:", request.POST)  

    if cgst != 0 and sgst != 0:
      if len(dis) == len(tax) == len(amount):
        mapped = zip(dis, tax, amount)
        mapped = list(mapped)
        print("Mapped:", mapped)

        for ele in mapped:
          Expense_list.objects.create(
              expense_id=data,
              discription=ele[0],
              tax=ele[1],
              amount=ele[2] 
          )

        
    else:
      if len(dis) == len(amount):
        mapped = zip(dis, amount)
        mapped = list(mapped)
        print("Mapped:", mapped)

        for ele in mapped:
          Expense_list.objects.create(
              expense_id=data,
              discription=ele[0],
              tax = 0.0,
              amount=ele[1] 
          )
          
    current_datetime = timezone.now()
    date =  current_datetime.date()
    ExpenseHistory.objects.create(
      staff=staff,
      expense=data,
      date=date,
      action = "Create",

    )    
    return redirect('expense')

  return redirect('add_party_in_expense')


def view_expense(request,eid):
  
  staff_id = request.session['staff_id']
  staff = staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  
  expenses = Expense.objects.filter(staff_id__company=staff.company).order_by('-id')
  allcat = Expense_Category.objects.filter(staff__company=staff.company).order_by('-id')
  


  category_totals = defaultdict(float)
  category_balances = defaultdict(float)


  for expense in expenses:
    category_totals[expense.expense_category_id.id] += expense.total
    category_balances[expense.expense_category_id.id] += expense.balance

  for category in allcat:
    
      category.grant = category_totals.get(category.id, 0)
      category.balance = category_balances.get(category.id, 0)
      category.save()


  first= Expense_Category.objects.filter(id=eid)
  for f in first:
    
      f.grant = category_totals.get(f.id, 0)
      f.balance = category_balances.get(f.id, 0)
      f.save()
      print(f)
 


  
  
 
  ex = Expense.objects.filter(staff_id__company=staff.company)


  context={'staff':staff,
           'allmodules':allmodules,
           'expenses':expenses,
           'first':first,
           'allcat':allcat,
           'ex':ex,
           
           }
  return render(request,'company/expense.html',context)

def expense_details(request,eid):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  expnse=Expense.objects.get(id=eid)
  elist = Expense_list.objects.filter(expense_id=eid)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  context={
    'staff':staff,
    'allmodules':allmodules,
    'expnse':expnse,
    'elist':elist

  }
  return render(request,'company/expense_details.html',context)


def edit_expense(request,eid):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  expense = Expense.objects.get(id=eid)
  parties = party.objects.all()
  Category=Expense_Category.objects.filter(staff__company=staff.company)
  expense_details=Expense_list.objects.filter(expense_id=eid)
  bank = BankModel.objects.filter(company=staff.company)

  context={
    'staff':staff,
    'allmodules':allmodules,
    'expense':expense,
    'parties':parties,
    'Category':Category,
    'expense_details':expense_details,
    'bank':bank,

    

  }

  return render(request,'company/edit_expense.html',context)

def edit_expense_action(request,eid):
  if request.method=='POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    expense = Expense.objects.get(id=eid)

    customer = request.POST['party']
    if customer != '':
      pid= party.objects.get(id=customer)


    ex_category = request.POST['ex_category']  
    cid=Expense_Category.objects.get(id=ex_category)
    ex_date = request.POST['ex_date']  
    payment_type = request.POST['pay_method']  
    subtotal = request.POST['sub_total']  
    cgst = request.POST['cgst']  
    sgst = request.POST['sgst']  
    taxamount = request.POST['tax_amount']  
    adjustment = request.POST['adjustment']  
    total = request.POST['total']  
    paid = request.POST['paid']  
    balance = request.POST['balance']

    expense.staff_id =  staff

    if customer != '':
      expense.party_id =  pid

    expense.expense_category_id = cid 

    expense.Sub_total = subtotal

    expense.payment_type = payment_type

    if cgst != 0:
      expense.cgst = cgst
      expense.sgst = sgst
      expense.tax_amount = taxamount

    expense.adjustment = adjustment  
    expense.total = total
    expense.paid = paid
    expense.balance = balance

    expense.save()

    data1=Expense.objects.filter(id=eid).update(action=1) 

    l=Expense_list.objects.filter(expense_id=eid)
    for i in l:
      i.delete()

    tax = tuple(request.POST.getlist("tax[]"))
    if tax == '':
      tax_value = 0.0
      tax_value1 = float(tax_value)
    dis = tuple(request.POST.getlist("dis[]"))
    amount = tuple(request.POST.getlist("amount[]"))

    print("Tax:", tax)  
    print("Dis:", dis)
    print("Amount:", amount)
    print("All POST data:", request.POST)  

    if cgst != 0 and sgst != 0:
      if len(dis) == len(tax) == len(amount):
        mapped = zip(dis, tax, amount)
        mapped = list(mapped)
        print("Mapped:", mapped)

        for ele in mapped:
          Expense_list.objects.create(
              expense_id=expense,
              discription=ele[0],
              tax=ele[1],
              amount=ele[2] 
          )

        return redirect('expense')
    else:
      if len(dis) == len(amount):
        mapped = zip(dis, amount)
        mapped = list(mapped)
        print("Mapped:", mapped)

        for ele in mapped:
          Expense_list.objects.create(
              expense_id=expense,
              discription=ele[0],
              tax = 0.0,
              amount=ele[1] 
          )
          

        return redirect('expense')
      
def delete_expense(request,eid):  
  data=Expense.objects.get(id=eid)
  data.delete()
  l=Expense_list.objects.filter(expense_id=eid)
  for i in l:
    i.delete()
  return redirect('expense')    



def import_expense(request):
  if request.method == 'POST' and 'billfile' in request.FILES and 'prdfile' in request.FILES:

    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(Expense.objects.last().EXP_NO)

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']

    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      print(billsheet)
      part = party.objects.get(id=billsheet[0],company=cmp)
      cat = Expense_Category.objects.get(id=billsheet[1])
      Expense.objects.create(party_id=part, 
                                  expense_category_id=cat,
                                  expense_date=billsheet[2],
                                  cgst =billsheet[3],
                                  sgst =billsheet[4],
                                  tax_amount =billsheet[5],
                                  payment_type =billsheet[6],
                                  Sub_total =billsheet[7],
                                  adjustment =billsheet[8],
                                  total =billsheet[9],
                                  paid =billsheet[10],
                                  balance =billsheet[11],
                                  EXP_NO = totval+1,
                                  staff_id=staff)
      Exp = Expense.objects.last()
      
      
  
    
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        
       
          
        Expense_list.objects.create(expense_id=Exp,
                                          
                                          discription=prdsheet[0],
                                          tax=prdsheet[1],
                                          amount=prdsheet[2]
                                          )
        

      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})
    
    
def create_purchaseorder(request):
  if request.method == 'POST': 
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    part = party.objects.get(id=request.POST.get('customername'))
    pord = PurchaseOrder(party=part, 
                          orderno=request.POST.get('ord_no'),
                          orderdate=request.POST.get('orderdate'),
                          duedate=request.POST.get('duedate'),
                          supplyplace =request.POST.get('placosupply'),
                          pay_method=request.POST.get("method"),
                          cheque_no=request.POST.get("cheque_id"),
                          upi_no=request.POST.get("upi_id"),
                          advance = request.POST.get("advance"),
                          balance = request.POST.get("balance"),
                          subtotal=float(request.POST.get('subtotal')),
                          igst = request.POST.get('igst'),
                          cgst = request.POST.get('cgst'),
                          sgst = request.POST.get('sgst'),
                          adjust = request.POST.get("adj"),
                          taxamount = request.POST.get("taxamount"),
                          grandtotal=request.POST.get('grandtotal'),
                          company=cmp,staff=staff)
    pord.save()
        
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    ordno = PurchaseOrder.objects.get(orderno=pord.orderno,company=cmp)

    if len(product)==len(qty)==len(discount)==len(total):
        mapped=zip(product,qty,discount,total)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          PurchaseOrderItem.objects.create(product=itm,qty=ele[1],discount=ele[2],total=ele[3],purchaseorder=ordno,company=cmp)

    PurchaseOrder.objects.filter(company=cmp).update(tot_ord_no=F('tot_ord_no') + 1)

    pord.tot_ord_no = pord.orderno
    pord.save()

    PurchaseOrderTransactionHistory.objects.create(purchaseorder=pord,staff=staff,company=cmp,action='Created')

    if 'Next' in request.POST:
      return redirect('add_purchaseorder')
    
    if "Save" in request.POST:
      return redirect('view_purchaseorder')
    
  else:
    return render(request,'company/purchaseorderadd.html')


def edit_purchaseorder(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")

  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pord = PurchaseOrder.objects.get(id=id,company=cmp)
  ordprd = PurchaseOrderItem.objects.filter(purchaseorder=pord,company=cmp)

  if pord.pay_method != 'Cash' and pord.pay_method != 'Cheque' and pord.pay_method != 'UPI':
    bankno = BankModel.objects.get(id = pord.pay_method,company=cmp,user=cmp.user)
  else:
    bankno = 0

  bdate = pord.orderdate.strftime("%Y-%m-%d")
  ddate = pord.duedate.strftime("%Y-%m-%d")
  context = {'staff':staff, 'allmodules':allmodules, 'pord':pord, 'ordprd':ordprd, 'cust':cust, 'item':item, 
             'tod':tod,'item_units':item_units, 'bdate':bdate, 'ddate':ddate, 'bank':bank, 'bankno':bankno}
  return render(request,'company/purchaseorderedit.html',context)


def update_purchaseorder(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = party.objects.get(id=request.POST.get('customername'))
    pord = PurchaseOrder.objects.get(id=id,company=cmp)
    pord.party = part
    pord.orderdate = request.POST.get('orderdate')
    pord.duedate = request.POST.get('duedate')
    pord.supplyplace  = request.POST.get('placosupply')
    pord.subtotal =float(request.POST.get('subtotal'))
    pord.grandtotal = request.POST.get('grandtotal')
    pord.igst = request.POST.get('igst')
    pord.cgst = request.POST.get('cgst')
    pord.sgst = request.POST.get('sgst')
    pord.taxamount = request.POST.get("taxamount")
    pord.adjust = request.POST.get("adj")
    pord.pay_method = request.POST.get("method")
    pord.cheque_no = request.POST.get("cheque_id")
    pord.upi_no = request.POST.get("upi_id")
    pord.advance = request.POST.get("advance")
    pord.balance = request.POST.get("balance")

    pord.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    PurchaseOrderItem.objects.filter(purchaseorder=pord,company=cmp).delete()
    if len(total)==len(discount)==len(qty):
      mapped=zip(product,qty,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        PurchaseOrderItem.objects.create(product=itm,qty=ele[1],discount=ele[2],total=ele[3],purchaseorder=pord,company=cmp)

    PurchaseOrderTransactionHistory.objects.create(purchaseorder=pord,staff=staff,company=cmp,action='Updated')
    return redirect('view_purchaseorder')

  return redirect('view_purchaseorder')


def details_purchaseorder(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules = modules_list.objects.get(company=staff.company,status='New')
  pord = PurchaseOrder.objects.get(id=id,company=cmp)
  oitm = PurchaseOrderItem.objects.filter(purchaseorder=pord,company=cmp)
  dis = 0
  for itm in oitm:
    dis += int(itm.discount)
  itm_len = len(oitm)

  context={'staff':staff,'allmodules':allmodules,'pord':pord,'oitm':oitm,'itm_len':itm_len,'dis':dis}
  return render(request,'company/purchaseorderdetails.html',context)


def delete_purchaseorder(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  pord = PurchaseOrder.objects.get(id=id,company=cmp)
  PurchaseOrderItem.objects.filter(purchaseorder=pord,company=cmp).delete()
  pord.delete()
  return redirect('view_purchaseorder')


def orderhistory(request):
  pid = request.POST.get('id')
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  pord = PurchaseOrder.objects.get(orderno=pid,company=cmp)
  hst = PurchaseOrderTransactionHistory.objects.filter(purchaseorder=pord,company=cmp).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid})

def convert_to_bill(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")

  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cust = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pord = PurchaseOrder.objects.get(id=id,company=cmp)
  ordprd = PurchaseOrderItem.objects.filter(purchaseorder=pord,company=cmp)

  if pord.pay_method != 'Cash' and pord.pay_method != 'Cheque' and pord.pay_method != 'UPI':
    bankno = BankModel.objects.get(id = pord.pay_method,company=cmp,user=cmp.user)
  else:
    bankno = 0

  last_bill = PurchaseBill.objects.filter(company=cmp).last()
  if last_bill:
    bill_no = last_bill.tot_bill_no + 1 
  else:
    bill_no = 1

  bdate = pord.orderdate.strftime("%Y-%m-%d")
  ddate = pord.duedate.strftime("%Y-%m-%d")
  context = {'staff':staff, 'allmodules':allmodules, 'pord':pord, 'ordprd':ordprd, 'cust':cust, 'item':item, 'bill_no':bill_no,
             'tod':tod,'item_units':item_units, 'bdate':bdate, 'ddate':ddate, 'bank':bank, 'bankno':bankno}
  return render(request,'company/ordertobill.html',context)


def import_purchase_order(request):
  if request.method == 'POST' and request.FILES['ordfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(PurchaseOrder.objects.filter(company=cmp).last().tot_ord_no) + 1

    excel_order = request.FILES['ordfile']
    excel_o = load_workbook(excel_order)
    eo = excel_o['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eo.max_row + 1):
      ordersheet = [eo.cell(row=row_number1, column=col_num).value for col_num in range(1, eo.max_column + 1)]
      part = party.objects.get(party_name=ordersheet[0],email=ordersheet[1],company=cmp)
      PurchaseOrder.objects.create(party=part,orderno=totval,
                                  orderdate=ordersheet[2],
                                  duedate=ordersheet[3],
                                  supplyplace =ordersheet[4],
                                  tot_ord_no = totval,
                                  company=cmp,staff=staff)
      
      pord = PurchaseOrder.objects.last()
      if ordersheet[5] == 'Cheque':
        pord.pay_method = 'Cheque'
        pord.cheque_no = ordersheet[5]
      elif ordersheet[5] == 'UPI':
        pord.pay_method = 'UPI'
        pord.upi_no = ordersheet[5]
      else:
        if ordersheet[5] != 'Cash':
          bank = BankModel.objects.get(bank_name=ordersheet[5],company=cmp)
          pord.pay_method = bank
        else:
          pord.pay_method = 'Cash'
      pord.save()

      PurchaseOrder.objects.all().update(tot_ord_no=totval + 1)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=int(prdsheet[2]),company=cmp)
          if ordersheet[3] =='State':
            taxval =itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval =itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount 
          PurchaseOrderItem.objects.create(purchaseorder=pord,
                                          company=cmp,
                                          product=itm,
                                          qty=prdsheet[3],
                                          discount=prdsheet[4],
                                          total=total)
          if ordersheet[4]=='State':
            gst = round((taxamount/2),2)
            pord.sgst=gst
            pord.cgst=gst
            pord.igst=0

          else:
            gst=round(taxamount,2)
            pord.igst=gst
            pord.cgst=0
            pord.sgst=0

      gtotal = subtotal + taxamount + float(ordersheet[7])
      balance = gtotal- float(ordersheet[8])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pord.subtotal=round(subtotal,2)
      pord.taxamount=round(taxamount,2)
      pord.adjust=round(ordersheet[7],2)
      pord.grandtotal=gtotal
      pord.advance=round(ordersheet[8],2)
      pord.balance=balance
      pord.save()

      PurchaseOrderTransactionHistory.objects.create(purchaseorder=pord,staff=pord.staff,company=pord.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})


def history_purchaseorder(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pord = PurchaseOrder.objects.get(id=id,company=cmp)
  hst= PurchaseOrderTransactionHistory.objects.filter(purchaseorder=pord,company=cmp)

  context = {'staff':staff,'allmodules':allmodules,'hst':hst,'pord':pord}
  return render(request,'company/purchaseorderhistory.html',context)


def order_to_bill(request,id):
  if request.method == 'POST': 
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    part = party.objects.get(id=request.POST.get('customername'))
    pord = PurchaseOrder.objects.get(id=id,company=cmp)
    pbill = PurchaseBill(party=part, 
                          billno=request.POST.get('bill_no'),
                          billdate=request.POST.get('billdate'),
                          duedate = request.POST.get('billdate'),
                          supplyplace =request.POST.get('placosupply'),
                          pay_method=request.POST.get("method"),
                          cheque_no=request.POST.get("cheque_id"),
                          upi_no=request.POST.get("upi_id"),
                          advance = request.POST.get("advance"),
                          balance = request.POST.get("balance"),
                          subtotal=float(request.POST.get('subtotal')),
                          igst = request.POST.get('igst'),
                          cgst = request.POST.get('cgst'),
                          sgst = request.POST.get('sgst'),
                          adjust = request.POST.get("adj"),
                          taxamount = request.POST.get("taxamount"),
                          grandtotal=request.POST.get('grandtotal'),
                          company=cmp,staff=staff)
    pbill.save()

    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    billno = PurchaseBill.objects.get(billno=pbill.billno,company=cmp)

    if len(product)==len(qty)==len(discount)==len(total):
        mapped=zip(product,qty,discount,total)
        mapped=list(mapped)
        print(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          PurchaseBillItem.objects.create(product=itm,qty=ele[1],discount=ele[2],total=ele[3],purchasebill=billno,company=cmp)

    PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=F('tot_bill_no') + 1)
    pbill.tot_bill_no = pbill.billno
    pbill.save()

    PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,company=cmp,staff=staff,action='Created')
    pord.convert = 1
    pord.convert_id = pbill
    pord.save()

    pord.balance = request.POST.get("balance")
    pord.save()
  
    return redirect('view_purchaseorder') 

# ===================Nasneen===========

def sale_order(request):
  
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  sale = salesorder.objects.filter(comp=staff.company)
  for i in sale:
      last_transaction = saleorder_transaction.objects.filter(sales_order=i).order_by('-id').first()
      i.last= last_transaction.action
      i.by=last_transaction.staff
      print(last_transaction.action)
      
      

  context={
    'sale':sale,'staff':staff,'allmodules':allmodules
  }
  return render(request, 'company/sale_order.html',context)

def saleorder_create(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  # cmp = staff.company
  cmp = company.objects.get(id=staff.company.id)
  par= party.objects.filter(company=staff.company)
  item = ItemModel.objects.filter(company=staff.company)
  bnk = BankModel.objects.filter(company=cmp)
  order = salesorder.next_orderno(company_id=staff.company.id)
  
  
  context={
    'party':par,'item':item,'staff':staff,'order':order,'bnk':bnk,'allmodule':allmodules
  }
  return render(request, 'company/saleorder_create.html',context)


def getparty(request):
    print("=======================")
    p_id = request.GET.get('id')
    print(p_id)
    par = party.objects.get(id=p_id)
    print(par.party_name)
    data7 = {'phone': par.contact,'balance':par.openingbalance,'payment':par.payment,'address':par.address}
    
    print(data7)
    return JsonResponse(data7)
    
def getacc(request):
    b_id = request.GET.get('id')
    print(b_id)
    par = BankModel.objects.get(id=b_id)
    data7 = {'acc': par.account_num}
    
    print(data7)
    return JsonResponse(data7)



def getproduct(request):
    p_id = request.GET.get('id')
    print(p_id)
    item = ItemModel.objects.get(id=p_id)
    data7 = {'hsn': item.item_hsn,'price':item.item_sale_price,'gst':item.item_gst,'igst':item.item_igst}
    
    print(data7)
    return JsonResponse(data7)

#  if 'staff_id' in request.session:
#     if request.session.has_key('staff_id'):
#       staff_id = request.session['staff_id']
           
#     else:
#       return redirect('/')
#   staff =  staff_details.objects.get(id=staff_id)

@login_required(login_url='login')
def create_saleorder(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  if request.method == 'POST':
    prtyid = request.POST.get('party')
    prty=party.objects.get(id=prtyid)
    # staff =  staff_details.objects.get(id=staff_id)
    cmp= staff.company
    payment = request.POST.get('paymethode')
    pos=request.POST.get('stateofsply')
    attach=request.FILES.get('attach')       
  
    print(request.POST.getlist("product[]"))
    print(request.POST.get('orderdate'))

    sale = salesorder(
      party=prty,
      partyname =prty.party_name,
      orderno=request.POST.get('orderno'),
      orderdate=request.POST.get('orderdate'),
      duedate=request.POST.get('duedate'),
      placeofsupply=pos,
      payment_method=payment,
      subtotal=request.POST.get('subtotal'),
      taxamount=request.POST.get('taxamount'),
      adjustment=request.POST.get('adj'),
      grandtotal=request.POST.get('grandtotal'),
      note=request.POST.get('note'),
      paid=request.POST.get('paid'),
      balance=request.POST.get('baldue'),
      file=attach,
      staff=staff,
      comp=cmp,
      

    )

    if payment == 'Cheque':
      sale.checkno = request.POST.get('checkno')
    elif payment == 'UPI':
      sale.UPI = request.POST.get('upiid')
    elif payment != 'Cheque' and payment != 'UPI 'and payment != 'Cash':
      sale.accno = request.POST.get('accno')
    
    if pos == 'State':
      sale.CGST=request.POST.get('cgst')
      sale.SGST=request.POST.get('sgst')
    elif pos == 'Other state':
      sale.IGST=request.POST.get('igst')

    sale.save()
    print("saved===================================")
    
    product = request.POST.getlist("product[]")
    hsn  = request.POST.getlist("hsn[]")
    qty = request.POST.getlist("qty[]")
    price = request.POST.getlist("price[]")
    tax = request.POST.getlist("tax1[]")
    discount = request.POST.getlist("discount[]")
    total = request.POST.getlist("total[]")
    taxamount=request.POST.getlist("taxamount[]")
    salesorderid=salesorder.objects.get(id =sale.id)
    print(product)
    print(len(hsn))
    print(len(qty))
    print(len(price))
   
    if len(product)==len(hsn)==len(qty) ==len(price)==len(tax)==len(discount)==len(total)==len(taxamount):
      mapped = zip(product, hsn, qty, price, tax, discount, total,taxamount)
      mapped = list(mapped)
      for ele in mapped:
        print(ele[0])
        prod=ItemModel.objects.get(id=ele[0])
        salesorderAdd = sales_item(
          product=prod,
          hsn=ele[1],
          qty=ele[2],
          price=ele[3],
          tax=ele[4],
          discount=ele[5],
          total=ele[6],
          taxamount=ele[7],
          sale_order=salesorderid,
          cmp=staff.company
            )
        salesorderAdd.save()
        print("item saved===================================")
      
      tran= saleorder_transaction(
        sales_order=salesorderid,staff=staff,company=cmp,action="Created",date=date.today()
      )
      tran.save()
    if request.POST.get('save_and_next'):
      return redirect('saleorder_create')
    elif request.POST.get('save'):
      return redirect('sale_order')
    
  return redirect('sale_order')
      
def saleorder_view(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)  
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  sale = salesorder.objects.get(id=id)
  item = sales_item.objects.filter(sale_order=sale)
  s = salesorder.objects.all()
  prty = party.objects.get(id=sale.party.id)
  
  context={
    'sale':sale,'item':item,'s':s,'prty':prty,'staff':staff,'allmodules':allmodules
  }
  return render(request, 'company/saleorder_view.html',context)

def delete_saleorder(request,id):
  sale = salesorder.objects.get(id=id)
  item = sales_item.objects.filter(sale_order=sale)
  tran = saleorder_transaction.objects.filter(sales_order=sale)
  for i in item:
    i.delete()
  for i in tran:
    i.delete()
  sale.delete()
  return redirect('sale_order')
  
def import_excel(request):
    if request.method == "POST" and request.FILES.get("file"):
      staff_id = request.session['staff_id']
      staff =  staff_details.objects.get(id=staff_id)
      print("open============================================")
      excel_file = request.FILES['file']
      if excel_file.name.endswith('.xlsx'):
        print("open1111111111111111111111")
        df = pd.read_excel(excel_file, engine='openpyxl')
        for index, row in df.iterrows():
          print(row['PARTY NAME'])
          s = salesorder(
                    partyname=row['PARTY NAME'],
                    orderno=row['NUMBER'],
                    orderdate=row['DATE'],
                    duedate=row['DUE DATE'],
                    grandtotal=row['TOTAL'],
                    balance=row['BALANCE'],
                    status=row['STATUS'],
                    action=row['ACTION'],
                    staff=staff,
                    comp=staff.company,
                    # Add other fields accordingly
                )
          s.save()
          tran= saleorder_transaction(
            sales_order=s,staff=staff,company=staff.company,action="Created",date=date.today()
            )
          tran.save()
        print("success============================================")
        return redirect('sale_order')  # Redirect to a success page
      print("end===========================")
    return redirect('sale_order')
  
  
def add_party(request):
  if request.method == 'POST':
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        
        party_name = request.POST['partyname']
        gst_no = request.POST['gstno']
        contact = request.POST['contact']
        gst_type = request.POST['gst']
        state = request.POST['state']
        address = request.POST['address']
        email = request.POST['email']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('creditlimit', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']
       
        if (
          not party_name
          
      ):
          return render(request, 'add_parties.html')

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=staff.company.user,company=staff.company)
        part.save()  
        print("===========saved")

        options = {}
        option_objects = party.objects.filter(company=staff.company)
        for option in option_objects:
            options[option.id] = [option.party_name]
        print("===========get")
        return JsonResponse(options) 
  else:
    return JsonResponse({'error': 'Invalid request'}, status=400)
  
def add_item(request):
  if request.method=='POST':
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    # user= request.user
    # user = User.objects.get(id=user.id)
    # company_user_data = company.objects.get(user=request.user)
    cmp= staff.company
    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' or None :
      item_opening_stock = 0
      item_current_stock = 0
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '' or None:
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == ''  or None:
      item_min_stock_maintain = 0
    item_data = ItemModel(user=cmp.user,
                          company=cmp,
                          item_name=item_name,
                          item_hsn=item_hsn,
                          item_unit=item_unit,
                          item_taxable=item_taxable,
                          item_gst=item_gst,
                          item_igst=item_igst,
                          item_sale_price=item_sale_price,
                          item_purchase_price=item_purchase_price,
                          item_opening_stock=item_opening_stock,
                          item_current_stock=item_current_stock,
                          item_at_price=item_at_price,
                          item_date=item_date,
                          item_min_stock_maintain=item_min_stock_maintain)
    item_data.save()
    options = {}
    option_objects = ItemModel.objects.filter(company=cmp,user=cmp.user)
    for option in option_objects:
      options[option.id] = [option.item_name]
    return JsonResponse(options) 
  else:
    return JsonResponse({'error': 'Invalid request'}, status=400)
  
def sales_transaction(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)  
  tr= saleorder_transaction.objects.filter(sales_order=id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  
  context={'tr':tr,'staff':staff,'allmodules':allmodules}
  return render(request,'company/sale_transaction.html',context)

# @login_required(login_url='login')
def saleorder_edit(request,id):
  
  staff_id = request.session['staff_id']
           
  staff =  staff_details.objects.get(id=staff_id)
  sale = salesorder.objects.get(id=id)
  cmp = staff.company
  par= party.objects.filter(company=staff.company)
  item = ItemModel.objects.filter(company=staff.company)
  sitem = sales_item.objects.filter(sale_order=sale)
  bnk = BankModel.objects.filter(company=cmp)
  allmodules= modules_list.objects.get(company=staff.company,status='New')

  
  context={
    'party':par,'item':item,'staff':staff,'bnk':bnk,'sale':sale,'sitem':sitem,
    'allmodules':allmodules,
  }
  return render(request, 'company/saleorder_edit.html',context)

def edit_saleorder(request,id):
  print("===========11111111111111111")
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  print("===========2222222")

  if request.method == 'POST':
    so = salesorder.objects.get(id=id)
    
    prtyid=request.POST.get('party')
    prty=party.objects.get(id=prtyid)
    so.party=prty
    so.partyname = prty.party_name
    so.orderno=request.POST.get('orderno')
    so.orderdate=request.POST.get('orderdate')
    so.duedate=request.POST.get('duedate')
    print(so.partyname)

    pos=request.POST.get('stateofsply')
    if pos != '':
      so.placeofsupply=pos
      if pos == 'State':
        so.CGST=request.POST.get('cgst')
        so.SGST=request.POST.get('sgst')
        so.IGST=''
      elif pos == 'Other state':
        so.IGST=request.POST.get('igst')
        so.CGST= ''
        so.SGST= ''
      
    payment = request.POST.get('paymethode')
    if payment != '':
      so.payment_method=payment
      if payment == 'Cheque':
        so.checkno = request.POST.get('checkno')
      elif payment == 'UPI':
        so.UPI = request.POST.get('upiid')
      elif payment != 'Cheque' and payment != 'UPI 'and payment != 'Cash':
        so.accno = request.POST.get('accno')
        
    if request.FILES.get('attach') is not None:
      new_file = request.FILES.get('attach')
      so.file = new_file
    
    so.note=request.POST.get('note')
    so.subtotal=request.POST.get('subtotal')
    so.taxamount=request.POST.get('taxamount')
    so.adjustment=request.POST.get('adj')
    so.grandtotal=request.POST.get('grandtotal')
    so.paid=request.POST.get('paid')
    so.balance=request.POST.get('baldue')
    
    so.save()
    print("updated===================================")
    salesorderid=salesorder.objects.get(id =so.id)

    
    product = request.POST.getlist("product[]")
    hsn  = request.POST.getlist("hsn[]")
    qty = request.POST.getlist("qty[]")
    price = request.POST.getlist("price[]")
    tax = request.POST.getlist("tax1[]")
    discount = request.POST.getlist("discount[]")
    total = request.POST.getlist("total[]")
    taxamount = request.POST.getlist("taxamount[]")
    # salesorderid=salesorder.objects.get(id =so.id)
    print(len(product))
    print(len(hsn))
    print(len(qty))
    print(len(price))
    
    objects_to_delete = sales_item.objects.filter(sale_order=salesorderid)
    objects_to_delete.delete()
   
    if len(product)==len(hsn)==len(qty) ==len(price)==len(tax)==len(discount)==len(total)==len(taxamount):
      mapped = zip(product, hsn, qty, price, tax, discount, total,taxamount)
      mapped = list(mapped)
      for ele in mapped:
        print(ele[0])
        prod=ItemModel.objects.get(id=ele[0])
        salesorderAdd = sales_item(
          product=prod,
          hsn=ele[1],
          qty=ele[2],
          price=ele[3],
          tax=ele[4],
          discount=ele[5],
          total=ele[6],
          taxamount=ele[7],
          sale_order=salesorderid,
          cmp=staff.company
            )
        salesorderAdd.save()
      
    tran= saleorder_transaction.objects.create(
      sales_order=salesorderid,staff=staff,company=staff.company,action="Updated",date=date.today()
    )
    # tran.save()
    
    return redirect('sale_order')
    
  return redirect('sale_order')


def saleorderto_invoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  company_instance = company.objects.get(id=staff.company.id)
  sale = salesorder.objects.get(id=id)
  itm = sales_item.objects.filter(sale_order=sale)
  par= party.objects.get(id=sale.party.id)
  Party=party.objects.filter(company=staff.company)
  print(Party)
  item=ItemModel.objects.filter(company=company_instance)
  bank=BankModel.objects.filter(company=company_instance)
  allmodules= modules_list.objects.get(company=staff.company.id,status='New')
  if SalesInvoice.objects.filter(company=company_instance).exists():
        invoice_count = SalesInvoice.objects.last().invoice_no
        next_count = invoice_count+1
  else:
        next_count=1
  
  context = {'staff':staff,'Party':Party,'item':item,
             'bank':bank,'allmodules':allmodules,'sale':sale,
             'next_count':next_count,'par':par,'itm':itm}


  return render(request, 'company/saleorderto_invoice.html',context)



def saleorder_convert(request, sid):
  if 'staff_id' in request.session:
    staff_id = request.session['staff_id']
  else:
    return redirect('/')
  staff = staff_details.objects.get(id=staff_id)
  company_instance = staff.company       
  partyid = request.POST.get('partyname')
  p= party.objects.get(id=partyid)
  contact = request.POST.get('contact')
  address = request.POST.get('address')
  invoice_no = request.POST.get('invoiceno')
  date = request.POST.get('date')
  state_of_supply = request.POST.get('state_of_supply')
  paymenttype = request.POST.get('bank')
  cheque = request.POST.get('chequeNumber')
  upi = request.POST.get('upiNumber')
  accountno = request.POST.get('accountNumber')
  item = request.POST.getlist('item[]')
  hsn = request.POST.getlist('hsn[]')
  quantity = request.POST.getlist('quantity[]')
  rate = request.POST.getlist('rate[]')
  discount = request.POST.getlist('discount[]')
  tax = request.POST.getlist('tax[]')
  totalamount = request.POST.getlist('amount[]')
  description = request.POST.get('description')
  subtotal = request.POST.get('subtotal')
  igst = request.POST.get('igst')
  cgst = request.POST.get('cgst')
  sgst = request.POST.get('sgst')
  total_taxamount = request.POST.get('total_taxamount')
  adjustment = request.POST.get('adjustment')
  grandtotal = request.POST.get('grandtotal')
  paidoff = request.POST.get('paidoff')
  totalbalance = request.POST.get('totalbalance')
  
  print(total_taxamount)

        
      
  sales_invoice = SalesInvoice(
    staff=staff,
    company=company_instance,
    party=p,
    party_name=p.party_name,
    contact=contact,
    address=address,
    invoice_no=invoice_no,
    date=date,
    state_of_supply=state_of_supply,
    paymenttype=paymenttype,
    cheque=cheque,
    upi=upi,
    accountno=accountno,
    description=description,
    subtotal=subtotal,
    total_taxamount=total_taxamount,
    adjustment=adjustment,
    grandtotal=grandtotal,
    paidoff=paidoff,
    totalbalance=totalbalance,
    )
  if state_of_supply == 'State':
    sales_invoice.cgst=cgst
    sales_invoice.sgst = sgst
  elif state_of_supply == 'Other state':
    sales_invoice.igst=igst
  sales_invoice.save()

  tr_history = SalesInvoiceTransactionHistory(company=company_instance,
                                              staff=staff,      
                                              salesinvoice=sales_invoice,
                                              action="CREATED",
                                              done_by_name=staff.first_name,
                                              )
  tr_history.save()

  sale = salesorder.objects.get(id=sid)
  sale.status= 'order completed'
  sale.action = 'converted to invoice no. '+sales_invoice.invoice_no
  sale.save()
  
  invoice = SalesInvoice.objects.get(id=sales_invoice.id)

  if len(item)==len(hsn)==len(quantity)==len(rate)==len(discount)==len(tax)==len(totalamount):
          mapped=zip(item,hsn,quantity,rate,discount,tax,totalamount)
          mapped=list(mapped)
          for ele in mapped:
            itm = ItemModel.objects.get(id=ele[0])
            SalesInvoiceItem.objects.create(item = itm,hsn=ele[1], quantity=ele[2],rate=ele[3],discount=ele[4],tax=ele[5],totalamount=ele[6],salesinvoice=invoice,company=company_instance)
          return redirect('sale_order')

  return redirect('sale_order')
  
#End

@require_POST
@csrf_exempt
def get_bill_date(request):
    selected_bill_no = request.POST.get('bill_no', None)

    try:
        # Get the latest PurchaseBill with the specified bill_number
        purchase_bill = PurchaseBill.objects.filter(billno=selected_bill_no).latest('billdate')
        bill_date = purchase_bill.billdate.strftime('%Y-%m-%d')
    except PurchaseBill.DoesNotExist:
        return JsonResponse({'error': 'Bill number not found'}, status=400)
    except PurchaseBill.MultipleObjectsReturned:
        # Handle the case where multiple PurchaseBills are found for the same bill_number
        return JsonResponse({'error': 'Multiple PurchaseBills found for the same bill number'}, status=400)

    return JsonResponse({'bill_date': bill_date})
    
def item_save_invoice(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  if not hsn:
    hsn = None

  itm = ItemModel(item_name=name, item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})
  
def item_invoicedropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  product = ItemModel.objects.filter(company=cmp)

  id_list = []
  product_list = []
  for p in product:
    id_list.append(p.id)
    product_list.append(p.item_name)
  return JsonResponse({'id_list':id_list, 'product_list':product_list})   
  
  
def expense_cat_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  cat = Expense_Category.objects.filter(staff__company=cmp)

  id_list = []
  cat_list = []
  for c in cat:
    id_list.append(c.id)
    cat_list.append(c.expense_category)

  return JsonResponse({'id_list':id_list, 'cat_list':cat_list })  
  
#--------------------------------------------Anuvinda K V---------------------------------------------#
@login_required
def view_paymentout(request):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    allmodules = modules_list.objects.get(company=cmp, status='New')
    
    # Assuming you want to display the latest PaymentOut records
    paymentouts = PaymentOut.objects.filter(company=cmp).order_by('ref_no')
    

    if not paymentouts:
        context = {'staff': staff, 'allmodules': allmodules}
        return render(request, 'company/paymentoutempty.html', context)

    context = {'staff': staff, 'allmodules': allmodules, 'paymentouts': paymentouts}
    return render(request, 'company/paymentoutlist.html', context)

def add_paymentout(request):
    toda = date.today()
    tod = toda.strftime("%Y-%m-%d")
    
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    cust = party.objects.filter(company=cmp, user=cmp.user)
    bank = BankModel.objects.filter(company=cmp, user=cmp.user)
    allmodules = modules_list.objects.get(company=staff.company, status='New')
    last_paymentout = PaymentOut.objects.filter(company=cmp).last()

    if last_paymentout:
        # Use the last_paymentout ref_no + 1
        bill_no = last_paymentout.ref_no + 1
    else:
        # Handle the case where there's no last_paymentout
        bill_no = 1

    # Debug code to print the ref_no
    print("Last PaymentOut Ref No:", last_paymentout.ref_no if last_paymentout else None)
    context = {'staff': staff, 'allmodules': allmodules, 'cust': cust, 'cmp': cmp, 'bill_no': bill_no, 'tod': tod, 'bank': bank,'last_paymentout': last_paymentout}
    return render(request, 'company/paymentoutadd.html', context)

def create_paymentout(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        part = party.objects.get(id=request.POST.get('customername'))
        # Find the maximum ref_no in the database
        max_ref_no = PaymentOut.objects.filter(company=cmp).aggregate(Max('ref_no'))['ref_no__max']

        # Use the maximum ref_no + 1 or set to 1 if there are no existing records
        bill_no = max_ref_no + 1 if max_ref_no is not None else 1

        pbill = PaymentOut(
            staff=staff,
            company=cmp,
            party=part,
            ref_no=bill_no,
            billdate=request.POST.get('billdate'),
            pay_method=request.POST.get("method"),
            cheque_no=request.POST.get("cheque_id"),
            upi_no=request.POST.get("upi_id"),
            balance=request.POST.get("balance"),
        )
        pbill.save()

          # Create PaymentOutDetails
        paid = request.POST.get('paid')
        description = request.POST.get('description')
        files = request.FILES.get('files')

        paymentout_details = PaymentOutDetails(
            paymentout=pbill,  # Set the foreign key relationship
            paid=paid,
            description=description,
            files=files
        )
        paymentout_details.save()
      # Record history for creation
        PaymentOutHistory.objects.create(paymentout=pbill, action='created')  
        
        if 'Next' in request.POST:
            return redirect('add_paymentout')

        if "Save" in request.POST:
            return redirect('view_paymentout')
    else:
        return render(request, 'error_page.html', {'error_message': 'Invalid request method'})

def delete_paymentout(request):
    if request.method == 'POST':
        paymentOutId = request.POST.get('paymentOutId')
        try:
            with transaction.atomic():
                # Perform the deletion, e.g., using the Django ORM
                payment_out = get_object_or_404(PaymentOut, id=paymentOutId)
                ref_no = payment_out.ref_no
                payment_out.delete()

                # Update the ref_no of subsequent records sequentially
                PaymentOut.objects.filter(ref_no__gt=ref_no).update(ref_no=models.F('ref_no') - 1)

            return JsonResponse({'success': True})
        except PaymentOut.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Payment Out not found'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})


def details_paymentout(request, id):
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    allmodules = modules_list.objects.get(company=cmp, status='New')

    paymentout = get_object_or_404(PaymentOut, id=id, company=cmp)

    context = {'staff': staff, 'allmodules': allmodules, 'paymentout': paymentout}
    return render(request, 'company/paymentoutdetails.html', context)

def add_pay(request):
    return render(request, 'company/add_pay.html')
    
def create_addpaymentout(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
       


        # Create PaymentOutDetails
        paid = request.POST.get('paid')
        description = request.POST.get('description')
        files = request.FILES.get('files')

        paymentout_details = PaymentOutDetails(
            paid=paid,
            description=description,
            files=files
        )
        paymentout_details.save()
        
        
        if 'Next' in request.POST:
            return redirect('add_pay')

        if "Save" in request.POST:
            return redirect('view_paymentout')
    else:
        return render(request, 'error_page.html', {'error_message': 'Invalid request method'}) 
    
def edit_paymentout(request, id):
    paymentout = get_object_or_404(PaymentOut, id=id)

    if request.method == 'POST':
        # Update the fields based on the form data
        paymentout.billdate = request.POST.get('billdate')
        paymentout.ref_no = request.POST.get('ref_no')
        paymentout.party.party_name = request.POST.get('party_name')
        # Update other fields as needed
        paymentout.party.contact=request.POST.get('contact')
        # Save the changes
        paymentout.save()

        # Update PaymentOutDetails
        paymentout_detail = paymentout.paymentoutdetails_set.first()  # Assuming there's only one PaymentOutDetails per PaymentOut
        if paymentout_detail:
            paymentout_detail.paid = request.POST.get('paid')
            paymentout_detail.save()
        # Record history for update
        PaymentOutHistory.objects.create(paymentout=paymentout, action='updated')
        return redirect('view_paymentout')

    context = {'paymentout': paymentout}
    return render(request, 'company/paymentoutedit.html', context)

def update_paymentout(request, id):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        paymentout = get_object_or_404(PaymentOut, id=id, company=cmp)

        # Update PaymentOut fields based on your form data
        paymentout.billdate = request.POST.get('billdate')
        paymentout.pay_method = request.POST.get('method')
        paymentout.cheque_no = request.POST.get('cheque_id')
        paymentout.upi_no = request.POST.get('upi_id')
        paymentout.balance = request.POST.get('balance')

        # Add more fields as needed...
        
        # Record history for update
        PaymentOutHistory.objects.create(paymentout=paymentout, action='updated')
        # Handle related items in a transaction to ensure consistency
        with transaction.atomic():
            # Update related PaymentOutDetails
            paymentout.paymentoutdetails_set.all().delete()  # Delete existing details

            # Iterate through form data to create new details
        
            
            # Iterate through form data to create new details
            for i in range(int(request.POST.get('total_items', 0))):
                paid = request.POST.get(f'paid_{i}')
                description = request.POST.get(f'description_{i}')
                # Handle file upload if needed
                file = request.FILES.get(f'file_{i}')
                print(f'Index: {i}, Paid: {paid}, Description: {description}, File: {file}')

                # Create new PaymentOutDetails
                PaymentOutDetails.objects.create(
                    paymentout=paymentout,
                    paid=paid,
                    description=description,
                    files=file
                )

        # Save the main PaymentOut object
        paymentout.save()

       
        # Redirect to the view page or list page
        return redirect('view_paymentout')

    # Handle the case where the request method is not POST
    return render(request, 'error_page.html', {'error_message': 'Invalid request method'})

def paymentout_history(request, id):
    paymentout_history = PaymentOutHistory.objects.filter(paymentout_id=id).order_by('-timestamp')
    return render(request, 'company/paymentout_history.html', {'paymentout_history': paymentout_history})

def get_party_details(request):
    party_id = request.GET.get('party_id')

    try:
        party = party.objects.get(id=party_id)
        # Customize the fields as needed based on your Party model
        data = {
            'success': True,
            'billing_address': party.billing_address,
            'phone_number': party.phone_number,
            'available_balance': party.available_balance,
        }
    except party.DoesNotExist:
        data = {'success': False}

    return JsonResponse(data)

@csrf_exempt  # For demonstration purposes, you might want to remove this in production and handle CSRF properly
def send_email(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode('utf-8'))
            email_ids = data.get('emailIds', '')
            email_message = data.get('emailMessage', '')

            # Your email sending logic here
            send_mail(
                'Subject',  # Replace with your subject
                email_message,  # Replace with your email message
                'your_email@example.com',  # Replace with your sender email
                [email_ids],  # Replace with your recipient email(s)
                fail_silently=False,
            )

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
#End

#Akshaya
def gstr3b(request):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  
  return render(request,'company/gstr3B.html',context)


def sharegstr3BToEmail(request):
    if request.method == "POST":
        sid = request.session.get('staff_id')
        staff =  staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        allmodules= modules_list.objects.get(company=cmp,status='New')
        context = {'staff' : staff,'allmodules':allmodules}
        my_subject = "GSTR 3B REPORT"
        emails_string = request.POST['email_ids']
        emails_list = [email.strip() for email in emails_string.split(',')]
        # recipient_email = request.POST.get('email_ids')
        html_message = render_to_string('company/gstr3B_pdf.html',context)#add ur html
        # vyaparapp\templates\index.html
        # vyaparapp\templates\company\gstr3B_pdf.html
        plain_message = strip_tags(html_message)
        pdf_content = BytesIO()
        pisa_document = pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content) 
        pdf_content.seek(0)
        # todo: need to update the from_email
        filename = f'gstr3B {staff.company.company_name}.pdf'
        message = EmailMultiAlternatives(
            subject=my_subject,
            body= f"Hi,\nPlease find the attached Gstr3B Report -  \n\n--\nRegards,\n{staff.company.company_name}\n{staff.company.address}\n{staff.company.state} - {staff.company.country}\n{staff.company.contact}",
            from_email='altostechnologies6@gmail.com',
            to=emails_list,  # Use the recipient_email variable here
            )
        message.attach(filename, pdf_content.read(), 'application/pdf')
        
        try:
            message.send()
            return HttpResponse('<script>alert("Report has been shared via successfully..!");window.location="/gstr3b"</script>')
        except Exception as e:
            # Handle the exception, log the error, or provide an error message
            return HttpResponse('<script>alert("Failed to send email!");window.location="/gstr3b"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/gstr3b"</script>') 


def gstr9(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'company/gstr9.html',context)



def sharegstr9ToEmail(request):
    if request.method == "POST":

        sid = request.session.get('staff_id')
        staff =  staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        allmodules= modules_list.objects.get(company=cmp,status='New')
        context = {'staff' : staff,'allmodules':allmodules}
        
        email_message = request.POST['email_message']
        my_subject = "GSTR9 REPORT"
        emails_string = request.POST['email_ids']
        emails_list = [email.strip() for email in emails_string.split(',')]
        # recipient_email = request.POST.get('email_ids')
        html_message = render_to_string('company/gstr9_pdf.html',context)#add ur html
        # vyaparapp\templates\index.html
        # vyaparapp\templates\company\gstr3B_pdf.html
        plain_message = strip_tags(html_message)
        pdf_content = BytesIO()
        pisa_document = pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content) 
        pdf_content.seek(0)
        # todo: need to update the from_email
        filename = f'gstr9 {staff.company.company_name}.pdf'
        message = EmailMultiAlternatives(
            subject=my_subject,
            body= f"Hi,\nPlease find the attached Gstr9 Report -  \n{email_message}\n--\nRegards,\n{staff.company.company_name}\n{staff.company.address}\n{staff.company.state} - {staff.company.country}\n{staff.company.contact}",
            from_email='altostechnologies6@gmail.com',
            to= emails_list ,  # Use the recipient_email variable here
            )
        message.attach(filename, pdf_content.read(), 'application/pdf')
        
        try:
            message.send()
            return HttpResponse('<script>alert("Report has been shared via successfully..!");window.location="/gstr9"</script>')
        except Exception as e:
            # Handle the exception, log the error, or provide an error message
            return HttpResponse('<script>alert("Failed to send email!");window.location="/gstr9"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/gstr9"</script>') 
    
#End

#______________Payment In__________________shemeem________________________________
def paymentIn(request):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  payments = PaymentIn.objects.filter(company = cmp)
  context = {
    'staff':staff,'allmodules':allmodules,'paymentIn':payments,
  }
  return render(request, 'company/payment_in.html',context)


def createPaymentIn(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    allmodules= modules_list.objects.get(company=com,status='New')
    try:
      parties = party.objects.filter(company = com)
      banks = BankModel.objects.filter(company = com)

      # Fetching last bill and assigning upcoming bill no as current + 1
      # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
      latest_bill = PaymentIn.objects.filter(company = com).order_by('-id').first()

      if latest_bill:
          last_number = int(latest_bill.rec_no)
          new_number = last_number + 1
      else:
          new_number = 1

      if DeletedPaymentIn.objects.filter(company = com).exists():
          deleted = DeletedPaymentIn.objects.get(company = com)
          if deleted:
              while int(deleted.rec_no) >= new_number:
                  new_number+=1
      
      context = {
        'staff':staff, 'company':com,'allmodules':allmodules, 'parties':parties, 'rec_no':new_number,'banks':banks,
      }
      return render(request, 'company/create_payment_in.html',context)
    except Exception as e:
      print(e)
      return redirect(paymentIn)

def getBankDetails(request):        
  try:
      bankId = request.POST.get('id')
      bankDetails = BankModel.objects.get(id = int(bankId))
      return JsonResponse({'status':"true", 'id':bankDetails.id, 'acc_number':bankDetails.account_num})
  except Exception as e:
      print(e)
      return JsonResponse({'status':'false'})


def createNewPaymentIn(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    try:
      if request.method == 'POST':
        payment = PaymentIn(
          staff = staff,
          company = com,
          party = party.objects.get(id = request.POST['party_name']),
          rec_no = request.POST['receipt_no'],
          date = request.POST['date'],
          party_name = party.objects.get(id = request.POST['party_name']).party_name,
          contact = request.POST['contact'],
          billing_address = request.POST['address'],
          description = request.POST['description'],
          payment_type = 'Payment In',
          payment_method = request.POST['payment_method'],
          payment_acc_number = None if request.POST['payment_acc_num'] == "" else request.POST['payment_acc_num'],
          payment_cheque_id = request.POST['payment_cheque_id'],
          payment_upi_id = request.POST['payment_upi_id'],
          total_amount = request.POST['payment_amount'],
          payment_received = request.POST['payment_amount'],
          balance = 0.0,
        )
        payment.save()

        #Transaction History
        history = PaymentInTransactionHistory(
          staff = staff,
          payment = payment,
          company = com,
          action = "Created",
          date = payment.date
        )
        history.save()
        print('saved...')
        if 'save_and_next' in request.POST:
          return redirect(createPaymentIn)
        return redirect(paymentIn)
    except Exception as e:
      print(e)
      return redirect(createPaymentIn)


def deletePaymentIn(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      pay = PaymentIn.objects.get(id = id)

      # Storing receipt number to deleted table
      # if entry exists and lesser than the current, update and save => Only one entry per company

      if DeletedPaymentIn.objects.filter(company = com).exists():
        deleted = DeletedPaymentIn.objects.get(company = com)
        if deleted:
          if int(pay.rec_no) > int(deleted.rec_no):
            deleted.rec_no = pay.rec_no
            deleted.save()
      else:
        deleted = DeletedPaymentIn(company = com, staff = staff, rec_no = pay.rec_no)
        deleted.save()

      pay.delete()
      messages.success(request, 'Payment In data deleted successfully.!')
      return redirect(paymentIn)
    except Exception as e:
      print(e)
      return redirect(paymentIn)
  return redirect('/')


def paymentHistory(request):
  pid = request.POST['id']
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  pay = PaymentIn.objects.get(rec_no=pid, company=cmp)
  hst = PaymentInTransactionHistory.objects.filter(payment = pay).last()
  name = hst.staff.first_name + ' ' + hst.staff.last_name 
  action = hst.action
  return JsonResponse({'name':name,'action':action,'pid':pid})


def viewPaymentIn(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  paymentInDetails = PaymentIn.objects.get(id = id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  context = {
    'payment':paymentInDetails,'staff':staff,'allmodules':allmodules,'company':cmp,
  }

  return render(request, 'company/payment_in_details.html',context)


def sharePaymentInToEmail(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      if request.method == 'POST':
        emails_string = request.POST['email_ids']

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(',')]
        email_message = request.POST['email_message']
        # print(emails_list)

        payment = PaymentIn.objects.get(id = id)
        context = {'payment': payment,'company':com}
        template_path = 'company/payment_in_pdf.html'
        template = get_template(template_path)

        html  = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f'Payment In - {payment.rec_no}.pdf'
        subject = f"Payment In Receipt - {payment.rec_no}"
        email = EmailMessage(subject, f"Hi,\nPlease find the attached Receipt of Payment In -{payment.rec_no}. \n{email_message}\n\n--\nRegards,\n{com.company_name}\n{com.address}\n{com.city} - {com.state}\n{com.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)

        messages.success(request, 'Receipt has been shared via email successfully..!')
        return redirect(viewPaymentIn,id)
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect(viewPaymentIn, id)


def editPaymentIn(request, id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      payment = PaymentIn.objects.get(id = id)
      allmodules= modules_list.objects.get(company=com,status='New')
      parties = party.objects.filter(company = com)
      banks = BankModel.objects.filter(company = com)
      context = {
        'payment':payment,'staff':staff,'allmodules':allmodules,'company':com,'parties':parties,'banks':banks,
      }
      return render(request, 'company/edit_payment_in.html',context)
    except Exception as e:
      print(e)
      return redirect(viewPaymentIn,id)



def updatePaymentIn(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)

    try:
      payment = PaymentIn.objects.get(id = id)
      if request.method == 'POST':
        payment.staff = staff
        payment.company = com
        payment.party = party.objects.get(id = request.POST['party_name'])
        payment.rec_no = request.POST['receipt_no']
        payment.date = request.POST['date']
        payment.party_name = party.objects.get(id = request.POST['party_name']).party_name
        payment.contact = request.POST['contact']
        payment.billing_address = request.POST['address']
        payment.description = request.POST['description']
        payment.payment_type = 'Payment In'
        payment.payment_method = request.POST['payment_method']
        payment.payment_acc_number = None if request.POST['payment_acc_num'] == "" else request.POST['payment_acc_num']
        payment.payment_cheque_id = request.POST['payment_cheque_id']
        payment.payment_upi_id = request.POST['payment_upi_id']
        payment.total_amount = request.POST['payment_amount']
        payment.payment_received = request.POST['payment_amount']
        payment.balance = 0.0
        payment.save()

        #Transaction History
        history = PaymentInTransactionHistory(
          staff = staff,
          payment = payment,
          company = com,
          action = "Updated",
          date = date.today()
        )
        history.save()

        return redirect(viewPaymentIn,id)
    except Exception as e:
      print(e)
      return redirect(editPaymentIn,id)

def paymentInHistory(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  pay = PaymentIn.objects.get(id=id,company=cmp)
  hst= PaymentInTransactionHistory.objects.filter(payment=pay,company=cmp)

  context = {'staff':staff,'allmodules':allmodules,'history':hst,'payment':pay}
  return render(request,'company/payment_in_history.html',context)


def downloadPaymentSampleImportFile(request):
  payment_table_data = [['SLNO','DATE','NAME','PAYMENT METHOD','ACCOUNT NUMBER','CHEQUE ID','UPI ID','TOTAL','RECEIVED','BALANCE','DESCRIPTION'], ['1', '2023-11-20', 'John Doe', 'Canara', '767676677667677','','','1000','500','0','Description']]

  wb = Workbook()

  sheet1 = wb.active
  sheet1.title = 'payment'

  # Populate the sheets with data
  for row in payment_table_data:
    sheet1.append(row)

  # Create a response with the Excel file
  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename=payment_sample_file.xlsx'

  # Save the workbook to the response
  wb.save(response)
  return response


def importPaymentFromExcel(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
            
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)    
    
    current_datetime = timezone.now()
    dateToday =  current_datetime.date()

    if request.method == "POST" and 'excel_file' in request.FILES:
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)

        # checking challan sheet columns
        try:
          ws = wb["payment"]
        except:
          print('sheet not found')
          messages.error(request,'`payment` sheet not found.! Please check.')
          return redirect(paymentIn)
        
        ws = wb["payment"]
        payment_columns = ['SLNO','DATE','NAME','PAYMENT METHOD','ACCOUNT NUMBER','CHEQUE ID','UPI ID','TOTAL','RECEIVED','BALANCE','DESCRIPTION']
        payment_sheet = [cell.value for cell in ws[1]]
        if payment_sheet != payment_columns:
          print('invalid sheet')
          messages.error(request,'`payment` sheet column names or order is not in the required formate.! Please check.')
          return redirect(paymentIn)

        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,name,payment_method,acc_num,cheque,upi,total,received,balance,description = row
          if slno is None or name is None or total is None or received is None:
            messages.error(request,'`payment` sheet entries missing required fields.! Please check.')
            return redirect(paymentIn)
        
        # getting data from estimate sheet and create estimate.
        incorrect_data = []
        ws = wb['payment']
        for row in ws.iter_rows(min_row=2, values_only=True):
          slno,date,name,payment_method,acc_num,cheque,upi,total,received,balance,description = row

          # Fetching last bill and assigning upcoming bill no as current + 1
          # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
          latest_bill = PaymentIn.objects.filter(company = com).order_by('-id').first()
          
          if latest_bill:
              last_number = int(latest_bill.rec_no)
              new_number = last_number + 1
          else:
              new_number = 1

          if DeletedPaymentIn.objects.filter(company = com).exists():
              deleted = DeletedPaymentIn.objects.get(company = com)
              if deleted:
                  while int(deleted.rec_no) >= new_number:
                      new_number+=1
          if not party.objects.filter(company = com, party_name = name).exists():
            incorrect_data.append(slno)
            continue
          try:
            prt = party.objects.get(company = com, party_name = name)
            cntct = prt.contact
            adrs = prt.address
          except:
            pass

          if date is None:
            date = dateToday

          payment = PaymentIn(
            staff = staff,
            company = com,
            date = date,
            rec_no = new_number,
            party = prt,
            party_name = name,
            contact = cntct,
            billing_address = adrs,
            description = description,
            payment_type = 'Payment',
            payment_method = payment_method,
            payment_acc_number = acc_num,
            payment_cheque_id = cheque,
            payment_upi_id = upi,
            total_amount = total,
            payment_received = received,
            balance = 0 if balance is None else balance,
          )
          payment.save()

          # Transaction history
          history = PaymentInTransactionHistory(
            staff = staff,
            payment = payment,
            company = com,
            action = "Created"
          )
          history.save()

    messages.success(request, 'Data imported successfully.!')
    if incorrect_data:
      messages.warning(request, f'Data with following Sl No could not import due to incorrect data provided - {", ".join(str(item) for item in incorrect_data)}')
    return redirect(paymentIn)
    
#End

def gstrr2(request):
    if 'staff_id' in request.session:
        staff_id = request.session['staff_id']
    else:
        return redirect('/')

    staff = staff_details.objects.get(id=staff_id)
    comp = company.objects.get(id=staff.company.id)

    # Filter PurchaseBill instances related to the specific company
    purchasebill = PurchaseBill.objects.filter(company=comp)

    # Filter party instances related to the specific company
    partydata = party.objects.filter(company=comp)
    allmodules= modules_list.objects.get(company=staff.company,status='New')

    return render(request, 'company/gstr_2.html', {'staff':staff,'company': comp,'purchasebill': purchasebill, 'partydata': partydata,'allmodules':allmodules})

def gstrnew1(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  comp =  company.objects.get(id = staff.company.id)

  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'company/gstr_1.html',{'staff':staff,'company':comp,'allmodules':allmodules})
       

def sharepurchaseBillToEmail(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    comp =  company.objects.get(id = staff.company.id)
    try:
      if request.method == 'POST':
        emails_string = request.POST['email_ids']

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(',')]
        email_message = request.POST['email_message']
        # print(emails_list)

        # comp = company.objects.get(user_id=request.user.id)
        purchasebill =  PurchaseBill.objects.all()
        partydata = party.objects.all()
        allmodules= modules_list.objects.get(company=staff.company,status='New')
        context = {'purchasebill': purchasebill,'partydata': partydata,'allmodules': allmodules, 'company': comp}


        template_path = 'company/gstr1_pdf.html'
        template = get_template(template_path)

        html  = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f'Purchase Bill - {comp.company_name}.pdf'
        subject = f"Purchase Bill Receipt - {comp.company_name}"
        email = EmailMessage(subject, f"Hi,\nPlease find the attached Receipt of Purchase Bill -{comp.company_name}. \n{email_message}\n\n--\nRegards,\n{comp.company_name}\n{comp.address}\n{comp.city} - {comp.state}\n{comp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)

        messages.success(request, 'Purchase Bill has been shared via email successfully..!')
        return redirect(gstrnew1)
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect(gstrnew1)  
    
def shareGSTR2purchaseBillToEmail(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    comp =  company.objects.get(id = staff.company.id)
    try:
      if request.method == 'POST':
        emails_string = request.POST['email_ids']

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(',')]
        email_message = request.POST['email_message']
        # print(emails_list)

        # comp = company.objects.get(user_id=request.user.id)
        purchasebill =  PurchaseBill.objects.all()
        partydata = party.objects.all()
        allmodules= modules_list.objects.get(company=staff.company,status='New')
        context = {'purchasebill': purchasebill,'partydata': partydata,'allmodules': allmodules, 'company': comp}


        template_path = 'company/gstr2_pdf.html'
        template = get_template(template_path)

        html  = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f'Purchase Bill - {comp.company_name}.pdf'
        subject = f"Purchase Bill Receipt - {comp.company_name}"
        email = EmailMessage(subject, f"Hi,\nPlease find the attached Receipt of Purchase Bill -{comp.company_name}. \n{email_message}\n\n--\nRegards,\n{comp.company_name}\n{comp.address}\n{comp.city} - {comp.state}\n{comp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)

        messages.success(request, 'Purchase Bill has been shared via email successfully..!')
        return redirect(gstrr2)
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect(gstrr2)  
        
        
def convertEstimateToSalesOrder(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    allmodules= modules_list.objects.get(company=staff.company,status='New')
    cmp = company.objects.get(id=staff.company.id)
    par= party.objects.filter(company=cmp)
    item = ItemModel.objects.filter(company=cmp)
    bnk = BankModel.objects.filter(company=cmp)
    estimate = Estimate.objects.get(id = id)
    est_items = Estimate_items.objects.filter(eid = estimate)
    order = salesorder.next_orderno()
    
    context={
      'party':par,'item':item,'staff':staff,'order':order,'bnk':bnk,'allmodules':allmodules,'estimate':estimate, 'est_items':est_items,
    }
    return render(request, 'company/estimate_to_salesorder.html',context)


def saveEstimateToSalesOrder(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    estimate = Estimate.objects.get(id = id)
    if request.method == 'POST':
      prty = request.POST.get('party')
      cmp= staff.company
      payment = request.POST.get('paymethode')
      pos=request.POST.get('stateofsply')
      attach=request.FILES.get('attach')       

      sale = salesorder(
        partyname=prty,
        orderno=request.POST.get('orderno'),
        orderdate=request.POST.get('orderdate'),
        duedate=request.POST.get('duedate'),
        placeofsupply=pos,
        payment_method=payment,
        subtotal=request.POST.get('subtotal'),
        taxamount=request.POST.get('taxamount'),
        adjustment=request.POST.get('adj'),
        grandtotal=request.POST.get('grandtotal'),
        note=request.POST.get('note'),
        paid=request.POST.get('paid'),
        balance=request.POST.get('baldue'),
        file=attach,
        staff=staff,
        comp=cmp,
      )

      if payment == 'Cheque':
        sale.checkno = request.POST.get('checkno')
      elif payment == 'UPI':
        sale.UPI = request.POST.get('upiid')
      elif payment != 'Cheque' and payment != 'UPI 'and payment != 'Cash':
        sale.accno = request.POST.get('accno')
      
      if pos == 'state':
        sale.CGST=request.POST.get('cgst')
        sale.SGST=request.POST.get('sgst')
      elif pos == 'other state':
        sale.IGST=request.POST.get('igst')

      sale.save()
      
      product = request.POST.getlist("product[]")
      hsn  = request.POST.getlist("hsn[]")
      qty = request.POST.getlist("qty[]")
      price = request.POST.getlist("price[]")
      tax = request.POST.getlist("tax1[]")
      discount = request.POST.getlist("discount[]")
      total = request.POST.getlist("total[]")
      salesorderid=salesorder.objects.get(id =sale.id)
    
      if len(product)==len(hsn)==len(qty) ==len(price)==len(tax)==len(discount)==len(total):
        mapped = zip(product, hsn, qty, price, tax, discount, total)
        mapped = list(mapped)
        for ele in mapped:
          print(ele[0])
          prod=ItemModel.objects.get(id=ele[0])
          salesorderAdd = sales_item(
            product=prod,
            hsn=ele[1],
            qty=ele[2],
            price=float(ele[3]),
            tax=ele[4],
            discount=float(ele[5]),
            total=float(ele[6]),
            sale_order=salesorderid,
            cmp=staff.company
          )
          salesorderAdd.save()

      tran= saleorder_transaction(
        sales_order=salesorderid,staff=staff,company=cmp,action="Created",date=date.today()
      )
      tran.save()

      estimate.status = 'Completed'
      estimate.is_converted = True
      estimate.balance = sale.balance
      estimate.sales_order = salesorder.objects.get(id = sale.id)
      estimate.save()


      return redirect(estimate_quotation)


def convertEstimateToInvoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    company_instance = company.objects.get(id=staff.company.id)

    Party=party.objects.filter(company=company_instance)
    item=ItemModel.objects.filter(company=company_instance)
    allmodules= modules_list.objects.get(company=staff.company.id,status='New')
    bank=BankModel.objects.filter(company=company_instance)
    if SalesInvoice.objects.filter(company=company_instance).exists():
          invoice_count = SalesInvoice.objects.last().invoice_no
          next_count = invoice_count+1
    else:
          next_count=1
    estimate = Estimate.objects.get(id = id)
    est_items = Estimate_items.objects.filter(eid = estimate)
    
    context={
      'staff':staff,'Party':Party,'item':item,'bank':bank,'count':next_count,'allmodules':allmodules,'estimate':estimate, 'est_items':est_items,
    }
    return render(request, 'company/estimate_to_invoice.html',context)


def saveEstimateToInvoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    estimate = Estimate.objects.get(id = id)
    company_instance = staff.company
    if request.method == 'POST':
      party_name = request.POST.get('partyname')
      contact = request.POST.get('contact')
      address = request.POST.get('address')
      invoice_no = request.POST.get('invoiceno')
      date = request.POST.get('date')
      state_of_supply = request.POST.get('state_of_supply')
      paymenttype = request.POST.get('bank')
      cheque = request.POST.get('chequeNumber')
      upi = request.POST.get('upiNumber')
      accountno = request.POST.get('accountNumber')
      product = tuple(request.POST.getlist("product[]"))
      hsn =  tuple(request.POST.getlist("hsn[]"))
      qty =  tuple(request.POST.getlist("qty[]"))
      rate =  tuple(request.POST.getlist("price[]"))
      discount =  tuple(request.POST.getlist("discount[]"))
      tax =  tuple(request.POST.getlist("tax[]"))
      total =  tuple(request.POST.getlist("total[]"))
      description = request.POST.get('description')
      advance = request.POST.get("advance")
      balance = request.POST.get("balance")
      subtotal = float(request.POST.get('subtotal'))
      igst = request.POST.get('igst')
      cgst = request.POST.get('cgst')
      sgst = request.POST.get('sgst')
      adjust = request.POST.get("adj")
      taxamount = request.POST.get("taxamount")
      grandtotal=request.POST.get('grandtotal')

      party_instance=party.objects.get(party_name=party_name)
      
    
      sales_invoice = SalesInvoice(
        staff=staff,
        company=company_instance,
        party=party_instance,
        party_name=party_name,
        contact=contact,
        address=address,
        invoice_no=invoice_no,
        date=date,
        state_of_supply=state_of_supply,
        paymenttype=paymenttype,
        cheque=cheque,
        upi=upi,
        accountno=accountno,
        description=description,
        subtotal=subtotal,
        igst=igst,
        cgst=cgst,
        sgst=sgst,
        total_taxamount=taxamount,
        adjustment=adjust,
        grandtotal=grandtotal,
        paidoff=advance,
        totalbalance=balance,
      )

      sales_invoice.save()

      tr_history = SalesInvoiceTransactionHistory(
        company=company_instance,
        staff=staff,      
        salesinvoice=sales_invoice,
        action="CREATED",
        done_by_name=staff.first_name,
      )
      tr_history.save()

      invoice = SalesInvoice.objects.get(id=sales_invoice.id)
      mapped = []  # Initialize mapped
      if len(product)==len(hsn)==len(qty)==len(rate)==len(discount)==len(tax)==len(total):
        mapped=zip(product, hsn, qty, rate, discount, tax, total)
        mapped=list(mapped)
        for ele in mapped: 
          itm = ItemModel.objects.get(id=ele[0])
          SalesInvoiceItem.objects.create(item=itm, hsn=ele[1], quantity=ele[2], rate=float(ele[3]), discount=float(ele[4]), tax=ele[5], totalamount=float(ele[6]), salesinvoice=invoice, company=company_instance)
      

      estimate.status = 'Completed'
      estimate.is_converted = True
      estimate.balance = sales_invoice.totalbalance
      estimate.invoice = SalesInvoice.objects.get(id = sales_invoice.id)
      estimate.save()

      return redirect(estimate_quotation)


def convertChallanToInvoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    company_instance = company.objects.get(id=staff.company.id)

    Party=party.objects.filter(company=company_instance)
    item=ItemModel.objects.filter(company=company_instance)
    allmodules= modules_list.objects.get(company=staff.company.id,status='New')
    bank=BankModel.objects.filter(company=company_instance)
    if SalesInvoice.objects.filter(company=company_instance).exists():
          invoice_count = SalesInvoice.objects.last().invoice_no
          next_count = invoice_count+1
    else:
          next_count=1
    challan = DeliveryChallan.objects.get(id = id)
    ch_items = DeliveryChallanItems.objects.filter(cid = challan)
    
    context={
      'staff':staff,'Party':Party,'item':item,'bank':bank,'count':next_count,'allmodules':allmodules,'challan':challan, 'ch_items':ch_items,
    }
    return render(request, 'company/challan_to_invoice.html',context)


def saveChallanToInvoice(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    challan = DeliveryChallan.objects.get(id = id)
    company_instance = staff.company
    if request.method == 'POST':
      party_name = request.POST.get('partyname')
      contact = request.POST.get('contact')
      address = request.POST.get('address')
      invoice_no = request.POST.get('invoiceno')
      date = request.POST.get('date')
      state_of_supply = request.POST.get('state_of_supply')
      paymenttype = request.POST.get('bank')
      cheque = request.POST.get('chequeNumber')
      upi = request.POST.get('upiNumber')
      accountno = request.POST.get('accountNumber')
      product = tuple(request.POST.getlist("product[]"))
      hsn =  tuple(request.POST.getlist("hsn[]"))
      qty =  tuple(request.POST.getlist("qty[]"))
      rate =  tuple(request.POST.getlist("price[]"))
      discount =  tuple(request.POST.getlist("discount[]"))
      tax =  tuple(request.POST.getlist("tax[]"))
      total =  tuple(request.POST.getlist("total[]"))
      description = request.POST.get('description')
      advance = request.POST.get("advance")
      balance = request.POST.get("balance")
      subtotal = float(request.POST.get('subtotal'))
      igst = request.POST.get('igst')
      cgst = request.POST.get('cgst')
      sgst = request.POST.get('sgst')
      adjust = request.POST.get("adj")
      taxamount = request.POST.get("taxamount")
      grandtotal=request.POST.get('grandtotal')

      party_instance=party.objects.get(party_name=party_name)
      
    
      sales_invoice = SalesInvoice(
        staff=staff,
        company=company_instance,
        party=party_instance,
        party_name=party_name,
        contact=contact,
        address=address,
        invoice_no=invoice_no,
        date=date,
        state_of_supply=state_of_supply,
        paymenttype=paymenttype,
        cheque=cheque,
        upi=upi,
        accountno=accountno,
        description=description,
        subtotal=subtotal,
        igst=igst,
        cgst=cgst,
        sgst=sgst,
        total_taxamount=taxamount,
        adjustment=adjust,
        grandtotal=grandtotal,
        paidoff=advance,
        totalbalance=balance,
      )

      sales_invoice.save()

      tr_history = SalesInvoiceTransactionHistory(
        company=company_instance,
        staff=staff,      
        salesinvoice=sales_invoice,
        action="CREATED",
        done_by_name=staff.first_name,
      )
      tr_history.save()

      invoice = SalesInvoice.objects.get(id=sales_invoice.id)
      mapped = []  # Initialize mapped
      if len(product)==len(hsn)==len(qty)==len(rate)==len(discount)==len(tax)==len(total):
        mapped=zip(product, hsn, qty, rate, discount, tax, total)
        mapped=list(mapped)
        for ele in mapped: 
          itm = ItemModel.objects.get(id=ele[0])
          SalesInvoiceItem.objects.create(item=itm, hsn=ele[1], quantity=ele[2], rate=float(ele[3]), discount=float(ele[4]), tax=ele[5], totalamount=float(ele[6]), salesinvoice=invoice, company=company_instance)
      

      challan.status = 'Completed'
      challan.is_converted = True
      challan.balance = sales_invoice.totalbalance
      challan.invoice = SalesInvoice.objects.get(id = sales_invoice.id)
      challan.save()

      return redirect(delivery_challan)
      
      
def shareinvoiceToEmail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

    
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)

                sid = request.session.get('staff_id')
                staff =  staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
               
                salesinvoice = SalesInvoice.objects.get(id=id,company=cmp)
                salesinvoiceitem = SalesInvoiceItem.objects.filter(salesinvoice=salesinvoice,company=cmp)
                        
                context = {'salesinvoice':salesinvoice, 'cmp':cmp,'salesinvoiceitem':salesinvoiceitem}
                template_path = 'company/salesinvoice_mailfile.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'INVOICE - {salesinvoice.id}.pdf'
                subject = f"INVOICE - {salesinvoice.id}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached INVOICE - File-{salesinvoice.id}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Invoice file has been shared via email successfully..!')
                return redirect(salesinvoice_billtemplate,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(salesinvoice_billtemplate, id)
            
def sales_report(request):
  id=request.session.get('staff_id')
  staff =  staff_details.objects.get(id=id)
  sale = salesorder.objects.filter(comp=staff.company)
  print(sale)
  c=sale.count()
  s=0
  for i in sale:
    s += float(i.grandtotal)
  content={
    'sale':sale,
    'staff':staff,
    'c':c,
    's':s,
  }
  return render(request,'company/sale_order_report.html',content)
#--------------------------------------------------------------------
def purchase_report(request):
  id=request.session.get('staff_id')
  staff=staff_details.objects.get(id=id)
  purchase_data=PurchaseBill.objects.filter(company=staff.company)
  debit_note=purchasedebit.objects.filter(company=staff.company)
  paid = unpaid = total=0
  for i in purchase_data:
    paid +=float(i.advance)
    unpaid +=float(i.balance)
    total +=float(i.grandtotal)
  content={
    'bill':purchase_data,
    'debit':debit_note,
    'staff':staff,
    'paid':paid,
    'unpaid':unpaid,
    'total':total
  }
  return render(request,'company/purchase_report.html',content)
#-------------------------------------------------------------------------------
def send_sale_report_via_mail(request):
  if request.method == 'POST':
    emails=request.POST['email']
    mess=request.POST['message']
    uploaded_file = request.FILES.get('file')
    if uploaded_file and uploaded_file.name.endswith('.pdf'):
      subject = 'PDF Report'
      message = mess
      email = EmailMessage(subject, message, to=[emails])
      email.attach(uploaded_file.name, uploaded_file.read(), 'application/pdf')
      email.send()
      messages.success(request, 'purchase report file has been shared via email successfully..!')
      return redirect('sales_report')
    else:
      messages.success(request, '!Invalid pdf')
      return redirect('sales_report')
  return redirect('purchase_report')
#------------------------------------------------------------------------------------
def send_purchase_report_via_mail(request):
  if request.method == 'POST':
    emails=request.POST['email']
    mess=request.POST['message']
    uploaded_file = request.FILES.get('file')
    if uploaded_file and uploaded_file.name.endswith('.pdf'):
      subject = 'PDF Report'
      message = mess
      email = EmailMessage(subject, message, to=[emails])
      email.attach(uploaded_file.name, uploaded_file.read(), 'application/pdf')
      email.send()
      messages.success(request, 'purchase report file has been shared via email successfully..!')
      return redirect('purchase_report')
    else:
      messages.success(request, '!Invalid pdf')
      return redirect('purchase_report')
  return redirect('purchase_report')
#-------------------------------------------------------------------------------
def day_book_report(request):
  id=request.session.get('staff_id')
  staff=staff_details.objects.get(id=id)
  return render(request,'company/day_book_report.html',{'staff':staff})
  
def add_loan_accounts_function(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        lender_bank = request.POST.get('lender_bank')
        current_balance = Decimal(request.POST.get('current_balance', 0))
        raw_date = request.POST.get('date')
        formatted_date = (
            datetime.strptime(raw_date, '%d/%m/%Y').strftime('%Y-%m-%d')
            if raw_date
            else None
        )
        loan_received = request.POST.get('loan_received')
        account_name = request.POST.get('Account_Name')   

    
        raw_interest_rate = request.POST.get('rate', '0.00')

        try:
            interest_rate = Decimal(raw_interest_rate)
        except (InvalidOperation, ConversionSyntax) as e:
            print(f"Error converting 'rate' to Decimal: {e}")
            interest_rate = Decimal(0.00)

        raw_fee = request.POST.get('fee', '0.00')

        try:
            fee = Decimal(raw_fee)
        except (InvalidOperation, ConversionSyntax) as e:
            print(f"Error converting 'fee' to Decimal: {e}")
            fee = Decimal(0.00)

        duration = request.POST.get('duration')
        description = request.POST.get('description')
        account_number = request.POST.get('account_number')
   
        lr = request.POST.get('lr')
        cheque_number = request.POST.get('cheque_number')
        upi_id = request.POST.get('upi_id')
        upi_id_for_fee = request.POST.get('upi_id_for_fee')
        cheque_number_for_fee = request.POST.get('cheque_number_for_fee')

        total_amount = fee + interest_rate + current_balance

        new_loan_account = LoanAccounts.objects.create(
            company=cmp,
            account_name=account_name,
            lender_bank=lender_bank,
            loan_amount=current_balance,
            date=formatted_date,
            loan_received=loan_received,
            interest_rate=interest_rate,
            duration=duration,
            description=description,
            proccessing_fee=fee,
            lr=lr,
            cheque_number=cheque_number,
            upi_id=upi_id,
            account_number=account_number,
            upi_id_for_fee=upi_id_for_fee,
            cheque_number_for_fee=cheque_number_for_fee,
            total_amount=str(total_amount)
        )

        BalanceAmount.objects.create(
            loan_account=new_loan_account,
            balance_amount=total_amount
        )

        LoanHistory.objects.create(loan_account=new_loan_account, date=datetime.now(), action='CREATED')

        return redirect('loan_accounts')

    parties = party.objects.all()
    return render(request, 'add_loan.html', {'staff': staff, 'parties': parties})
 
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from .models import LoanAccounts, party
from datetime import datetime

def import_loan_accounts(request):
    staff_id = request.session.get('staff_id')
    staff = staff_details.objects.get(id=staff_id)

    if request.method == 'POST' and 'exceladd' in request.FILES:
        excel_file = request.FILES['exceladd']

        excel_data = load_workbook(excel_file, read_only=True)
        sheet = excel_data.active  

        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_date, lender_bank, account_name, account_number, description, loan_amount, loan_received, \
            cheque_number, upi_id, interest_rate_raw, duration, fee, lr, cheque_number_for_fee, upi_id_for_fee = row

            if interest_rate_raw is not None:
                try:
                    interest_rate = float(interest_rate_raw)
                except ValueError:
                    interest_rate = float(str(interest_rate_raw).rstrip('%'))
            else:
                interest_rate = None

            formatted_date = (
                raw_date.strftime('%Y-%m-%d') if raw_date else None
            )
            
            new_loan_account=LoanAccounts.objects.create(
  
                cheque_number_for_fee=cheque_number_for_fee,
                lender_bank=lender_bank,
                account_number=account_number,
                loan_amount=loan_amount,
                date=formatted_date,
                loan_received=loan_received,
                interest_rate=interest_rate,
                duration=duration,
                description=description,
                proccessing_fee=fee,
                lr=lr,
                upi_id=upi_id,
                upi_id_for_fee=upi_id_for_fee,
                account_name=account_name,
                cheque_number=cheque_number,
                company_id=staff.company.id
            )
            total_amount = fee + interest_rate + loan_amount
            BalanceAmount.objects.create(
            loan_account=new_loan_account,
            balance_amount=total_amount
        )

        return redirect('loan_accounts')


def edit_loan_page_function(request, eid):
    loan_account = LoanAccounts.objects.get(id=eid)

    if request.method == 'POST':
        raw_date = request.POST.get('date')
        formatted_date = (
            datetime.strptime(raw_date, '%d-%m-%Y').strftime('%Y-%m-%d')
            if raw_date
            else None
        )

        loan_account.account_number = request.POST.get('account_number')
        loan_account.lender_bank = request.POST.get('lender_bank')

        try:
            loan_account.loan_amount = Decimal(request.POST.get('loan_amount', '0'))
            loan_account.interest_rate = Decimal(request.POST.get('rate', '0'))
            loan_account.proccessing_fee = Decimal(request.POST.get('fee', '0'))
        except InvalidOperation as e:
            print(f"Error converting to Decimal: {e}")
            # Handle the error, e.g., set a default value or return an error response

        loan_account.duration = request.POST.get('duration')
        loan_account.description = request.POST.get('description')
        loan_account.loan_received = request.POST.get('loan_received')
        loan_account.lr = request.POST.get('lr')

        loan_account.cheque_number = request.POST.get('cheque_number')
        loan_account.upi_id = request.POST.get('upi_id')

        loan_account.cheque_number_for_fee = request.POST.get('cheque_number_for_fee')
        loan_account.upi_id_for_fee = request.POST.get('upi_id_for_fee')

        loan_account.save()

        existing_entry = LoanHistory.objects.filter(
            loan_account=loan_account,
            date__date=datetime.now().date(),
            action='EDITED'
        ).exists()

        if not existing_entry:
            LoanHistory.objects.create(
                loan_account=loan_account,
                date=datetime.now(),
                action='EDITED'
            )

        return redirect('loan_accounts')

    return render(request, 'edit_loan_page.html', {'data': loan_account})


def loan_accounts (request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
  cmp = company.objects.get(id=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)

  data = LoanAccounts.objects.filter(company=cmp)
  data2 = makepayment.objects.filter(loan_account__company=cmp)
  data3 = AdditionalLOan.objects.filter(loan_account__company=cmp)

  return render(request, 'company/loan_accounts.html',{'data':data, 'allmodules':allmodules,'staff':staff,'data2':data2,'data3':data3,'bank':bank})

def  add_loan_accounts(request):
  data = LoanAccounts.objects.all()
  parties = party.objects.all()

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  cmp = company.objects.get(id=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
  
  data = LoanAccounts.objects.all()
  parties = party.objects.all()
  return render(request, 'company/add_loan_accounts.html',{'data':data,'parties':parties,'allmodules':allmodules,'staff':staff,'bank':bank})

def  edit_loan_page(request,eid):
  data = LoanAccounts.objects.get(id=eid)
  parties = party.objects.all()

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
  cmp = company.objects.get(id=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)

  return render(request, 'company/edit_loan_page.html',{'data':data,'parties':parties,'allmodules':allmodules,'staff':staff,'bank':bank})

def  loan_accounts_view_page(request,eid):
  data = LoanAccounts.objects.get(id=eid)
  data2= BalanceAmount.objects.get(loan_account=eid) 
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  return render(request, 'company/loan_accounts_view_page.html',{'data':data,'allmodules':allmodules,'staff':staff ,'data2':data2})


from django.db.models import Sum

def make_payment(request, eid):
    loan_account_instance = get_object_or_404(LoanAccounts, id=eid)

    if request.method == 'POST':
        principal_amount = float(request.POST.get('principal_amount', 0))
        
        # Validate interest_amount input
        raw_interest_amount = request.POST.get('interest_amount', '0')
        if raw_interest_amount:
            interest_amount = float(raw_interest_amount)
        else:
            interest_amount = 0.0

        total_amount = principal_amount + interest_amount
        loan_received = request.POST.get('loan_received')

        cheque_number = request.POST.get('cheque_number')
        upi_id = request.POST.get('upi_id')
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        raw_date = request.POST.get('date')
        formatted_date = (
            datetime.strptime(raw_date, '%d/%m/%Y').strftime('%Y-%m-%d')
            if raw_date
            else None
        )

        with transaction.atomic():
            payment = makepayment(
                loan_account=loan_account_instance,
                principal_amount=principal_amount,
                interest_amount=interest_amount,
                total_amount=total_amount,
                loan_received=loan_received,
                date=formatted_date,
                cheque_number=cheque_number,
                upi_id=upi_id
            )
            payment.save()

            update_balance_amount(loan_account_instance)

        return redirect('loan_accounts')

    data = {}
    return render(request, 'loan_accounts.html', {'data': data}) 



from datetime import datetime

def additional_loan(request):
    if request.method == 'POST':
        loan_account_id = request.POST.get('loan_account')
        raw_date = request.POST.get('date')


        formatted_date = (
            datetime.strptime(raw_date, '%d/%m/%Y').strftime('%Y-%m-%d')
            if raw_date
            else None
        )

        additional_loan = request.POST.get('additional_loan')
        interest_amount = float(request.POST.get('interest_amount'))
        total_loan = float(additional_loan) + interest_amount

        try:
            loan_account = LoanAccounts.objects.get(id=loan_account_id)
        except LoanAccounts.DoesNotExist:
            loan_account = None

        makepayment_instance = makepayment.objects.create(
            loan_account=loan_account,
            date=formatted_date,
            interest_amount=interest_amount,
            total_amount=total_loan,
            loan_received=request.POST.get('loan_received')
        )

        AdditionalLOan.objects.create(
            loan_account=loan_account,
            makepayment=makepayment_instance,
            additional_loan=additional_loan,
            date=formatted_date,
            interest_amount=interest_amount,
            total_loan=total_loan
        )

    return redirect('loan_accounts')


def ShareLoanStatementMail(request,eid):
    if request.method == "POST":
        sid = request.session.get('staff_id')
        staff =  staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)
        data= LoanAccounts.objects.get(id=eid)
        context = {'staff' : staff,'data':data}
        my_subject = "LOAN ACCOUNT STATEMENT"
        emails_string = request.POST['email_ids']
        emails_list = [email.strip() for email in emails_string.split(',')]
 
        html_message = render_to_string('company/LoanStatement_pdf.html',context) 
     
        plain_message = strip_tags(html_message)
        pdf_content = BytesIO()
        pisa_document = pisa.CreatePDF(html_message.encode("UTF-8"), pdf_content) 
        pdf_content.seek(0)
   
        filename = f'LoanStatement {staff.company.company_name}.pdf'
        message = EmailMultiAlternatives(
            subject=my_subject,
            body= f"Hi,\nPlease find the attached Loan statement -  \n\n--\nRegards,\n{staff.company.company_name}\n{staff.company.address}\n{staff.company.state} - {staff.company.country}\n{staff.company.contact}",
            from_email='gokulkrishnagokul6@gmail.com',
            to=emails_list,  
            )
        message.attach(filename, pdf_content.read(), 'application/pdf')
        
        try:
            message.send()
            return HttpResponse('<script>alert("Report has been shared via successfully..!");window.location="/loan_accounts"</script>')
        except Exception as e:
         
            return HttpResponse('<script>alert("Failed to send email!");window.location="/loan_accounts"</script>')

    return HttpResponse('<script>alert("Invalid Request!");window.location="/loan_accounts"</script>')


def loan_account_history(request, id):

    loan_account_instance = get_object_or_404(LoanAccounts, id=id)

  
    loan_history_entries = LoanHistory.objects.filter(loan_account=loan_account_instance)

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    return render(request, 'company/loan_account_history.html', {'loan_account': loan_account_instance, 'loan_history_entries': loan_history_entries,'allmodules':allmodules,'staff':staff})


def LoanAccountDelete(request,id):
  data=LoanAccounts.objects.get(id=id)
  data.delete()
  return redirect('loan_accounts')
  

def create_sale(request):
    toda = date.today()
    tod = toda.strftime("%Y-%m-%d")
   
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    Party = party.objects.filter(company=cmp,user=cmp.user)
    bank = BankModel.objects.filter(company=cmp,user=cmp.user)
    allmodules= modules_list.objects.get(company=staff.company,status='New')
    
    last_credit = CreditNote.objects.filter(company=cmp).count()

    if last_credit:
      credit_note = last_credit + 1 
    else:
      credit_note = 1

    item = ItemModel.objects.filter(company=cmp,user=cmp.user)
    item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)
            
       

      
    context = {'staff':staff, 'allmodules':allmodules, 'party':Party, 'cmp':cmp,'credit_note':credit_note,'tod':tod,'item':item, 'item_units':item_units,'bank':bank}
    return render(request, 'company/create_sale.html', context)

def add_creditnote(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)    
    party_id = request.POST.get('partyname')
    print(party_id)
    part = party.objects.get(id=party_id)
    return_no=request.POST.get('creditno')
    partmob=request.POST.get('partyPhoneNumber')
    creditdate=request.POST.get('cr_date')
    invoice_date=request.POST.get('inv_date')
    supplyplace =request.POST.get('destination')
    pay_method=request.POST.get("payment_method")
    cheque_no=request.POST.get("cheque_id")
    upi_no=request.POST.get("upi_id")
    bank_acc=request.POST.get("bnk_id")
    advance = request.POST.get("advance")
    balance = request.POST.get("balance")
    subtotal=float(request.POST.get('subtotal'))
    igst = request.POST.get('igst')
    cgst = request.POST.get('cgst')
    sgst = request.POST.get('sgst')
    adjust = request.POST.get("adj")
    taxamount = request.POST.get("taxamount")
    grandtotal=request.POST.get('grandtotal')
    descptn=request.POST.get('description')
    creditnote=CreditNote(party=part,retrn_no = return_no,partymob=partmob,date=creditdate,invoice_date=invoice_date,
                          supplyplace=supplyplace,pay_method=pay_method,cheque_no=cheque_no,upi_no=upi_no,
                          bankaccount=bank_acc, subtotal=subtotal,advance=advance, balance=balance, igst=igst,
                          cgst=cgst,sgst=sgst,adjust=adjust,taxamount=taxamount,grandtotal=grandtotal, description=descptn,
                          company=cmp,staff=staff,)
    
    creditnote.save()
    
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    rate=tuple(request.POST.getlist("price[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    if request.POST.get('destination') =='State':
      tax =  tuple(request.POST.getlist("tax1[]"))
    else:
      tax =  tuple(request.POST.getlist("tax2[]"))
    total =  tuple(request.POST.getlist("total[]"))
    return_no = CreditNote.objects.filter(retrn_no=creditnote.retrn_no, company=cmp).first()


    if len(product)==len(qty)==len(tax)==len(discount)==len(total):
      mapped=zip(product,qty,tax,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        CreditNoteItem.objects.create(product = itm,qty=ele[1], tax=ele[2],discount=ele[3],total=ele[4],creditnote=return_no,company=cmp)

    
    CreditNote.objects.filter(company=cmp).update(tot_credit_no=F('tot_credit_no') + 1)
    
    creditnote.tot_credit_no = creditnote.retrn_no
    creditnote.save()
    CreditNoteTransactionHistory.objects.create(creditnote=creditnote,company=cmp,staff=staff,action='Created')
    if 'save_and_new' in request.POST:
      return redirect('create_sale')
    if 'save' in request.POST:
      return redirect('creditnote_list')
  return render(request, 'company/create_sale.html')


def new_creditnote_item(request):
  print('items')
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' or None :
      item_opening_stock = 0
      item_current_stock = 0
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '' or None:
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == ''  or None:
      item_min_stock_maintain = 0
    item_data = ItemModel(user=user,
                          company=company_user_data,
                          item_name=item_name,
                          item_hsn=item_hsn,
                          item_unit=item_unit,
                          item_taxable=item_taxable,
                          item_gst=item_gst,
                          item_igst=item_igst,
                          item_sale_price=item_sale_price,
                          item_purchase_price=item_purchase_price,
                          item_opening_stock=item_opening_stock,
                          item_current_stock=item_current_stock,
                          item_at_price=item_at_price,
                          item_date=item_date,
                          item_min_stock_maintain=item_min_stock_maintain)
    item_data.save()
    return JsonResponse({
            'status': 'success',
            'message': 'Item added successfully',
            'item_id': item_data.id,
            'item_name': item_data.item_name,
            'item_hsn': item_data.item_hsn,
        })
  else:
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'})
  

def get_hsn_for_item(request):
 
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})


def get_party_number(request):
    selected_party_id = request.GET.get('partyname')
    party_instance = get_object_or_404(party, id=selected_party_id)
    phone_number = party_instance.contact
    party_id = party_instance.id
    balnce =party_instance.openingbalance
    
    return JsonResponse({'phone': phone_number, 'id': party_id,'balance':balnce})


def creditnote_list(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  Part = party.objects.filter(company=cmp,user=cmp.user)
  credit = CreditNote.objects.filter(company=cmp)

  if not credit:
    context = {'staff':staff, 'allmodules':allmodules,'party':Part}
    return render(request,'company/salesfirst.html',context)
  
  context = {'staff':staff,'allmodules':allmodules,'credit':credit,'party':Part}
  return render(request,'company/creditlist.html',context)


def party_dropdown(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = party.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  party_list = []
  for p in part:
    id_list.append(p.id)
    party_list.append(p.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })

def saveparty(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  party_name = request.POST['name']
  email = request.POST['email']
  contact = request.POST['mobile']
  state = request.POST['splystate']
  address = request.POST['baddress']
  gst_type = request.POST['gsttype']
  gst_no = request.POST['gstin']
  current_date = request.POST['partydate']
  openingbalance = request.POST.get('openbalance')
  payment = request.POST.get('paytype')
  creditlimit = request.POST.get('credit_limit')
  End_date = request.POST.get('enddate', None)
  additionalfield1 = request.POST['add1']
  additionalfield2 = request.POST['add2']
  additionalfield3 = request.POST['add3']

  part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,
                payment=payment,creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,
                additionalfield3=additionalfield3,company=cmp,user=cmp.user)
  part.save() 
  return JsonResponse({'success': True})



def credit_bankdetails(request):
  bid = request.POST['id']
  bank = BankModel.objects.filter(id=bid) 
  bank_no = bank.account_num 
  bank_name = bank.bank_name
  return JsonResponse({'bank_no':bank_no,'bank_name':bank_name})



def detail_creditnote(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.get(company=staff.company,status='New')
  credit = CreditNote.objects.get(id=id,company=cmp)
  citm = CreditNoteItem.objects.filter(creditnote=credit,company=cmp)
  dis = 0
  for itm in citm:
    dis += int(itm.discount)
  itm_len = len(citm)

  context={'staff':staff,'allmodules':allmodules,'credit':credit,'citm':citm,'itm_len':itm_len,'dis':dis}
  return render(request,'company/creditnotedetails.html',context)

def import_creditnote(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(CreditNote.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      CreditNote.objects.create(party=part,retrn_no=totval,
                                  date=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      credit = CreditNote.objects.last()
      if billsheet[4] == 'Cheque':
        credit.pay_method = 'Cheque'
        credit.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        credit.pay_method = 'UPI'
        credit.upi_no = billsheet[5]
      else:
        if billsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=billsheet[4],company=cmp)
          credit.pay_method = bank
        else:
          credit.pay_method = 'Cash'
      credit.save()

      CreditNote.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],item_hsn=prdsheet[2])
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[5])
      CreditNoteItem.objects.create(creditnote=credit,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                tax=prdsheet[4],
                                discount=prdsheet[5],
                                total=total)

      temp = prdsheet[4].split('[')
      if billsheet[3] =='State':
        tax=int(temp[0][3:])
      else:
        tax=int(temp[0][4:])

        subtotal += total
        tamount = total *(tax / 100)
        taxamount += tamount
                
      if billsheet[3]=='State':
        gst = round((taxamount/2),2)
        credit.sgst=gst
        credit.cgst=gst
        credit.igst=0

      else:
        gst=round(taxamount,2)
        credit.igst=gst
        credit.cgst=0
        credit.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      credit.subtotal=round(subtotal,2)
      credit.taxamount=round(taxamount,2)
      credit.adjust=round(billsheet[6],2)
      credit.grandtotal=gtotal
      credit.advance=round(billsheet[7],2)
      credit.balance=balance
      credit.save()

      CreditNoteTransactionHistory.objects.create(creditnote=credit,staff=credit.staff,company=credit.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})


def delete_CreditNote(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  credit = CreditNote.objects.get(id=id)
  CreditNoteItem.objects.filter(creditnote=credit,company=cmp).delete()
  credit.delete()
  return redirect('creditnote_list')

def edit_creditnote(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part= party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(user=cmp.user,company=staff.company.id)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  bank = BankModel.objects.filter(company=cmp,user=cmp.user)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  credit = CreditNote.objects.get(id=id,company=cmp)
  credit_itm = CreditNoteItem.objects.filter(creditnote=credit,company=cmp)


  if credit.pay_method != 'Cash' and credit.pay_method != 'Cheque' and credit.pay_method != 'UPI':
    bankno = BankModel.objects.get(id= credit.pay_method,company=cmp,user=cmp.user)
  else:
    bankno = 0

  bdate = credit.date.strftime("%Y-%m-%d")
  context = {'staff':staff, 'allmodules':allmodules, 'credit':credit, 'credititm':credit_itm,'tod':tod,
             'party':part, 'product':item, 'item_units':item_units, 'bdate':bdate,'bank':bank, 'bankno':bankno}
  return render(request,'company/edit_creditnot.html',context)


def update_creditnote(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    part = party.objects.get(id=request.POST.get('partyname'))
    credit = CreditNote.objects.get(id=id,company=cmp)
    credit.party = part
    credit.date = request.POST.get('cr_date')
    credit.supplyplace  = request.POST.get('destination')
    credit.subtotal =float(request.POST.get('subtotal'))
    credit.grandtotal = request.POST.get('grandtotal')
    credit.igst = request.POST.get('igst')
    credit.cgst = request.POST.get('cgst')
    credit.sgst = request.POST.get('sgst')
    credit.taxamount = request.POST.get("taxamount")
    credit.adjust = request.POST.get("adj")
    credit.pay_method = request.POST.get("method")
    credit.cheque_no = request.POST.get("cheque_id")
    credit.upi_no = request.POST.get("upi_id")
    credit.advance = request.POST.get("advance")
    credit.balance = request.POST.get("balance")

    credit.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    if request.POST.get('placosupply') == 'State':
      tax =tuple( request.POST.getlist("tax1[]"))
    else:
      tax = tuple(request.POST.getlist("tax2[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    CreditNoteItem.objects.filter(creditnote=credit,company=cmp).delete()
    if len(total)==len(discount)==len(qty)==len(tax):
      mapped=zip(product,qty,tax,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        CreditNoteItem.objects.create(product =itm,qty=ele[1], tax=ele[2],discount=ele[3],total=ele[4],creditnote=credit,company=cmp)

    CreditNoteTransactionHistory.objects.create(creditnote=credit,company=cmp,staff=staff,action='Updated')
    return redirect('creditnote_list')

  return redirect('creditnote_list')


def get_inv_date(request):
    selected_bill_no = request.POST.get('bill_no', None)

    try:
        # Get the latest PurchaseBill with the specified bill_number
        sales_inv =SalesInvoice.objects.filter(invoice_no=selected_bill_no).latest('inv_date')
        bill_date = sales_inv.date.strftime('%Y-%m-%d')
    except SalesInvoice.DoesNotExist:
        return JsonResponse({'error': 'Bill number not found'}, status=400)
    except SalesInvoice.MultipleObjectsReturned:
        # Handle the case where multiple PurchaseBills are found for the same bill_number
        return JsonResponse({'error': 'Multiple PurchaseBills found for the same bill number'}, status=400)

    return JsonResponse({'bill_date': bill_date})
def salesinvoicedata(request):
    try:
        party_name = request.POST['id']
        party_instance = party.objects.get(id=party_name)

        # Initialize lists to store multiple bill numbers and dates
        bill_numbers = []
        bill_dates = []

        try:
            # Retrieve all PurchaseBill instances for the party
            bill_instances = SalesInvoice.objects.filter(party=party_instance)

            # Loop through each PurchaseBill instance and collect bill numbers and dates
            for bill_instance in bill_instances:
                bill_numbers.append(bill_instance.invoice_no)
                bill_dates.append(bill_instance.date)

        except SalesInvoice.DoesNotExist:
            pass

        # Return a JSON response with the list of bill numbers and dates
        if not bill_numbers and not bill_dates:
            return JsonResponse({'bill_numbers': ['nobill'], 'bill_dates': ['nodate']})

        return JsonResponse({'bill_numbers': bill_numbers, 'bill_dates': bill_dates})

    except KeyError:
        return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

    except party.DoesNotExist:
        return JsonResponse({'error': 'Party not found.'})
    
  

def  creditnote_item_unit(request):
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_unit_name = request.POST.get('item_unit_name')
    unit_data = UnitModel(user=user,company=company_user_data,unit_name=item_unit_name)
    unit_data.save()
  return JsonResponse({'message':'asdasd'})


def history_creditnote(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)   
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  credit = CreditNote.objects.get(id=id,company=cmp)
  hst= CreditNoteTransactionHistory.objects.filter(creditnote=credit,company=cmp)

  context = {'staff':staff,'allmodules':allmodules,'hst':hst,'credit':credit}
  return render(request,'company/creditnotehistory.html',context)


def  credititemdetails(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})
  
  
from decimal import Decimal, getcontext
getcontext().prec = 10   

def calculate_balance_amount(loan_account_instance):
    payments_sum = makepayment.objects.filter(loan_account=loan_account_instance).aggregate(Sum('total_amount'))
    total_payments = payments_sum['total_amount__sum'] or Decimal('0')

    additional_loans_sum = AdditionalLOan.objects.filter(loan_account=loan_account_instance).aggregate(Sum('total_loan'))
    total_additional_loans = additional_loans_sum['total_loan__sum'] or Decimal('0')

    loan_amount = Decimal(str(loan_account_instance.loan_amount))
    balance_amount = loan_amount - Decimal(str(total_payments)) + total_additional_loans
    return balance_amount
    
def update_balance_amount(loan_account_instance):
    balance_amount = calculate_balance_amount(loan_account_instance)

    print(f"Updated Balance Amount: {balance_amount}")

    balance_entry, created = BalanceAmount.objects.get_or_create(
        loan_account=loan_account_instance,
        makepayment=None,
        additional_loan=None,
        defaults={'balance_amount': balance_amount}
    )

    if not created:
        balance_entry.balance_amount = balance_amount
        balance_entry.save()
        
def additional_loan_function(request, eid):
    loan_account_instance = get_object_or_404(LoanAccounts, id=eid)

    if request.method == 'POST':

        loan_received= request.POST.get("loan_received")
        cheque_number= request.POST.get("cheque_number")
        upi_id= request.POST.get("upi_id")
        raw_date = request.POST.get('date')
        formatted_date = (
            datetime.strptime(raw_date, '%d/%m/%Y').strftime('%Y-%m-%d')
            if raw_date
            else None
        )
        additional_loan = Decimal(request.POST.get('additional_loan', '0')) if request.POST.get('additional_loan') else Decimal('0')
        interest_amount = Decimal(request.POST.get('interest_amount', '0')) if request.POST.get('interest_amount') else Decimal('0')
        total_loan = Decimal(additional_loan) + Decimal(interest_amount)


 

        data=AdditionalLOan.objects.create(
            loan_account=loan_account_instance,
      
            additional_loan=additional_loan,
            date=formatted_date,
            interest_amount=interest_amount,
            total_loan=total_loan,
            loan_received=loan_received,
            cheque_number=cheque_number,
            upi_id=upi_id
        )
        data.save()

   
        update_balance_amount(loan_account_instance)

        print("Data inserted successfully!")
        return redirect('loan_accounts')

 
    return render(request, 'your_template.html', {'loan_account_instance': loan_account_instance})
    
    
def email_saleorder(request,id):
  if request.method == 'POST':
    print("ggggggggggggggggggggg")
    emails_string = request.POST['email_ids']
                # Split the string by commas and remove any leading or trailing whitespace
    emails_list = [email.strip() for email in emails_string.split(',')]
    email_message = request.POST['email_message']
    print(emails_list)

    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id) 
    sale = salesorder.objects.get(id=id,staff=staff)
    saleitem= sales_item.objects.filter(sale_order=sale.id)
    context = {'sale':sale, 'cmp':cmp,'saleitem':saleitem}
    template_path = 'company/saleorder_file_mail.html'
    
    template = get_template(template_path)
    html  = template.render(context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    pdf = result.getvalue()
    filename = f'SALE ORDER - {sale.orderno}.pdf'
    subject = f"SALE ORDER - {sale.orderno}"
    email = EmailMessage(subject, f"Hi,\nPlease find the attached SALE ORDER - File-{sale.orderno}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
    email.attach(filename, pdf, "application/pdf")
    email.send(fail_silently=False)
    # msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
    return redirect(saleorder_view,id)


def dropp(request):
  return render(request, 'drop.html')







